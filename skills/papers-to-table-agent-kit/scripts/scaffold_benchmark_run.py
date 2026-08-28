from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from review_package_common import review_input_path, safe_filename  # noqa: E402

SCRATCH_DIRNAME = "scratch_delete_after_success"
SCRATCH_ROOT_MARKER = ".papers_to_table_scratch_root"
SCRATCH_RUN_MARKER = ".papers_to_table_scratch"
EXTRACTION_MODES = {"fill_blanks", "fill_and_verify"}
PROTECTED_BENCHMARK_TABLES = {"table_gold.csv"}


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_xlsx(path: Path, sheet_name: str | None = None) -> list[tuple[str, list[str], list[dict[str, str]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            f"Cannot inspect workbook {path}; install openpyxl or export the authoritative sheet to CSV."
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    names = [sheet_name] if sheet_name else list(workbook.sheetnames)
    if sheet_name and sheet_name not in workbook.sheetnames:
        raise ValueError(f"Workbook {path} has no sheet named {sheet_name!r}.")
    tables: list[tuple[str, list[str], list[dict[str, str]]]] = []
    for name in names:
        rows = list(workbook[name].iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        if not any(headers):
            continue
        records = [
            {header: "" if value is None else str(value) for header, value in zip(headers, values) if header}
            for values in rows[1:]
        ]
        tables.append((name, headers, records))
    return tables


def _tabular_views(path: Path, sheet_name: str | None = None) -> list[tuple[str | None, list[str], list[dict[str, str]]]]:
    if path.suffix.lower() == ".csv":
        headers, rows = read_csv(path)
        return [(None, headers, rows)]
    if path.suffix.lower() == ".xlsx":
        return [(name, headers, rows) for name, headers, rows in _read_xlsx(path, sheet_name)]
    raise ValueError(f"Authoritative tables must be CSV or XLSX: {path}")


def _identity(value: object) -> str:
    return " ".join(str(value or "").casefold().split())


def _match_rows(template_rows: list[dict[str, str]], candidate_rows: list[dict[str, str]]) -> tuple[dict[int, dict[str, str]], list[int]]:
    by_row_id: dict[str, list[dict[str, str]]] = {}
    by_title: dict[str, list[dict[str, str]]] = {}
    for row in candidate_rows:
        row_id = _identity(row.get("row_id"))
        title = _identity(row.get("Title") or row.get("title"))
        if row_id:
            by_row_id.setdefault(row_id, []).append(row)
        if title:
            by_title.setdefault(title, []).append(row)
    matched: dict[int, dict[str, str]] = {}
    ambiguous: list[int] = []
    for index, row in enumerate(template_rows):
        candidates = by_row_id.get(_identity(row.get("row_id")), [])
        if not candidates:
            candidates = by_title.get(_identity(row.get("Title") or row.get("title")), [])
        if len(candidates) == 1:
            matched[index] = candidates[0]
        elif len(candidates) > 1:
            ambiguous.append(index)
    return matched, ambiguous


def _candidate_assessment(
    path: Path,
    sheet: str | None,
    headers: list[str],
    rows: list[dict[str, str]],
    template_rows: list[dict[str, str]],
    target_columns: list[str],
) -> tuple[dict[str, object], dict[int, dict[str, str]]]:
    matched, ambiguous = _match_rows(template_rows, rows)
    shared_targets = [name for name in target_columns if name in headers]
    missing_values = 0
    conflicts = 0
    populated = 0
    for index, source_row in matched.items():
        template_row = template_rows[index]
        for name in shared_targets:
            source_value = str(source_row.get(name) or "")
            template_value = str(template_row.get(name) or "")
            if source_value.strip():
                populated += 1
                if not template_value.strip():
                    missing_values += 1
                elif source_value != template_value:
                    conflicts += 1
    assessment = {
        "path": str(path.resolve()),
        "sheet": sheet,
        "matched_rows": len(matched),
        "ambiguous_rows": len(ambiguous),
        "shared_target_columns": shared_targets,
        "populated_target_cells": populated,
        "missing_from_template_cells": missing_values,
        "conflicting_target_cells": conflicts,
        "requires_authority_decision": bool(missing_values or conflicts),
    }
    return assessment, matched


def _discover_companion_tables(
    dataset_dir: Path,
    table_path: Path,
    schema_path: Path | None,
    run_dir: Path,
    output_root: Path | None,
    template_rows: list[dict[str, str]],
    target_columns: list[str],
) -> list[dict[str, object]]:
    excluded_roots = [path.resolve() for path in (run_dir, output_root) if path is not None]
    assessments: list[dict[str, object]] = []
    for path in sorted(dataset_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in {".csv", ".xlsx"}:
            continue
        resolved = path.resolve()
        if resolved == table_path.resolve() or (schema_path is not None and resolved == schema_path.resolve()):
            continue
        if path.name.casefold() in PROTECTED_BENCHMARK_TABLES:
            assessments.append({"path": str(resolved), "status": "protected_benchmark_gold_ignored"})
            continue
        if any(root == resolved or root in resolved.parents for root in excluded_roots):
            continue
        try:
            views = _tabular_views(path)
        except Exception as exc:
            assessments.append({"path": str(resolved), "status": "inspection_error", "error": str(exc)})
            continue
        for sheet, headers, rows in views:
            if not any(name in headers for name in target_columns):
                continue
            assessment, _matched = _candidate_assessment(
                path, sheet, headers, rows, template_rows, target_columns
            )
            if assessment["matched_rows"]:
                assessment["status"] = "compatible"
                assessments.append(assessment)
    return assessments


def _write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _parse_allowed_values(value: object) -> list[str] | None:
    if value is None:
        return None
    if isinstance(value, list):
        parsed = [str(item).strip() for item in value if str(item).strip()]
        return parsed or None
    raw = str(value).strip()
    if not raw:
        return None
    if raw.startswith("["):
        try:
            decoded = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise ValueError(f"Invalid JSON-array allowed_values: {raw}") from exc
        if not isinstance(decoded, list):
            raise ValueError(f"JSON allowed_values must be an array: {raw}")
        parsed = [str(item).strip() for item in decoded if str(item).strip()]
        return parsed or None
    parsed = [item.strip() for item in raw.split("|") if item.strip()]
    return parsed or None


def _normalize_schema_columns(columns: list[dict[str, object]]) -> list[dict[str, object]]:
    normalized: list[dict[str, object]] = []
    for item in columns:
        column = dict(item)
        allowed_values = _parse_allowed_values(column.get("allowed_values"))
        if allowed_values is not None:
            column["allowed_values"] = allowed_values
        elif "allowed_values" in column:
            column.pop("allowed_values")
        normalized.append(column)
    return normalized


def read_schema_columns(schema_path: Path | None, table_headers: list[str]) -> list[dict[str, object]]:
    if schema_path is None:
        return [
            {"column_name": header, "field_type": "text"}
            for header in table_headers
            if header not in {"row_id", "row_index", "pdf_id"}
        ]

    if schema_path.suffix.lower() == ".json":
        data = json.loads(schema_path.read_text(encoding="utf-8-sig"))
        if isinstance(data, dict) and isinstance(data.get("columns"), list):
            return _normalize_schema_columns(data["columns"])
        if isinstance(data, list):
            return _normalize_schema_columns(data)
        raise ValueError(f"Cannot infer columns from JSON schema: {schema_path}")

    _, rows = read_csv(schema_path)
    columns: list[dict[str, object]] = []
    for row in rows:
        name = row.get("column_name") or row.get("name") or row.get("column")
        if not name:
            continue
        column: dict[str, object] = {"column_name": name}
        if row.get("description"):
            column["description"] = row["description"]
        if row.get("field_type"):
            column["field_type"] = row["field_type"]
        allowed_values = _parse_allowed_values(row.get("allowed_values"))
        if allowed_values is not None:
            column["allowed_values"] = allowed_values
        columns.append(column)
    return columns


def choose_existing(dataset_dir: Path, names: list[str]) -> Path | None:
    for name in names:
        candidate = dataset_dir / name
        if candidate.exists():
            return candidate
    return None


def _explicit_pdf_value(row: dict[str, str]) -> str:
    return str(row.get("pdf_id") or row.get("PDF") or row.get("pdf") or row.get("pdf_file") or "").strip()


def assign_row_pdf_ids(
    table_rows: list[dict[str, str]],
    pdf_entries: list[dict[str, str]],
    *,
    allow_positional_pdf_fallback: bool,
) -> tuple[list[str | None], dict[str, object]]:
    pdf_by_alias: dict[str, str] = {}
    for pdf in pdf_entries:
        canonical = str(pdf["pdf_id"])
        aliases = {canonical, str(pdf["label"]), Path(str(pdf["path"])).name, Path(str(pdf["path"])).stem}
        for alias in aliases:
            pdf_by_alias[alias.casefold()] = canonical

    explicit_values = [_explicit_pdf_value(row) for row in table_rows]
    if allow_positional_pdf_fallback:
        if any(explicit_values):
            raise ValueError("--allow-positional-pdf-fallback cannot be combined with explicit table PDF mappings.")
        if len(table_rows) != len(pdf_entries):
            raise ValueError(
                "Positional PDF fallback requires exactly one table row per PDF; "
                f"found {len(table_rows)} rows and {len(pdf_entries)} PDFs."
            )
        row_pdf_ids: list[str | None] = [str(pdf["pdf_id"]) for pdf in pdf_entries]
        mapping_mode = "positional_explicit_opt_in"
    else:
        row_pdf_ids = []
        for index, (row, explicit) in enumerate(zip(table_rows, explicit_values)):
            if not explicit:
                row_pdf_ids.append(None)
                continue
            canonical = pdf_by_alias.get(explicit.casefold()) or pdf_by_alias.get(Path(explicit).stem.casefold())
            if canonical is None:
                row_id = row.get("row_id") or f"row_{index + 1}"
                raise ValueError(f"Row {row_id!r} references a PDF that is not present in pdfs/: {explicit!r}")
            row_pdf_ids.append(canonical)
        mapping_mode = "explicit"

    assigned = [pdf_id for pdf_id in row_pdf_ids if pdf_id]
    duplicates = sorted(pdf_id for pdf_id in set(assigned) if assigned.count(pdf_id) > 1)
    if duplicates:
        raise ValueError(f"Each PDF may map to at most one row; duplicate assignments: {duplicates}")
    unused_pdf_ids = sorted({str(pdf["pdf_id"]) for pdf in pdf_entries} - set(assigned))
    if unused_pdf_ids:
        preview = ", ".join(unused_pdf_ids[:8])
        suffix = "..." if len(unused_pdf_ids) > 8 else ""
        raise ValueError(
            f"{len(unused_pdf_ids)} PDF(s) are not explicitly mapped to table rows: {preview}{suffix}. "
            "Match PDFs from DOI/title/authors/year and add a pdf_id column before scaffolding. "
            "Use --allow-positional-pdf-fallback only when one-to-one row order has been independently verified."
        )

    return row_pdf_ids, {
        "mapping_mode": mapping_mode,
        "mapped_rows": len(assigned),
        "unmapped_rows": len(table_rows) - len(assigned),
        "unused_pdfs": len(unused_pdf_ids),
    }


def scaffold(
    dataset_dir: Path,
    run_dir: Path,
    *,
    force: bool = False,
    output_root: Path | None = None,
    extraction_mode: str = "fill_blanks",
    allow_positional_pdf_fallback: bool = False,
    authoritative_table: Path | None = None,
    authoritative_sheet: str | None = None,
    allow_template_only: bool = False,
) -> dict[str, object]:
    dataset_dir = dataset_dir.resolve()
    run_dir = run_dir.resolve()
    output_root = output_root.resolve() if output_root is not None else None
    if extraction_mode not in EXTRACTION_MODES:
        raise ValueError(f"Unsupported extraction mode: {extraction_mode}")
    if not dataset_dir.exists():
        raise FileNotFoundError(f"Dataset directory does not exist: {dataset_dir}")
    if run_dir.exists() and any(run_dir.iterdir()) and not force:
        raise FileExistsError(f"Run directory is not empty: {run_dir}. Re-run with --force to add/overwrite scaffold files.")

    pdf_source_dir = dataset_dir / "pdfs"
    if not pdf_source_dir.exists():
        raise FileNotFoundError(f"Expected PDFs in: {pdf_source_dir}")
    pdf_files = sorted(path for path in pdf_source_dir.iterdir() if path.suffix.lower() == ".pdf")
    if not pdf_files:
        raise FileNotFoundError(f"No PDF files found in: {pdf_source_dir}")

    table_path = choose_existing(dataset_dir, ["table_template.csv", "source_table.csv"])
    if table_path is None:
        raise FileNotFoundError(f"Expected table_template.csv or source_table.csv in: {dataset_dir}")

    schema_path = choose_existing(dataset_dir, ["schema.json", "schema.csv"])
    table_headers, table_rows = read_csv(table_path)

    pdf_entries: list[dict[str, str]] = []
    for pdf_path in pdf_files:
        pdf_entries.append({"pdf_id": pdf_path.stem, "path": str(pdf_path.resolve()), "label": pdf_path.stem})

    columns = read_schema_columns(schema_path, table_headers)
    target_columns = [str(column.get("column_name") or "") for column in columns if column.get("column_name")]
    companion_tables = _discover_companion_tables(
        dataset_dir,
        table_path,
        schema_path,
        run_dir,
        output_root,
        table_rows,
        target_columns,
    )
    risky_companions = [item for item in companion_tables if item.get("requires_authority_decision")]
    if risky_companions and authoritative_table is None and not allow_template_only:
        examples = ", ".join(
            f"{item['path']}" + (f" [{item['sheet']}]" if item.get("sheet") else "")
            for item in risky_companions[:4]
        )
        raise ValueError(
            "Potential pre-existing human-reviewed target values were found in compatible companion table(s): "
            f"{examples}. Re-run with --authoritative-table PATH (and --authoritative-sheet for XLSX), "
            "or use --allow-template-only only after confirming those values must be disregarded."
        )

    original_table_path = table_path
    effective_rows = [dict(row) for row in table_rows]
    authoritative_source: Path | None = None
    authoritative_source_sheet: str | None = None
    restored_cells = 0
    overridden_cells = 0
    authoritative_assessment: dict[str, object] | None = None
    if authoritative_table is not None:
        authoritative_source = authoritative_table.resolve()
        if not authoritative_source.exists():
            raise FileNotFoundError(f"Authoritative table does not exist: {authoritative_source}")
        compatible: list[tuple[str | None, list[str], list[dict[str, str]], dict[str, object], dict[int, dict[str, str]]]] = []
        for sheet, headers, rows in _tabular_views(authoritative_source, authoritative_sheet):
            assessment, matched = _candidate_assessment(
                authoritative_source, sheet, headers, rows, table_rows, target_columns
            )
            if assessment["matched_rows"] and assessment["shared_target_columns"]:
                compatible.append((sheet, headers, rows, assessment, matched))
        if len(compatible) != 1:
            raise ValueError(
                f"Authoritative table must resolve to exactly one compatible sheet; found {len(compatible)}. "
                "Pass --authoritative-sheet when an XLSX workbook is ambiguous."
            )
        authoritative_source_sheet, _headers, _rows, authoritative_assessment, matched = compatible[0]
        if authoritative_assessment["ambiguous_rows"]:
            raise ValueError(
                f"Authoritative table has {authoritative_assessment['ambiguous_rows']} ambiguous row matches; "
                "add unique row_id values or disambiguate the source table."
            )
        for index, source_row in matched.items():
            for name in target_columns:
                source_value = str(source_row.get(name) or "")
                if not source_value.strip():
                    continue
                current = str(effective_rows[index].get(name) or "")
                if not current.strip():
                    restored_cells += 1
                elif current != source_value:
                    overridden_cells += 1
                effective_rows[index][name] = source_value
        table_rows = effective_rows
        for name in target_columns:
            if name not in table_headers:
                table_headers.append(name)

    row_pdf_ids, mapping_summary = assign_row_pdf_ids(
        table_rows,
        pdf_entries,
        allow_positional_pdf_fallback=allow_positional_pdf_fallback,
    )
    row_ids = [row.get("row_id") or f"row_{index + 1}" for index, row in enumerate(table_rows)]
    duplicate_row_ids = sorted(row_id for row_id in set(row_ids) if row_ids.count(row_id) > 1)
    if duplicate_row_ids:
        raise ValueError(f"Duplicate row_id values are not allowed: {duplicate_row_ids}")
    mapped_table_rows = [row for row, pdf_id in zip(table_rows, row_pdf_ids) if pdf_id]
    populated_target_cells = sum(
        1
        for row in mapped_table_rows
        for column_name in target_columns
        if str(row.get(column_name) or "").strip()
    )
    total_target_cells = len(mapped_table_rows) * len(target_columns)
    blank_target_cells = total_target_cells - populated_target_cells
    rows_with_populated_targets = sum(
        1 for row in mapped_table_rows if any(str(row.get(column_name) or "").strip() for column_name in target_columns)
    )

    # Do not create a partial run workspace until table/PDF matching and the
    # source contracts have passed their fail-closed checks.
    run_dir.mkdir(parents=True, exist_ok=True)
    effective_table_path = original_table_path.resolve()
    if authoritative_source is not None:
        effective_table_path = run_dir / "extraction" / "authoritative_baseline.csv"
        _write_csv(effective_table_path, table_rows, table_headers)
    manifest_path = run_dir / "extraction" / "baseline_manifest.json"
    manifest = {
        "schema_version": "papers_to_table.baseline_manifest.v1",
        "original_template_path": str(original_table_path.resolve()),
        "original_template_sha256": _sha256(original_table_path),
        "authoritative_source_table_path": str(authoritative_source) if authoritative_source else None,
        "authoritative_source_sheet": authoritative_source_sheet,
        "authoritative_source_sha256": _sha256(authoritative_source) if authoritative_source else None,
        "effective_source_table_path": str(effective_table_path.resolve()),
        "target_columns": target_columns,
        "companion_tables": companion_tables,
        "template_only_override": bool(allow_template_only and risky_companions and authoritative_source is None),
        "authoritative_assessment": authoritative_assessment,
        "restored_cells": restored_cells,
        "overridden_cells": overridden_cells,
        "preexisting_human_reviewed_cells": sum(
            1 for row in table_rows for name in target_columns if str(row.get(name) or "").strip()
        ),
    }
    manifest["effective_source_table_sha256"] = _sha256(effective_table_path)
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    if output_root is not None:
        (output_root / "runs").mkdir(parents=True, exist_ok=True)
        scratch_root = output_root / SCRATCH_DIRNAME
        scratch_run = scratch_root / run_dir.name
        scratch_run.mkdir(parents=True, exist_ok=True)
        (scratch_root / SCRATCH_ROOT_MARKER).write_text("papers-to-table scratch root\n", encoding="utf-8")
        (scratch_run / SCRATCH_RUN_MARKER).write_text("delete-after-success scratch\n", encoding="utf-8")
        (output_root / "logs").mkdir(parents=True, exist_ok=True)
    rows: list[dict[str, object]] = []
    for index, (row, row_pdf_id) in enumerate(zip(table_rows, row_pdf_ids)):
        row_id = row_ids[index]
        values = {key: value for key, value in row.items() if value not in {"", None}}
        label = row.get("Title") or row.get("title") or row_id
        rows.append(
            {
                "row_id": row_id,
                "pdf_id": row_pdf_id,
                "label": label,
                "values": values,
            }
        )

    output_table_name = f"{safe_filename(dataset_dir.name, 'filled_table')}_filled.csv"
    review_input = {
        "schema_version": "papers_to_table.review_input.v1",
        "extraction_mode": extraction_mode,
        "run_id": run_dir.name,
        "output_table_name": output_table_name,
        "source_table_path": str(effective_table_path.resolve()),
        "original_template_path": str(original_table_path.resolve()),
        "authoritative_source_table_path": str(authoritative_source) if authoritative_source else None,
        "authoritative_source_sheet": authoritative_source_sheet,
        "baseline_manifest_path": str(manifest_path.resolve()),
        "schema_path": str(schema_path.resolve()) if schema_path is not None else None,
        "pdfs": pdf_entries,
        "columns": columns,
        "rows": rows,
        "proposals": [],
    }
    if output_root is not None:
        review_input["output_table_path"] = str(output_root / output_table_name)
    filled_table = Path(str(review_input["output_table_path"])) if "output_table_path" in review_input else run_dir / output_table_name
    path = review_input_path(run_dir)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(review_input, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    return {
        "run_dir": str(run_dir),
        "review_input": str(path),
        "filled_table": str(filled_table),
        "output_root": str(output_root) if output_root is not None else None,
        "pdfs": len(pdf_entries),
        "rows": len(rows),
        "columns": len(columns),
        "extraction_mode": extraction_mode,
        **mapping_summary,
        "total_target_cells": total_target_cells,
        "source_table_target_cells": len(table_rows) * len(target_columns),
        "table_only_target_cells": (len(table_rows) - len(mapped_table_rows)) * len(target_columns),
        "populated_target_cells": populated_target_cells,
        "blank_target_cells": blank_target_cells,
        "eligible_target_cells": blank_target_cells if extraction_mode == "fill_blanks" else total_target_cells,
        "rows_with_populated_targets": rows_with_populated_targets,
        "source_table": str(table_path.resolve()),
        "baseline_manifest": str(manifest_path.resolve()),
        "baseline_status": (
            "authoritative_baseline_applied"
            if authoritative_source is not None
            else "template_only_explicit_override"
            if allow_template_only and risky_companions
            else "template_verified_no_missing_companion_values"
        ),
        "companion_candidate_count": len(risky_companions),
        "authoritative_restored_cells": restored_cells,
        "authoritative_overridden_cells": overridden_cells,
        "preexisting_human_reviewed_cells": manifest["preexisting_human_reviewed_cells"],
        "schema": str(schema_path.resolve()) if schema_path is not None else None,
        "status": "scaffolded_incomplete_until_proposals_are_added",
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Scaffold a papers-to-table review run from a benchmark dataset folder.")
    parser.add_argument("--dataset-dir", required=True, type=Path, help="Benchmark dataset folder containing pdfs/, table_template.csv, and schema.*")
    parser.add_argument("--run", type=Path, help="Run directory to create or update with scaffold files. Defaults to OUTPUT_ROOT/runs/<dataset> when --output-root is set.")
    parser.add_argument("--output-root", type=Path, help="Output workspace root for final *_filled.csv files and organized runs/scratch/logs subfolders.")
    parser.add_argument(
        "--extraction-mode",
        choices=sorted(EXTRACTION_MODES),
        default="fill_blanks",
        help="Fill only blank target cells, or explicitly verify populated cells as review proposals.",
    )
    parser.add_argument(
        "--authoritative-table",
        type=Path,
        help="CSV/XLSX containing pre-existing approved target values to overlay onto the extraction template.",
    )
    parser.add_argument(
        "--authoritative-sheet",
        help="Worksheet name when --authoritative-table is an XLSX workbook with multiple compatible sheets.",
    )
    parser.add_argument(
        "--allow-template-only",
        action="store_true",
        help="Explicitly disregard compatible companion-table values after independently confirming the template is authoritative.",
    )
    parser.add_argument(
        "--allow-positional-pdf-fallback",
        action="store_true",
        help="Opt into row-order PDF assignment only when row/PDF counts and one-to-one ordering were independently verified.",
    )
    parser.add_argument("--force", action="store_true", help="Allow writing scaffold files into a non-empty run directory")
    parser.add_argument("--json", action="store_true", help="Print machine-readable JSON")
    args = parser.parse_args(argv)

    if args.run is None and args.output_root is None:
        parser.error("Pass --run or --output-root.")
    run_dir = args.run
    if run_dir is None:
        run_dir = args.output_root / "runs" / safe_filename(args.dataset_dir.name, "dataset")  # type: ignore[operator]

    result = scaffold(
        args.dataset_dir,
        run_dir,
        force=args.force,
        output_root=args.output_root,
        extraction_mode=args.extraction_mode,
        allow_positional_pdf_fallback=args.allow_positional_pdf_fallback,
        authoritative_table=args.authoritative_table,
        authoritative_sheet=args.authoritative_sheet,
        allow_template_only=args.allow_template_only,
    )
    if args.json:
        print(json.dumps(result, indent=2))
    else:
        print(f"Scaffolded incomplete review input: {result['review_input']}")
        print(f"PDFs: {result['pdfs']} Rows: {result['rows']} Columns: {result['columns']}")
        print(
            f"Mapping: {result['mapping_mode']} ({result['mapped_rows']} mapped rows, "
            f"{result['unmapped_rows']} table-only rows)"
        )
        print(
            f"Target cells: {result['blank_target_cells']} blank, "
            f"{result['populated_target_cells']} populated; mode={result['extraction_mode']}"
        )
        print("Add evidence-backed proposals, then run build_and_serve_review.py.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        raise SystemExit(1)
