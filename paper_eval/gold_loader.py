from __future__ import annotations

import csv
from csv import Error as CsvError
from pathlib import Path
from typing import Any

from paper_eval.contracts import GoldCell, GoldDataset
from paper_eval.errors import ContractError
from paper_eval.normalize import is_empty_value


def load_gold(path: Path, *, sheet_name: str | None = None) -> GoldDataset:
    if not path.exists():
        raise ContractError(f"Gold input does not exist: {path}")
    if not path.is_file():
        raise ContractError(f"Gold input is not a file: {path}")
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        return _load_csv_gold(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx_gold(path, sheet_name=sheet_name)
    raise ContractError(f"Unsupported gold file type: {path.suffix}")


def _load_csv_gold(path: Path) -> GoldDataset:
    try:
        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)
            fieldnames = list(reader.fieldnames or [])
    except CsvError as exc:
        raise ContractError(f"Gold CSV could not be parsed at {path}: {exc}") from exc
    if not fieldnames:
        raise ContractError(f"Gold CSV '{path}' is empty or missing a header row.")
    cells = _rows_to_gold_cells(rows, fieldnames, sheet_name=None)
    return GoldDataset(source_path=path, sheet_name=None, cells=cells)


def _load_xlsx_gold(path: Path, *, sheet_name: str | None) -> GoldDataset:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ContractError(
            "XLSX gold inputs require openpyxl; install requirements.txt before scoring XLSX files."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        selected_sheet_name = sheet_name or workbook.sheetnames[0]
        if selected_sheet_name not in workbook.sheetnames:
            raise ContractError(
                f"Worksheet '{selected_sheet_name}' was not found in {path.name}. "
                f"Available sheets: {', '.join(workbook.sheetnames)}"
            )

        worksheet = workbook[selected_sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ContractError(f"Gold worksheet '{selected_sheet_name}' is empty.")
        fieldnames = [str(value).strip() if value is not None else "" for value in rows[0]]
        data_rows = [dict(zip(fieldnames, values)) for values in rows[1:]]
        cells = _rows_to_gold_cells(data_rows, fieldnames, sheet_name=selected_sheet_name)
        return GoldDataset(source_path=path, sheet_name=selected_sheet_name, cells=cells)
    finally:
        workbook.close()


def _rows_to_gold_cells(
    rows: list[dict[str, Any]],
    fieldnames: list[str],
    *,
    sheet_name: str | None,
) -> list[GoldCell]:
    fieldname_set = set(fieldnames)
    if {"row_id", "column_name", "gold_value"}.issubset(fieldname_set):
        cells = _load_long_form_rows(rows, sheet_name=sheet_name)
    else:
        cells = _load_wide_form_rows(rows, fieldnames, sheet_name=sheet_name)
    _validate_unique_gold_join_keys(cells)
    return cells


def _load_long_form_rows(rows: list[dict[str, Any]], *, sheet_name: str | None) -> list[GoldCell]:
    cells: list[GoldCell] = []
    for row in rows:
        row_id = _required_join_value(row.get("row_id"), "row_id")
        column_name = _required_join_value(row.get("column_name"), "column_name")
        raw_value = row.get("gold_value")
        cells.append(
            GoldCell(
                row_id=row_id,
                column_name=column_name,
                cell_id=_optional_text(row.get("cell_id")),
                row_index=_optional_int(row.get("row_index")),
                raw_value=raw_value,
                is_present=not is_empty_value(raw_value),
                sheet_name=sheet_name,
            )
        )
    return cells


def _load_wide_form_rows(
    rows: list[dict[str, Any]],
    fieldnames: list[str],
    *,
    sheet_name: str | None,
) -> list[GoldCell]:
    if "row_id" not in fieldnames:
        raise ContractError(
            "Gold wide-format inputs must include a 'row_id' column to support stable joins."
        )
    reserved_columns = {"row_id", "row_index"}
    data_columns = [
        name
        for name in fieldnames
        if name and name not in reserved_columns and not name.endswith("__cell_id")
    ]
    cells: list[GoldCell] = []
    for row in rows:
        row_id = _required_join_value(row.get("row_id"), "row_id")
        row_index = _optional_int(row.get("row_index"))
        for column_name in data_columns:
            raw_value = row.get(column_name)
            cells.append(
                GoldCell(
                    row_id=row_id,
                    column_name=column_name,
                    cell_id=_optional_text(row.get(f"{column_name}__cell_id")),
                    row_index=row_index,
                    raw_value=raw_value,
                    is_present=not is_empty_value(raw_value),
                    sheet_name=sheet_name,
                )
            )
    return cells


def _required_join_value(value: Any, field_name: str) -> str:
    text = _optional_text(value)
    if not text:
        raise ContractError(f"Gold input is missing required stable join field '{field_name}'.")
    return text


def _optional_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _optional_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    return int(value)


def _validate_unique_gold_join_keys(cells: list[GoldCell]) -> None:
    seen: dict[tuple[str, str], GoldCell] = {}
    for cell in cells:
        join_key = cell.join_key
        if join_key in seen:
            raise ContractError(
                "Gold input publishes duplicate stable join keys; each scored cell must have a unique "
                f"row_id + column_name pair. Duplicate: row_id='{cell.row_id}', column_name='{cell.column_name}'."
            )
        seen[join_key] = cell
