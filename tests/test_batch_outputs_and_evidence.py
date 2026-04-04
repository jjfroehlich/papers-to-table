import csv
import json
import tempfile
import unittest
from pathlib import Path

import pyarrow.parquet as pq
from openpyxl import load_workbook

from paper_eval.cli import main
from paper_eval.evidence import validate_evidence_anchors
from paper_eval.run_loader import load_run


class EvidenceValidationTests(unittest.TestCase):
    def test_anchor_is_valid_when_quote_is_locatable(self) -> None:
        result = validate_evidence_anchors(
            [
                self._evidence_item(
                    page=1,
                    quote_text="result was positive",
                )
            ],
            page_text_by_page={1: "The result was positive in the trial cohort."},
            page_count=3,
        )

        self.assertEqual(result.outcome, "anchor_valid")
        self.assertTrue(result.anchor_valid)
        self.assertFalse(result.evidence_present_but_unvalidated)

    def test_evidence_present_but_unvalidated_when_text_is_unavailable(self) -> None:
        result = validate_evidence_anchors(
            [
                self._evidence_item(
                    page=1,
                    quote_text="result was positive",
                )
            ],
            page_text_by_page={},
            page_count=3,
        )

        self.assertEqual(result.outcome, "evidence_present_but_unvalidated")
        self.assertFalse(result.anchor_valid)
        self.assertTrue(result.evidence_present_but_unvalidated)

    def test_anchor_is_invalid_when_quote_cannot_be_located(self) -> None:
        result = validate_evidence_anchors(
            [
                self._evidence_item(
                    page=2,
                    quote_text="result was positive",
                )
            ],
            page_text_by_page={2: "A different sentence appears here."},
            page_count=3,
        )

        self.assertEqual(result.outcome, "anchor_invalid")
        self.assertFalse(result.anchor_valid)
        self.assertFalse(result.evidence_present_but_unvalidated)

    def _evidence_item(self, *, page: int, quote_text: str) -> object:
        from paper_eval.contracts import EvidenceItem

        return EvidenceItem(page=page, quote_text=quote_text)


class BatchEvaluationTests(unittest.TestCase):
    def test_run_loader_flattens_metadata_from_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            run_dir = self._create_run_bundle(
                base / "run-a",
                run_payload={
                    "run_id": "run-a",
                    "mode": "batch",
                    "provider": {"model_id": "text-model-a", "vision_model_id": "vision-a", "token": "provider-a"},
                    "parser": {"identity": "parser-a", "version": "1.2.3"},
                    "prompt": {"id": "prompt-17", "hash": "prompt-hash-a"},
                    "schema": {"version": "schema-v2"},
                },
                config_payload={"config": {"hash": "cfg-a"}},
                include_page_text=True,
            )

            loaded_run = load_run(run_dir)
            metadata = loaded_run.metadata.flat_metadata()

            self.assertEqual(metadata["mode"], "batch")
            self.assertEqual(metadata["model_id"], "text-model-a")
            self.assertEqual(metadata["vision_model_id"], "vision-a")
            self.assertEqual(metadata["parser_identity"], "parser-a")
            self.assertEqual(metadata["parser_version"], "1.2.3")
            self.assertEqual(metadata["parser_identity_version"], "parser-a@1.2.3")
            self.assertEqual(metadata["prompt_identity"], "prompt-17")
            self.assertEqual(metadata["schema_identity"], "schema-v2")
            self.assertEqual(metadata["config_hash"], "cfg-a")
            self.assertEqual(metadata["run__provider__model_id"], "text-model-a")
            self.assertEqual(metadata["config_snapshot__config__hash"], "cfg-a")

    def test_evaluate_writes_batch_outputs_with_one_row_per_run(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            run_a = self._create_run_bundle(
                base / "run-a",
                run_payload={
                    "run_id": "run-a",
                    "mode": "batch",
                    "provider": {"model_id": "text-model-a", "vision_model_id": "vision-a"},
                    "parser": {"identity": "parser-a", "version": "1.2.3"},
                    "prompt": {"id": "prompt-17"},
                    "schema": {"version": "schema-v2"},
                },
                config_payload={"config": {"hash": "cfg-a"}},
                include_page_text=True,
            )
            run_b = self._create_run_bundle(
                base / "run-b",
                run_payload={
                    "run_id": "run-b",
                    "run_mode": "batch",
                    "provider_text_model_id": "text-model-b",
                    "provider_vision_model_id": "vision-b",
                    "parser_identity": "parser-b",
                    "parser_version": "2.0.0",
                    "prompt_hash": "prompt-hash-b",
                    "schema_hash": "schema-hash-b",
                },
                config_payload={"config_hash": "cfg-b"},
                include_page_text=False,
            )
            gold_path = base / "gold.csv"
            gold_path.write_text("row_id,status\nrow-1,yes\n", encoding="utf-8")
            output_dir = base / "out"

            exit_code = main(
                [
                    "evaluate",
                    "--run",
                    str(run_a),
                    "--run",
                    str(run_b),
                    "--gold",
                    str(gold_path),
                    "--out",
                    str(output_dir),
                ]
            )

            self.assertEqual(exit_code, 0)
            compare_dir = output_dir / "compare"
            csv_path = compare_dir / "runs_comparison.csv"
            xlsx_path = compare_dir / "runs_comparison.xlsx"
            parquet_path = compare_dir / "runs_comparison.parquet"
            self.assertTrue(csv_path.exists())
            self.assertTrue(xlsx_path.exists())
            self.assertTrue(parquet_path.exists())

            csv_rows = self._read_csv(csv_path)
            self.assertEqual(len(csv_rows), 2)
            row_a = self._find_row(csv_rows, "run-a")
            row_b = self._find_row(csv_rows, "run-b")
            self.assertEqual(row_a["model_id"], "text-model-a")
            self.assertEqual(row_a["prompt_identity"], "prompt-17")
            self.assertEqual(row_a["anchor_valid_rate"], "1.0")
            self.assertEqual(row_a["correct_and_anchored_rate"], "1.0")
            self.assertEqual(row_b["model_id"], "text-model-b")
            self.assertEqual(row_b["prompt_identity"], "prompt-hash-b")
            self.assertEqual(row_b["anchor_valid_rate"], "0.0")
            self.assertEqual(row_b["evidence_present_but_unvalidated_count"], "1")

            workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
            worksheet = workbook["runs_comparison"]
            workbook_rows = list(worksheet.iter_rows(values_only=True))
            self.assertEqual(len(workbook_rows), 3)
            self.assertIn("run_id", workbook_rows[0])

            parquet_rows = pq.read_table(parquet_path).to_pylist()
            self.assertEqual(len(parquet_rows), 2)
            self.assertEqual({row["run_id"] for row in parquet_rows}, {"run-a", "run-b"})

            summary_a = json.loads((output_dir / "per-run" / "run-a" / "run_summary.json").read_text(encoding="utf-8"))
            summary_b = json.loads((output_dir / "per-run" / "run-b" / "run_summary.json").read_text(encoding="utf-8"))
            self.assertEqual(summary_a["metrics"]["anchor_valid_rate"], 1.0)
            self.assertEqual(summary_b["metrics"]["anchor_valid_rate"], 0.0)
            self.assertEqual(summary_b["metrics"]["evidence_present_but_unvalidated_count"], 1)

    def test_compare_command_rebuilds_comparison_outputs_from_per_run_summaries(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            run_dir = self._create_run_bundle(
                base / "run-a",
                run_payload={"run_id": "run-a", "provider_text_model_id": "model-a"},
                config_payload={"config_hash": "cfg-a"},
                include_page_text=True,
            )
            gold_path = base / "gold.csv"
            gold_path.write_text("row_id,status\nrow-1,yes\n", encoding="utf-8")
            output_dir = base / "out"
            rebuilt_dir = base / "rebuilt-compare"

            self.assertEqual(
                main(
                    [
                        "evaluate",
                        "--run",
                        str(run_dir),
                        "--gold",
                        str(gold_path),
                        "--out",
                        str(output_dir),
                    ]
                ),
                0,
            )

            self.assertEqual(
                main(
                    [
                        "compare",
                        "--summaries",
                        str(output_dir / "per-run"),
                        "--out",
                        str(rebuilt_dir),
                    ]
                ),
                0,
            )

            rebuilt_rows = self._read_csv(rebuilt_dir / "runs_comparison.csv")
            self.assertEqual(len(rebuilt_rows), 1)
            self.assertEqual(rebuilt_rows[0]["run_id"], "run-a")

    def _create_run_bundle(
        self,
        run_dir: Path,
        *,
        run_payload: dict[str, object],
        config_payload: dict[str, object],
        include_page_text: bool,
    ) -> Path:
        (run_dir / "proposals").mkdir(parents=True)
        (run_dir / "inputs").mkdir(parents=True)
        (run_dir / "summaries").mkdir(parents=True)
        (run_dir / "run.json").write_text(json.dumps(run_payload), encoding="utf-8")
        (run_dir / "config.snapshot.json").write_text(json.dumps(config_payload), encoding="utf-8")
        (run_dir / "inputs" / "input_summary.json").write_text(json.dumps({"page_count": 3}), encoding="utf-8")
        (run_dir / "summaries" / "run_summary.json").write_text(json.dumps({"status": "complete"}), encoding="utf-8")
        proposal_rows = [
            {
                "run_id": run_payload["run_id"],
                "row_id": "row-1",
                "column_name": "status",
                "cell_id": "cell-status-1",
                "proposed_value": "yes",
                "field_type": "boolean",
                "evidence": [{"page": 1, "quote_text": "result was positive"}],
            }
        ]
        with (run_dir / "proposals" / "proposals.jsonl").open("w", encoding="utf-8") as handle:
            for row in proposal_rows:
                handle.write(json.dumps(row))
                handle.write("\n")
        if include_page_text:
            (run_dir / "evidence").mkdir(parents=True)
            (run_dir / "evidence" / "page_text.json").write_text(
                json.dumps({"1": "The result was positive in the trial cohort."}),
                encoding="utf-8",
            )
        return run_dir

    def _read_csv(self, path: Path) -> list[dict[str, str]]:
        with path.open("r", encoding="utf-8", newline="") as handle:
            return list(csv.DictReader(handle))

    def _find_row(self, rows: list[dict[str, str]], run_id: str) -> dict[str, str]:
        for row in rows:
            if row["run_id"] == run_id:
                return row
        raise AssertionError(f"Missing row for {run_id}")


if __name__ == "__main__":
    unittest.main()
