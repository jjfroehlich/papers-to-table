"""Tests for the staged asyncio runner."""
from __future__ import annotations

import asyncio
import pathlib
from datetime import datetime, timezone
from types import SimpleNamespace

import openpyxl
import pytest
import respx
import httpx
import pandas as pd

from backend.app.artifacts import (
    get_artifact_summary_path,
    get_config_snapshot_path,
    get_input_summary_path,
    get_provider_diagnostics_path,
    get_provider_mode_path,
    get_provider_probe_path,
    get_run_json_path,
    get_run_stats_path,
    get_reviewer_summary_path,
    get_run_summary_path,
    init_run_bundle,
    read_json,
    write_json,
)
from backend.app.extraction import ProposalRecord, load_proposals, persist_proposal
from backend.app.config import RunConfig
from backend.app.runner import get_initial_run_data, run_pipeline
from backend.app.ids import generate_proposal_id
from backend.app.matching import MatchResult
from backend.app.schemas import MatchOutcome, ProposalState, RunStatus, SupportLabel

FIXTURE_TABLE = "tests/fixtures/tables/literature_fixture.xlsx"
FIXTURE_SCHEMA = "tests/fixtures/tables/literature_fixture_schema.csv"
FIXTURE_PDF_DIR = "tests/fixtures/papers"
_CAPTURED_ROW_CONTEXTS: list[dict] = []


def make_config(tmp_path: pathlib.Path, **kwargs) -> RunConfig:
    data = {
        "table_path": FIXTURE_TABLE,
        "schema_path": FIXTURE_SCHEMA,
        "pdf_dir": FIXTURE_PDF_DIR,
        "output_dir": str(tmp_path / "runs"),
        "verify_mode": False,
        "provider": {
            "token": "lm_studio",
            "base_url": "http://localhost:1234",
            "text_model": {
                "model_id": "qwen/qwen3-30b-a3b-2507",
            },
        },
        **kwargs,
    }
    return RunConfig.model_validate(data)


class TestGetInitialRunData:
    def test_has_required_fields(self, tmp_path):
        config = make_config(tmp_path)
        data = get_initial_run_data("run_test_123", config, "config.json")
        assert data["run_id"] == "run_test_123"
        assert data["status"] == RunStatus.created.value
        assert data["config_path"] == "config.json"
        assert data["verify_mode"] is False
        assert data["eval_mode"] is False
        assert data["run_mode"] == "normal"
        assert data["provider_token"] == "lm_studio"
        assert data["retrieval_mode"] == "lexical"
        assert data["prompt_hash"] is not None
        assert data["prompt_bundle_id"] == "default"
        assert data["prompt_manifest_hash"] is not None
        assert data["prompt_bundle_hash"] is not None
        assert data["prompt_bundle_path"] is not None
        assert data["prompt_keys_used"]
        assert data["prompt_files"]

    def test_timestamps(self, tmp_path):
        config = make_config(tmp_path)
        data = get_initial_run_data("run_test_123", config, None)
        assert data["started_at"] is None
        assert data["completed_at"] is None
        assert data["created_at"] is not None


class TestRunPipeline:
    @pytest.mark.asyncio
    @respx.mock
    async def test_completes_with_valid_inputs(self, tmp_path, monkeypatch):
        """Happy-path: pipeline completes when LM Studio is reachable."""
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)
        config = make_config(tmp_path)
        run_id = "run_test_happy"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        assert run_data["status"] in (
            RunStatus.completed.value,
            RunStatus.completed_with_warnings.value,
        )
        assert run_data["completed_at"] is not None

    @pytest.mark.asyncio
    @respx.mock
    async def test_writes_config_snapshot(self, tmp_path, monkeypatch):
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)
        config = make_config(tmp_path)
        run_id = "run_snap"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        snap_path = get_config_snapshot_path(output_dir, run_id)
        assert snap_path.exists()
        snap = read_json(snap_path)
        assert snap["provider"]["token"] == "lm_studio"
        assert snap["retrieval"]["mode"] == "lexical"
        assert "strategy" not in snap["retrieval"]

    @pytest.mark.asyncio
    @respx.mock
    async def test_writes_run_stats_summary(self, tmp_path, monkeypatch):
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)
        config = make_config(tmp_path)
        run_id = "run_stats"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        stats_path = get_run_stats_path(output_dir, run_id)
        assert stats_path.exists()
        stats = read_json(stats_path)
        assert stats["retrieval_mode"] == "lexical"
        assert stats["per_run"]["run_total_ms"] is not None
        assert "parse" in stats["per_run"]["stage_ms"]
        assert "provider_request_counts" in stats["counters"]

    @pytest.mark.asyncio
    @respx.mock
    async def test_writes_artifact_and_provider_diagnostic_summaries(self, tmp_path, monkeypatch):
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)
        monkeypatch.setattr("backend.app.runner.parse_pdf", _fake_parse_pdf)
        monkeypatch.setattr("backend.app.runner.run_matching", lambda **kwargs: [])
        monkeypatch.setattr("backend.app.runner.persist_match_artifacts", lambda *args, **kwargs: None)
        monkeypatch.setattr("backend.app.runner.run_style_profiles_stage", _fake_style_profiles)
        config = make_config(tmp_path)
        run_id = "run_artifact_summary"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        artifact_summary = read_json(get_artifact_summary_path(output_dir, run_id))
        provider_diagnostics = read_json(get_provider_diagnostics_path(output_dir, run_id))
        provider_probe = read_json(get_provider_probe_path(output_dir, run_id))
        run_data = read_json(get_run_json_path(output_dir, run_id))

        assert artifact_summary["files"]["provider_diagnostics"]["present"] is True
        assert artifact_summary["files"]["provider_probe"]["present"] is True
        assert artifact_summary["directories"]["exports"]["file_count"] == 0
        assert artifact_summary["directories"]["review"]["file_count"] == 0
        assert artifact_summary["directories"]["logs"]["file_count"] == 0
        assert artifact_summary["directories"]["diagnostics"]["file_count"] >= 3
        assert "diagnostics/provider_diagnostics.json" in artifact_summary["sections"]["diagnostics"]
        assert provider_diagnostics["attempt_count"] == 0
        assert provider_probe["provider"] == "lm_studio"
        assert run_data["artifact_summary_path"] == "summaries/artifact_summary.json"
        assert run_data["provider_diagnostics_path"] == "diagnostics/provider_diagnostics.json"
        assert run_data["provider_probe_path"] == "diagnostics/provider_probe.json"

    @pytest.mark.asyncio
    @respx.mock
    async def test_writes_input_summary(self, tmp_path, monkeypatch):
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)
        config = make_config(tmp_path)
        run_id = "run_inputs"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        input_path = get_input_summary_path(output_dir, run_id)
        assert input_path.exists()
        summary = read_json(input_path)
        assert summary["table_rows"] is not None
        assert summary["table_rows"] > 0
        assert summary["pdf_count"] is not None
        assert summary["run_mode"] == "normal"
        assert summary["retrieval_mode"] == "lexical"
        assert summary["prompt_hash"] is not None
        assert summary["prompt_bundle_id"] == "default"
        assert summary["prompt_manifest_hash"] is not None
        assert summary["prompt_bundle_hash"] is not None
        assert summary["prompt_bundle_path"] is not None
        assert summary["prompt_keys_used"]
        assert summary["prompt_files"]
        assert summary["config_hash"] is not None

    @pytest.mark.asyncio
    @respx.mock
    async def test_eval_mode_persists_masked_and_gold_artifacts(self, tmp_path, monkeypatch):
        _CAPTURED_ROW_CONTEXTS.clear()
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)

        workbook_path = tmp_path / "gold.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Title", "Authors", "Publication Year", "assay"])
        sheet.append(["Matched Paper", "Smith, J.", "2024", "STARR-seq"])
        workbook.save(workbook_path)
        pdf_dir = tmp_path / "pdfs"
        pdf_dir.mkdir()
        (pdf_dir / "paper_1.pdf").write_bytes(b"%PDF-1.4\n%stub")
        schema_path = tmp_path / "schema.csv"
        schema_path.write_text("column_name,description\nassay,Assay description\n", encoding="utf-8")

        config = make_config(
            tmp_path,
            table_path=str(workbook_path),
            schema_path=str(schema_path),
            pdf_dir=str(pdf_dir),
            eval_mode=True,
        )
        run_id = "run_eval_artifacts"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        monkeypatch.setattr("backend.app.runner.parse_pdf", _fake_parse_pdf)
        monkeypatch.setattr("backend.app.runner.run_matching", lambda **kwargs: [
            MatchResult(
                pdf_id="paper_1",
                pdf_path="paper_1.pdf",
                outcome=MatchOutcome.matched,
                matched_row_index=0,
                matched_row_title="Matched Paper",
                score=0.9,
                runner_up_score=0.1,
                runner_up_row_index=None,
                reasoning="clear match",
                blocked=False,
                matched_at=datetime.now(timezone.utc).isoformat(),
            )
        ])
        monkeypatch.setattr("backend.app.runner.persist_match_artifacts", lambda *args, **kwargs: None)
        monkeypatch.setattr("backend.app.runner.run_style_profiles_stage", _fake_style_profiles)
        monkeypatch.setattr("backend.app.runner.run_retrieval_for_cell", lambda **kwargs: None)
        monkeypatch.setattr("backend.app.runner.extract_cell", _fake_extract_cell_capture)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        input_summary = read_json(get_input_summary_path(output_dir, run_id))
        proposals = load_proposals(pathlib.Path(output_dir) / run_id)
        masked_workbook_path = pathlib.Path(output_dir) / run_id / "inputs" / "masked_working_table.xlsx"
        masked_sheet = openpyxl.load_workbook(masked_workbook_path).active
        header_row = [cell.value for cell in masked_sheet[1]]
        assay_column = header_row.index("assay") + 1

        assert run_data["run_mode"] == "eval"
        assert run_data["eval_mode"] is True
        assert input_summary["run_mode"] == "eval"
        assert run_data["prompt_hash"] is not None
        assert run_data["config_hash"] is not None
        assert run_data["schema_hash"] is not None
        assert run_data["eval_artifacts"]["gold_table"]["source_reference"] == str(workbook_path)
        assert run_data["eval_artifacts"]["gold_table"]["snapshot_path"] == "inputs/gold_table.xlsx"
        assert run_data["eval_artifacts"]["masked_working_table"]["path"] == "inputs/masked_working_table.xlsx"
        assert proposals[0].run_mode == "eval"
        assert proposals[0].prompt_hash == run_data["prompt_hash"]
        assert proposals[0].gold_table_hash == run_data["eval_artifacts"]["gold_table"]["content_hash"]
        assert proposals[0].masked_working_table_hash == run_data["eval_artifacts"]["masked_working_table"]["content_hash"]

        artifact_summary = read_json(get_artifact_summary_path(output_dir, run_id))
        assert artifact_summary["eval_artifact_parity"]["expected"] is True
        assert artifact_summary["eval_artifact_parity"]["gold_table_snapshot_present"] is True
        assert artifact_summary["eval_artifact_parity"]["masked_working_table_present"] is True
        assert artifact_summary["proposal_metadata_coverage"]["gold_table_snapshot_path_present"] == 1
        assert artifact_summary["proposal_metadata_coverage"]["masked_working_table_path_present"] == 1

        original_sheet = openpyxl.load_workbook(workbook_path).active
        original_value = original_sheet.cell(row=2, column=assay_column).value
        masked_value = masked_sheet.cell(row=2, column=assay_column).value
        assert original_value == "STARR-seq"
        assert masked_value in ("", None)
        assert _CAPTURED_ROW_CONTEXTS[-1]["assay"] == ""

    @pytest.mark.asyncio
    @respx.mock
    async def test_fails_with_unreachable_provider(self, tmp_path):
        respx.get("http://localhost:1234/v1/models").mock(
            side_effect=httpx.ConnectError("refused")
        )
        config = make_config(tmp_path)
        run_id = "run_fail"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        assert run_data["status"] == RunStatus.failed.value
        assert run_data["error_message"] is not None
        assert run_data["current_stage"] is None

    @pytest.mark.asyncio
    async def test_input_summary_written_before_readiness_check(self, tmp_path):
        """Input summary should exist even when readiness fails."""
        config = make_config(tmp_path, pdf_dir="/nonexistent/dir")
        run_id = "run_early_summary"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        assert run_data["status"] == RunStatus.failed.value
        assert run_data["current_stage"] is None

        # Input summary should still exist
        input_path = get_input_summary_path(output_dir, run_id)
        assert input_path.exists()

    @pytest.mark.asyncio
    @respx.mock
    async def test_records_total_rows_and_eligible_cells(self, tmp_path, monkeypatch):
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)
        config = make_config(tmp_path)
        run_id = "run_counts"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        assert run_data["total_rows"] > 0
        assert run_data["eligible_cells"] >= 0

    @pytest.mark.asyncio
    async def test_provider_init_failure_does_not_generate_skipped_proposals(self, tmp_path, monkeypatch):
        config = make_config(tmp_path)
        run_id = "run_no_skipped"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        monkeypatch.setattr("backend.app.runner.check_readiness", _ready_ok)
        monkeypatch.setattr("backend.app.runner.initialize_provider", _raise_provider_error)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        proposals_dir = pathlib.Path(output_dir) / run_id / "proposals"

        assert run_data["status"] == RunStatus.failed.value
        assert run_data["error_message"] == "provider offline"
        assert list(proposals_dir.glob("*.json")) == []
        assert not (proposals_dir / "proposals.jsonl").exists()

    @pytest.mark.asyncio
    @respx.mock
    async def test_json_object_fallback_surfaces_provider_degraded_warning(self, tmp_path, monkeypatch):
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        config = make_config(tmp_path)
        run_id = "run_provider_degraded"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider_json_object)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        warnings = run_data.get("warnings", [])
        degraded = [w for w in warnings if w.get("category") == "provider_degraded"]
        assert len(degraded) >= 1
        assert "json_object" in degraded[0].get("message", "")
        assert degraded[0].get("context", {}).get("structured_output_reason") == "structured_backend_incompatible"
        assert run_data.get("structured_output_mode") == "json_object"
        assert run_data.get("structured_output_fallback_used") is True
        assert run_data.get("provider_readiness_reason") == "structured_backend_incompatible"

        provider_mode = read_json(get_provider_mode_path(output_dir, run_id))
        assert provider_mode.get("structured_output_mode") == "json_object"
        assert provider_mode.get("structured_output_fallback_used") is True

    @pytest.mark.asyncio
    @respx.mock
    async def test_prompt_only_json_fallback_surfaces_provider_degraded_warning(self, tmp_path, monkeypatch):
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        config = make_config(tmp_path)
        run_id = "run_provider_prompt_only"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider_prompt_only)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        warnings = run_data.get("warnings", [])
        degraded = [w for w in warnings if w.get("category") == "provider_degraded"]
        assert len(degraded) >= 1
        assert "prompt-only" in degraded[0].get("message", "")
        assert degraded[0].get("context", {}).get("structured_output_reason") == "structured_backend_incompatible"
        assert run_data.get("structured_output_mode") == "none"
        assert run_data.get("structured_output_fallback_used") is True
        assert run_data.get("provider_readiness_reason") == "structured_backend_incompatible"

        provider_mode = read_json(get_provider_mode_path(output_dir, run_id))
        assert provider_mode.get("structured_output_mode") == "none"
        assert provider_mode.get("structured_output_fallback_used") is True

    @pytest.mark.asyncio
    async def test_provider_init_model_unavailable_is_not_classified_as_unreachable(self, tmp_path, monkeypatch):
        config = make_config(tmp_path)
        run_id = "run_model_unavailable"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        monkeypatch.setattr("backend.app.runner.check_readiness", _ready_ok)
        monkeypatch.setattr("backend.app.runner.initialize_provider", _raise_model_unavailable_error)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        assert run_data["status"] == RunStatus.failed.value
        assert run_data["provider_readiness_reason"] == "model_unavailable"
        categories = [w.get("category") for w in run_data.get("warnings", [])]
        assert "model_unavailable" in categories
        assert "provider_unreachable" not in categories

    @pytest.mark.asyncio
    async def test_only_usable_matched_rows_create_proposal_artifacts(self, tmp_path, monkeypatch):
        config = make_config(tmp_path)
        run_id = "run_matched_only"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        df = pd.DataFrame([
            {
                "Title": "Matched Paper",
                "Authors": "Smith, J.",
                "Publication Year": "2024",
                "assay": "",
            },
            {
                "Title": "Unmatched Paper",
                "Authors": "Doe, A.",
                "Publication Year": "2023",
                "assay": "",
            },
        ])
        schema = [{"column_name": "assay", "description": "Assay description"}]
        eligible = [
            {"row_index": 0, "column_name": "assay", "current_value": "", "eligibility": "eligible"},
            {"row_index": 1, "column_name": "assay", "current_value": "", "eligibility": "eligible"},
        ]
        match_results = [
            MatchResult(
                pdf_id="paper_1",
                pdf_path="paper_1.pdf",
                outcome=MatchOutcome.matched,
                matched_row_index=0,
                matched_row_title="Matched Paper",
                score=0.9,
                runner_up_score=0.2,
                runner_up_row_index=1,
                reasoning="clear match",
                blocked=False,
                matched_at=datetime.now(timezone.utc).isoformat(),
            ),
            MatchResult(
                pdf_id="paper_2",
                pdf_path="paper_2.pdf",
                outcome=MatchOutcome.unmatched,
                matched_row_index=None,
                matched_row_title=None,
                score=0.0,
                runner_up_score=0.0,
                runner_up_row_index=None,
                reasoning="unmatched",
                blocked=True,
                blocked_reason="unmatched",
                matched_at=datetime.now(timezone.utc).isoformat(),
            ),
        ]

        monkeypatch.setattr("backend.app.runner.check_readiness", _ready_ok)
        monkeypatch.setattr("backend.app.runner.load_table", lambda path: df)
        monkeypatch.setattr("backend.app.runner.validate_metadata_columns", lambda df: [])
        monkeypatch.setattr("backend.app.runner.load_schema", lambda schema_path, table_path: schema)
        monkeypatch.setattr("backend.app.runner.validate_schema_columns", lambda schema: [])
        monkeypatch.setattr(
            "backend.app.runner.get_eligible_cells",
            lambda df, schema, verify_mode, eval_mode=False: eligible,
        )
        monkeypatch.setattr("backend.app.runner.parse_pdf", _fake_parse_pdf)
        monkeypatch.setattr("backend.app.runner.run_matching", lambda **kwargs: match_results)
        monkeypatch.setattr("backend.app.runner.persist_match_artifacts", lambda *args, **kwargs: None)
        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)
        monkeypatch.setattr("backend.app.runner.run_style_profiles_stage", _fake_style_profiles)
        monkeypatch.setattr("backend.app.runner.run_retrieval_for_cell", lambda **kwargs: None)
        monkeypatch.setattr("backend.app.runner.extract_cell", _fake_extract_cell)

        await run_pipeline(run_id, config, "config.json", output_dir)

        proposals_dir = pathlib.Path(output_dir) / run_id / "proposals"
        run_data = read_json(get_run_json_path(output_dir, run_id))

        assert run_data["proposals_generated"] == 1
        assert (proposals_dir / "proposals.jsonl").exists()
        assert (proposals_dir / "proposal_index.json").exists()

    @pytest.mark.asyncio
    async def test_parse_and_evidence_fallback_truth_surfaces_in_run_warnings(self, tmp_path, monkeypatch):
        config = make_config(tmp_path)
        run_id = "run_warning_truth"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        df = pd.DataFrame([
            {
                "Title": "Matched Paper",
                "Authors": "Smith, J.",
                "Publication Year": "2024",
                "assay": "",
            }
        ])
        schema = [{"column_name": "assay", "description": "Assay description"}]
        eligible = [{"row_index": 0, "column_name": "assay", "current_value": "", "eligibility": "eligible"}]
        match_results = [
            MatchResult(
                pdf_id="paper_1",
                pdf_path="paper_1.pdf",
                outcome=MatchOutcome.matched,
                matched_row_index=0,
                matched_row_title="Matched Paper",
                score=0.9,
                runner_up_score=0.2,
                runner_up_row_index=None,
                reasoning="clear match",
                blocked=False,
                matched_at=datetime.now(timezone.utc).isoformat(),
            ),
        ]

        monkeypatch.setattr("backend.app.runner.check_readiness", _ready_ok)
        monkeypatch.setattr("backend.app.runner.load_table", lambda path: df)
        monkeypatch.setattr("backend.app.runner.validate_metadata_columns", lambda df: [])
        monkeypatch.setattr("backend.app.runner.load_schema", lambda schema_path, table_path: schema)
        monkeypatch.setattr("backend.app.runner.validate_schema_columns", lambda schema: [])
        monkeypatch.setattr(
            "backend.app.runner.get_eligible_cells",
            lambda df, schema, verify_mode, eval_mode=False: eligible,
        )
        monkeypatch.setattr("backend.app.runner.parse_pdf", _fake_parse_pdf_with_warnings)
        monkeypatch.setattr("backend.app.runner.run_matching", lambda **kwargs: match_results)
        monkeypatch.setattr("backend.app.runner.persist_match_artifacts", lambda *args, **kwargs: None)
        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)
        monkeypatch.setattr("backend.app.runner.run_style_profiles_stage", _fake_style_profiles)
        monkeypatch.setattr("backend.app.runner.run_retrieval_for_cell", lambda **kwargs: None)
        monkeypatch.setattr("backend.app.runner.extract_cell", _fake_extract_cell_with_fallback)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        warning_messages = [warning["message"] for warning in run_data["warnings"]]

        assert any("Parser fallback used" in message for message in warning_messages)
        assert any("OCR fallback used" in message for message in warning_messages)
        assert any("require evidence fallback review" in message for message in warning_messages)


async def _ready_ok(*args, **kwargs):
    return SimpleNamespace(ok=True, errors=[])


async def _raise_provider_error(*args, **kwargs):
    from backend.app.provider import ProviderError

    raise ProviderError("provider offline")


async def _raise_model_unavailable_error(*args, **kwargs):
    from backend.app.provider import ProviderError

    raise ProviderError("configured model not loaded", reason="model_unavailable")


async def _fake_initialize_provider(*args, **kwargs):
    return object(), SimpleNamespace(
        mode="live_local",
        locality="local",
        readiness_error=None,
        readiness_reason=None,
        structured_output_mode="json_schema",
        structured_output_fallback_used=False,
        capabilities=None,
        model_dump=lambda: {
            "mode": "live_local",
            "locality": "local",
            "readiness_error": None,
            "readiness_reason": None,
            "structured_output_mode": "json_schema",
            "structured_output_fallback_used": False,
        },
    )


async def _fake_initialize_provider_json_object(*args, **kwargs):
    return object(), SimpleNamespace(
        mode="live_local",
        locality="local",
        readiness_error=None,
        readiness_reason="structured_backend_incompatible",
        structured_output_mode="json_object",
        structured_output_fallback_used=True,
        capabilities=SimpleNamespace(
            structured_output_mode="json_object",
            structured_output_reason="structured_backend_incompatible",
            structured_output_error="LM Studio rejected structured-output grammar/regex constraints: Failed to process regex",
        ),
        model_dump=lambda: {
            "mode": "live_local",
            "locality": "local",
            "readiness_error": None,
            "readiness_reason": "structured_backend_incompatible",
            "structured_output_mode": "json_object",
            "structured_output_fallback_used": True,
            "capabilities": {
                "structured_output_mode": "json_object",
                "structured_output_reason": "structured_backend_incompatible",
                "structured_output_error": "LM Studio rejected structured-output grammar/regex constraints: Failed to process regex",
            },
        },
    )


async def _fake_initialize_provider_prompt_only(*args, **kwargs):
    return object(), SimpleNamespace(
        mode="live_local",
        locality="local",
        readiness_error=None,
        readiness_reason="structured_backend_incompatible",
        structured_output_mode="none",
        structured_output_fallback_used=True,
        capabilities=SimpleNamespace(
            structured_output_mode="none",
            structured_output_reason="structured_backend_incompatible",
            structured_output_error="LM Studio rejected structured-output grammar/regex constraints: Failed to process regex",
        ),
        model_dump=lambda: {
            "mode": "live_local",
            "locality": "local",
            "readiness_error": None,
            "readiness_reason": "structured_backend_incompatible",
            "structured_output_mode": "none",
            "structured_output_fallback_used": True,
            "capabilities": {
                "structured_output_mode": "none",
                "structured_output_reason": "structured_backend_incompatible",
                "structured_output_error": "LM Studio rejected structured-output grammar/regex constraints: Failed to process regex",
            },
        },
    )


async def _fake_style_profiles(**kwargs):
    return {}


def _fake_parse_pdf(**kwargs):
    pdf_id = kwargs["pdf_id"]
    doc = SimpleNamespace(model_dump=lambda: {"pdf_id": pdf_id, "blocks": []})
    diagnostics = SimpleNamespace(
        fallback_used=False,
        actual_parser_used="docling",
        configured_parser="docling",
        ocr_used=False,
        ocr_reason=None,
        parse_warnings=[],
        major_extraction_gaps=[],
    )
    return doc, diagnostics, []


def _fake_parse_pdf_with_warnings(**kwargs):
    pdf_id = kwargs["pdf_id"]
    doc = SimpleNamespace(model_dump=lambda: {"pdf_id": pdf_id, "blocks": []})
    diagnostics = SimpleNamespace(
        fallback_used=True,
        actual_parser_used="pypdfium2",
        configured_parser="docling",
        ocr_used=True,
        ocr_reason="low text extraction",
        parse_warnings=["Low text extraction (scanned PDF)."],
        major_extraction_gaps=["Sparse extracted text"],
    )
    return doc, diagnostics, []


async def _fake_extract_cell(**kwargs):
    proposal_id = generate_proposal_id(kwargs["run_id"], kwargs["cell_id"])
    artifact_context = kwargs.get("artifact_context") or {}
    proposal = ProposalRecord(
        proposal_id=proposal_id,
        run_id=kwargs["run_id"],
        pdf_id=kwargs["pdf_id"],
        row_id=kwargs["row_id"],
        column_name=kwargs["column_name"],
        cell_id=kwargs["cell_id"],
        state=ProposalState.unclear,
        support=SupportLabel.weak_evidence,
        proposed_value="candidate",
        rationale="- extracted",
        evidence_ids=[],
        warning_flags=[],
        needs_more_evidence=False,
        is_verify_mode=False,
        run_mode=artifact_context.get("run_mode", "normal"),
        prompt_version=artifact_context.get("prompt_version"),
        prompt_hash=artifact_context.get("prompt_hash"),
        schema_hash=artifact_context.get("schema_hash"),
        schema_version=artifact_context.get("schema_version"),
        config_hash=artifact_context.get("config_hash"),
        config_snapshot_path=artifact_context.get("config_snapshot_path"),
        parser_identity=artifact_context.get("parser_identity"),
        parser_version=artifact_context.get("parser_version"),
        text_model_id=kwargs.get("text_model_id"),
        vision_model_id=kwargs.get("vision_model_id"),
        gold_table_source_reference=artifact_context.get("gold_table_source_reference"),
        gold_table_hash=artifact_context.get("gold_table_hash"),
        gold_table_snapshot_path=artifact_context.get("gold_table_snapshot_path"),
        masked_working_table_path=artifact_context.get("masked_working_table_path"),
        masked_working_table_hash=artifact_context.get("masked_working_table_hash"),
        created_at=datetime.now(timezone.utc).isoformat(),
    )
    persist_proposal(kwargs["run_dir"], proposal)
    return proposal


async def _fake_extract_cell_with_fallback(**kwargs):
    proposal_id = generate_proposal_id(kwargs["run_id"], kwargs["cell_id"])
    proposal = ProposalRecord(
        proposal_id=proposal_id,
        run_id=kwargs["run_id"],
        pdf_id=kwargs["pdf_id"],
        row_id=kwargs["row_id"],
        column_name=kwargs["column_name"],
        cell_id=kwargs["cell_id"],
        state=ProposalState.unclear,
        support=SupportLabel.weak_evidence,
        proposed_value="candidate",
        rationale="- extracted",
        evidence_ids=[],
        warning_flags=["fallback_evidence_used"],
        needs_more_evidence=False,
        is_verify_mode=False,
        created_at=datetime.now(timezone.utc).isoformat(),
    )
    persist_proposal(kwargs["run_dir"], proposal)
    return proposal


async def _fake_extract_cell_capture(**kwargs):
    _CAPTURED_ROW_CONTEXTS.append(dict(kwargs["row_context"]))
    return await _fake_extract_cell(**kwargs)
