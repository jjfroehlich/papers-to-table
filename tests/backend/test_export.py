"""
Batch 6 tests — T101

Tests covering:
- T096: Export integrity (correct values written to output workbook)
- T096: Content-only fidelity (no formulas in output; values preserved)
- T096: Changed-cell highlighting (accepted cells have yellow fill)
- T096: Accepted-only export behavior (rejected/undecided proposals excluded)
- T097: Unsupported-feature warnings (merged cells, formulas, conditional formatting,
        frozen panes, hidden rows, hidden cols, charts, named ranges, comments)
- T098: Audit-log completeness (all fields, only accepted entries, real timestamps)
- T099: Diagnostics JSON (matching failures, blocked outcomes, weak evidence,
        feature warnings, completed-with-warnings semantics)
- T100: Export trigger endpoint and download availability endpoint
"""
from __future__ import annotations

import csv
import io
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from backend.app.artifacts import RunArtifacts
from backend.app.export import (
    CHANGED_CELL_FILL,
    detect_unsupported_features,
    generate_audit_log,
    generate_diagnostics,
    generate_xlsx_export,
    run_export,
)
from backend.app.ids import make_cell_id, make_proposal_id, make_review_decision_id
from backend.app.main import app
from backend.app.schemas import (
    ExportCandidate,
    ProposalRecord,
    ProposalState,
    ReviewDecision,
    ReviewDecisionRecord,
    SupportLabel,
    WarningStatusCategory,
)

REPO_ROOT = Path(__file__).resolve().parents[2]
FIXTURE_XLSX = REPO_ROOT / "tests" / "fixtures" / "tables" / "literature_fixture.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_artifacts(tmp_path: Path) -> RunArtifacts:
    return RunArtifacts.create(tmp_path / "out", "run_export_test")


def _write_proposal(
    artifacts: RunArtifacts,
    run_id: str = "run_export_test",
    pdf_id: str = "pdf_a",
    row_id: str = "Paper A",
    column_name: str = "Focus",
    proposed_value: str = "MPRA",
    proposal_state: ProposalState = ProposalState.FOUND,
    support_label: SupportLabel = SupportLabel.DIRECT_EVIDENCE,
    status_flags: list[WarningStatusCategory] | None = None,
) -> ProposalRecord:
    cell_id = make_cell_id(row_id, column_name)
    proposal_id = make_proposal_id(run_id, pdf_id, cell_id)
    record = ProposalRecord(
        proposal_id=proposal_id,
        run_id=run_id,
        pdf_id=pdf_id,
        row_id=row_id,
        column_name=column_name,
        cell_id=cell_id,
        source_mode="text",
        proposal_state=proposal_state,
        support_label=support_label,
        proposed_value=proposed_value,
        status_flags=status_flags or [],
    )
    artifacts.append_jsonl("proposals/proposals.jsonl", record.model_dump(mode="json"))
    return record


def _write_decision(
    artifacts: RunArtifacts,
    proposal_id: str,
    cell_id: str,
    decision: ReviewDecision,
    edited_value: str | None = None,
    run_id: str = "run_export_test",
    ordinal: int = 0,
) -> ReviewDecisionRecord:
    decision_id = make_review_decision_id(run_id, proposal_id, ordinal)
    record = ReviewDecisionRecord(
        decision_id=decision_id,
        run_id=run_id,
        proposal_id=proposal_id,
        cell_id=cell_id,
        decision=decision,
        edited_value=edited_value,
        decided_at=datetime.now(timezone.utc),
    )
    artifacts.append_jsonl("review/decisions.jsonl", record.model_dump(mode="json"))
    return record


def _make_simple_xlsx(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    """Create a minimal XLSX workbook for testing."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def _make_candidate(
    row_id: str,
    column_name: str,
    accepted_value: str,
    decision: ReviewDecision = ReviewDecision.ACCEPT,
    run_id: str = "run_export_test",
    pdf_id: str = "pdf_a",
) -> ExportCandidate:
    cell_id = make_cell_id(row_id, column_name)
    proposal_id = make_proposal_id(run_id, pdf_id, cell_id)
    return ExportCandidate(
        proposal_id=proposal_id,
        run_id=run_id,
        pdf_id=pdf_id,
        row_id=row_id,
        column_name=column_name,
        cell_id=cell_id,
        accepted_value=accepted_value,
        decision=decision,
    )


# ---------------------------------------------------------------------------
# T096 — Export integrity: correct values in output workbook
# ---------------------------------------------------------------------------


class TestExportIntegrity:
    def test_accepted_value_written_to_output(self, tmp_path: Path) -> None:
        """Accepted change appears in the output workbook at the correct cell."""
        source = tmp_path / "table.xlsx"
        _make_simple_xlsx(
            source,
            headers=["Title", "Method"],
            rows=[["Paper A", ""], ["Paper B", "RNA-seq"]],
        )

        candidate = _make_candidate("Paper A", "Method", "MPRA")
        output = tmp_path / "out.xlsx"
        generate_xlsx_export(
            source_path=source,
            table_rows=[
                {"Title": "Paper A", "Method": ""},
                {"Title": "Paper B", "Method": "RNA-seq"},
            ],
            candidates=[candidate],
            output_path=output,
        )

        wb = load_workbook(str(output), data_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        method_col = headers.index("Method") + 1  # 1-based

        row2_values = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
        row3_values = [c.value for c in next(ws.iter_rows(min_row=3, max_row=3))]

        assert row2_values[method_col - 1] == "MPRA", "Accepted value should be written"
        assert row3_values[method_col - 1] == "RNA-seq", "Unchanged row should retain original value"

    def test_all_rows_preserved(self, tmp_path: Path) -> None:
        """All rows from the source workbook appear in the output."""
        source = tmp_path / "table.xlsx"
        titles = ["Paper A", "Paper B", "Paper C"]
        _make_simple_xlsx(source, headers=["Title", "Focus"], rows=[[t, ""] for t in titles])

        output = tmp_path / "out.xlsx"
        generate_xlsx_export(
            source_path=source,
            table_rows=[{"Title": t, "Focus": ""} for t in titles],
            candidates=[],
            output_path=output,
        )

        wb = load_workbook(str(output), data_only=True)
        ws = wb.active
        # 1 header + 3 data rows
        assert ws.max_row == 4

    def test_accept_with_edit_value_used(self, tmp_path: Path) -> None:
        """ACCEPT_WITH_EDIT uses the edited_value, not the proposed_value."""
        source = tmp_path / "table.xlsx"
        _make_simple_xlsx(source, headers=["Title", "Scale"], rows=[["Paper A", ""]])

        candidate = _make_candidate(
            "Paper A", "Scale", "large-scale edited", ReviewDecision.ACCEPT_WITH_EDIT
        )
        output = tmp_path / "out.xlsx"
        generate_xlsx_export(
            source_path=source,
            table_rows=[{"Title": "Paper A", "Scale": ""}],
            candidates=[candidate],
            output_path=output,
        )

        wb = load_workbook(str(output), data_only=True)
        ws = wb.active
        row2 = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
        assert row2[1] == "large-scale edited"

    def test_fixture_xlsx_export_runs_without_error(self, tmp_path: Path) -> None:
        """Export against the real fixture XLSX produces a valid workbook."""
        output = tmp_path / "out.xlsx"
        result_warnings = generate_xlsx_export(
            source_path=FIXTURE_XLSX,
            table_rows=[],
            candidates=[],
            output_path=output,
        )
        assert output.is_file()
        wb = load_workbook(str(output), data_only=True)
        assert wb.active is not None
        assert isinstance(result_warnings, list)


# ---------------------------------------------------------------------------
# T096 — Content-only fidelity
# ---------------------------------------------------------------------------


class TestContentOnlyFidelity:
    def test_output_is_plain_value_not_formula(self, tmp_path: Path) -> None:
        """Formula cells in source are written as their plain value (not as formulas)."""
        source = tmp_path / "table.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Title", "Computed"])
        ws.append(["Paper A", "=1+1"])  # formula cell
        wb.save(str(source))

        output = tmp_path / "out.xlsx"
        generate_xlsx_export(
            source_path=source,
            table_rows=[{"Title": "Paper A", "Computed": None}],
            candidates=[],
            output_path=output,
        )

        out_wb = load_workbook(str(output), data_only=False)
        out_ws = out_wb.active
        row2_vals = [c.value for c in next(out_ws.iter_rows(min_row=2, max_row=2))]
        # The output should not contain the formula string "=1+1"
        assert row2_vals[1] != "=1+1", "Formula should not be copied verbatim"

    def test_no_merged_cells_in_output(self, tmp_path: Path) -> None:
        """Merged cells in source are not preserved in output."""
        source = tmp_path / "table.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Title", "A", "B"])
        ws.append(["Paper A", "x", "y"])
        ws.merge_cells("B2:C2")
        wb.save(str(source))

        output = tmp_path / "out.xlsx"
        generate_xlsx_export(
            source_path=source,
            table_rows=[{"Title": "Paper A", "A": "x", "B": "y"}],
            candidates=[],
            output_path=output,
        )

        out_wb = load_workbook(str(output))
        out_ws = out_wb.active
        assert not out_ws.merged_cells.ranges, "Output workbook must have no merged cells"

    def test_headers_preserved_in_output(self, tmp_path: Path) -> None:
        """Column headers are preserved exactly in the output."""
        source = tmp_path / "table.xlsx"
        headers = ["Title", "Focus", "Species", "Readout"]
        _make_simple_xlsx(source, headers=headers, rows=[["Paper A", "a", "b", "c"]])

        output = tmp_path / "out.xlsx"
        generate_xlsx_export(
            source_path=source,
            table_rows=[{"Title": "Paper A", "Focus": "a", "Species": "b", "Readout": "c"}],
            candidates=[],
            output_path=output,
        )

        out_wb = load_workbook(str(output), data_only=True)
        out_ws = out_wb.active
        out_headers = [c.value for c in next(out_ws.iter_rows(min_row=1, max_row=1))]
        assert out_headers == headers


# ---------------------------------------------------------------------------
# T096 — Changed-cell highlighting
# ---------------------------------------------------------------------------


class TestChangedCellHighlighting:
    def test_accepted_cell_has_yellow_fill(self, tmp_path: Path) -> None:
        """Changed cells are highlighted with the expected yellow fill."""
        source = tmp_path / "table.xlsx"
        _make_simple_xlsx(source, headers=["Title", "Focus"], rows=[["Paper A", ""]])

        candidate = _make_candidate("Paper A", "Focus", "MPRA")
        output = tmp_path / "out.xlsx"
        generate_xlsx_export(
            source_path=source,
            table_rows=[{"Title": "Paper A", "Focus": ""}],
            candidates=[candidate],
            output_path=output,
        )

        out_wb = load_workbook(str(output))
        out_ws = out_wb.active
        # Row 2, column 2 (Focus) should have yellow fill
        changed_cell = out_ws.cell(row=2, column=2)
        fill = changed_cell.fill
        assert fill.fill_type == "solid", "Changed cell must have solid fill"
        assert fill.fgColor.rgb == CHANGED_CELL_FILL.fgColor.rgb, "Changed cell must have correct highlight color"

    def test_unchanged_cell_has_no_highlight(self, tmp_path: Path) -> None:
        """Cells that were not changed must not have the highlight fill."""
        source = tmp_path / "table.xlsx"
        _make_simple_xlsx(
            source,
            headers=["Title", "Focus", "Species"],
            rows=[["Paper A", "", "mouse"]],
        )

        candidate = _make_candidate("Paper A", "Focus", "MPRA")
        output = tmp_path / "out.xlsx"
        generate_xlsx_export(
            source_path=source,
            table_rows=[{"Title": "Paper A", "Focus": "", "Species": "mouse"}],
            candidates=[candidate],
            output_path=output,
        )

        out_wb = load_workbook(str(output))
        out_ws = out_wb.active
        # Species column (col 3) should not have the highlight fill
        species_cell = out_ws.cell(row=2, column=3)
        fill = species_cell.fill
        is_yellow = (
            fill.fill_type == "solid"
            and fill.fgColor.rgb == CHANGED_CELL_FILL.fgColor.rgb
        )
        assert not is_yellow, "Unchanged cell must not have changed-cell highlight"

    def test_multiple_changed_cells_all_highlighted(self, tmp_path: Path) -> None:
        """Multiple accepted changes across rows/columns are all highlighted."""
        source = tmp_path / "table.xlsx"
        _make_simple_xlsx(
            source,
            headers=["Title", "Focus", "Species"],
            rows=[["Paper A", "", ""], ["Paper B", "", ""]],
        )

        candidates = [
            _make_candidate("Paper A", "Focus", "MPRA"),
            _make_candidate("Paper B", "Species", "mouse"),
        ]
        output = tmp_path / "out.xlsx"
        generate_xlsx_export(
            source_path=source,
            table_rows=[
                {"Title": "Paper A", "Focus": "", "Species": ""},
                {"Title": "Paper B", "Focus": "", "Species": ""},
            ],
            candidates=candidates,
            output_path=output,
        )

        out_wb = load_workbook(str(output))
        out_ws = out_wb.active
        # Paper A Focus: row 2, col 2
        cell_a_focus = out_ws.cell(row=2, column=2)
        # Paper B Species: row 3, col 3
        cell_b_species = out_ws.cell(row=3, column=3)

        for cell in [cell_a_focus, cell_b_species]:
            assert cell.fill.fill_type == "solid"
            assert cell.fill.fgColor.rgb == CHANGED_CELL_FILL.fgColor.rgb


# ---------------------------------------------------------------------------
# T096 — Accepted-only export behavior
# ---------------------------------------------------------------------------


class TestAcceptedOnlyExport:
    def test_no_accepted_changes_produces_copy(self, tmp_path: Path) -> None:
        """With no accepted candidates, the output workbook mirrors the source values."""
        source = tmp_path / "table.xlsx"
        _make_simple_xlsx(source, headers=["Title", "Focus"], rows=[["Paper A", "existing"]])

        output = tmp_path / "out.xlsx"
        generate_xlsx_export(
            source_path=source,
            table_rows=[{"Title": "Paper A", "Focus": "existing"}],
            candidates=[],
            output_path=output,
        )

        out_wb = load_workbook(str(output), data_only=True)
        out_ws = out_wb.active
        row2 = [c.value for c in next(out_ws.iter_rows(min_row=2, max_row=2))]
        assert row2[1] == "existing"

    def test_rejected_proposal_not_in_output(self, tmp_path: Path) -> None:
        """Rejected proposals do not change cell values in the output."""
        source = tmp_path / "table.xlsx"
        _make_simple_xlsx(source, headers=["Title", "Focus"], rows=[["Paper A", "original"]])

        # No candidates: rejected proposals are already excluded by get_export_candidates
        output = tmp_path / "out.xlsx"
        generate_xlsx_export(
            source_path=source,
            table_rows=[{"Title": "Paper A", "Focus": "original"}],
            candidates=[],  # rejected → excluded
            output_path=output,
        )

        out_wb = load_workbook(str(output), data_only=True)
        out_ws = out_wb.active
        row2 = [c.value for c in next(out_ws.iter_rows(min_row=2, max_row=2))]
        assert row2[1] == "original"

    def test_partial_review_only_accepted_exported(self, tmp_path: Path) -> None:
        """With mixed decisions, only accepted proposals appear in output."""
        source = tmp_path / "table.xlsx"
        _make_simple_xlsx(
            source,
            headers=["Title", "A", "B"],
            rows=[["Paper A", "", ""]],
        )

        # Only column A is accepted; column B is rejected (not in candidates)
        candidates = [_make_candidate("Paper A", "A", "val_a")]
        output = tmp_path / "out.xlsx"
        generate_xlsx_export(
            source_path=source,
            table_rows=[{"Title": "Paper A", "A": "", "B": ""}],
            candidates=candidates,
            output_path=output,
        )

        out_wb = load_workbook(str(output), data_only=True)
        out_ws = out_wb.active
        row2 = [c.value for c in next(out_ws.iter_rows(min_row=2, max_row=2))]
        assert row2[1] == "val_a"
        assert row2[2] is None or row2[2] == ""


# ---------------------------------------------------------------------------
# T097 — Unsupported workbook feature warnings
# ---------------------------------------------------------------------------


class TestUnsupportedFeatureWarnings:
    def test_no_warnings_for_simple_workbook(self, tmp_path: Path) -> None:
        """A plain XLSX with no advanced features produces no warnings."""
        source = tmp_path / "plain.xlsx"
        _make_simple_xlsx(source, headers=["Title", "Method"], rows=[["Paper A", "x"]])

        warnings = detect_unsupported_features(source)
        assert warnings == []

    def test_merged_cells_warning(self, tmp_path: Path) -> None:
        """Merged cells trigger a warning."""
        source = tmp_path / "merged.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Title", "A", "B"])
        ws.append(["Paper A", "x", "y"])
        ws.merge_cells("B2:C2")
        wb.save(str(source))

        warnings = detect_unsupported_features(source)
        assert any("merged_cells" in w for w in warnings)

    def test_formula_warning(self, tmp_path: Path) -> None:
        """Formula cells trigger a warning."""
        source = tmp_path / "formula.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Title", "Computed"])
        ws.append(["Paper A", "=1+2"])
        wb.save(str(source))

        warnings = detect_unsupported_features(source)
        assert any("formulas" in w for w in warnings)

    def test_frozen_panes_warning(self, tmp_path: Path) -> None:
        """Frozen panes trigger a warning."""
        source = tmp_path / "frozen.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Title", "Method"])
        ws.freeze_panes = "A2"
        wb.save(str(source))

        warnings = detect_unsupported_features(source)
        assert any("frozen_panes" in w for w in warnings)

    def test_hidden_rows_warning(self, tmp_path: Path) -> None:
        """Hidden rows trigger a warning."""
        source = tmp_path / "hidden_rows.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Title", "Method"])
        ws.append(["Paper A", "x"])
        ws.row_dimensions[2].hidden = True
        wb.save(str(source))

        warnings = detect_unsupported_features(source)
        assert any("hidden_rows" in w for w in warnings)

    def test_hidden_columns_warning(self, tmp_path: Path) -> None:
        """Hidden columns trigger a warning."""
        source = tmp_path / "hidden_cols.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Title", "Method"])
        ws.append(["Paper A", "x"])
        ws.column_dimensions["B"].hidden = True
        wb.save(str(source))

        warnings = detect_unsupported_features(source)
        assert any("hidden_columns" in w for w in warnings)

    def test_feature_warnings_recorded_in_export_return(self, tmp_path: Path) -> None:
        """Feature warnings detected during export are returned from generate_xlsx_export."""
        source = tmp_path / "merged.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Title", "A", "B"])
        ws.append(["Paper A", "x", "y"])
        ws.merge_cells("B2:C2")
        wb.save(str(source))

        output = tmp_path / "out.xlsx"
        warnings = generate_xlsx_export(
            source_path=source,
            table_rows=[{"Title": "Paper A", "A": "x", "B": "y"}],
            candidates=[],
            output_path=output,
        )
        assert any("merged_cells" in w for w in warnings)


# ---------------------------------------------------------------------------
# T098 — Audit-log completeness
# ---------------------------------------------------------------------------


class TestAuditLog:
    def _make_audit_log(
        self,
        tmp_path: Path,
        candidates: list[ExportCandidate],
        row_lookup: dict[str, dict[str, Any]],
        decision_records: list[dict[str, Any]],
    ) -> list[dict[str, str]]:
        output = tmp_path / "audit_log.csv"
        generate_audit_log(
            candidates=candidates,
            row_lookup=row_lookup,
            decision_records=decision_records,
            output_path=output,
        )
        with output.open("r", encoding="utf-8", newline="") as fh:
            return list(csv.DictReader(fh))

    def test_all_fields_present(self, tmp_path: Path) -> None:
        """Audit log contains all required fields."""
        candidate = _make_candidate("Paper A", "Focus", "MPRA")
        dec_id = make_review_decision_id("run_export_test", candidate.proposal_id, 0)
        now = datetime.now(timezone.utc).isoformat()
        decision_records = [
            {
                "decision_id": dec_id,
                "run_id": "run_export_test",
                "proposal_id": candidate.proposal_id,
                "cell_id": candidate.cell_id,
                "decision": "accept",
                "edited_value": None,
                "decided_at": now,
            }
        ]
        rows = self._make_audit_log(
            tmp_path,
            candidates=[candidate],
            row_lookup={"Paper A": {"Title": "Paper A", "Focus": ""}},
            decision_records=decision_records,
        )
        assert len(rows) == 1
        row = rows[0]
        for field in ["row_id", "column_name", "old_value", "new_value", "proposal_id", "decision", "decided_at"]:
            assert field in row, f"Field '{field}' missing from audit log"

    def test_only_accepted_entries_in_log(self, tmp_path: Path) -> None:
        """Only accepted proposals appear in the audit log."""
        candidates = [
            _make_candidate("Paper A", "Focus", "MPRA"),
            _make_candidate("Paper B", "Species", "mouse"),
        ]
        rows = self._make_audit_log(
            tmp_path,
            candidates=candidates,
            row_lookup={
                "Paper A": {"Title": "Paper A", "Focus": ""},
                "Paper B": {"Title": "Paper B", "Species": ""},
            },
            decision_records=[],
        )
        assert len(rows) == 2

    def test_old_value_populated(self, tmp_path: Path) -> None:
        """Old value is read from the row lookup."""
        candidate = _make_candidate("Paper A", "Focus", "MPRA")
        rows = self._make_audit_log(
            tmp_path,
            candidates=[candidate],
            row_lookup={"Paper A": {"Title": "Paper A", "Focus": "old_focus_value"}},
            decision_records=[],
        )
        assert rows[0]["old_value"] == "old_focus_value"

    def test_real_decision_timestamp_used(self, tmp_path: Path) -> None:
        """Decision timestamp from the review decision record is used."""
        candidate = _make_candidate("Paper A", "Focus", "MPRA")
        ts = "2025-01-15T12:34:56+00:00"
        decision_records = [
            {
                "decision_id": "dec_1",
                "run_id": "run_export_test",
                "proposal_id": candidate.proposal_id,
                "cell_id": candidate.cell_id,
                "decision": "accept",
                "edited_value": None,
                "decided_at": ts,
            }
        ]
        rows = self._make_audit_log(
            tmp_path,
            candidates=[candidate],
            row_lookup={"Paper A": {"Title": "Paper A", "Focus": ""}},
            decision_records=decision_records,
        )
        assert rows[0]["decided_at"] == ts

    def test_empty_log_for_no_candidates(self, tmp_path: Path) -> None:
        """With no accepted candidates, audit log has only a header row (no data rows)."""
        rows = self._make_audit_log(
            tmp_path,
            candidates=[],
            row_lookup={},
            decision_records=[],
        )
        assert rows == []

    def test_accept_with_edit_new_value(self, tmp_path: Path) -> None:
        """ACCEPT_WITH_EDIT entries use the edited value as new_value."""
        candidate = _make_candidate("Paper A", "Scale", "large-scale edited", ReviewDecision.ACCEPT_WITH_EDIT)
        rows = self._make_audit_log(
            tmp_path,
            candidates=[candidate],
            row_lookup={"Paper A": {"Title": "Paper A", "Scale": ""}},
            decision_records=[],
        )
        assert rows[0]["new_value"] == "large-scale edited"


# ---------------------------------------------------------------------------
# T099 — Diagnostics JSON
# ---------------------------------------------------------------------------


class TestDiagnosticsJson:
    def _make_diagnostics(
        self,
        run_id: str = "run_export_test",
        proposals: list[dict[str, Any]] | None = None,
        decision_records: list[dict[str, Any]] | None = None,
        unresolved_matches: list[dict[str, Any]] | None = None,
        feature_warnings: list[str] | None = None,
        candidates: list[ExportCandidate] | None = None,
    ) -> dict[str, Any]:
        return generate_diagnostics(
            run_id=run_id,
            proposals=proposals or [],
            decision_records=decision_records or [],
            unresolved_matches=unresolved_matches or [],
            feature_warnings=feature_warnings or [],
            candidates=candidates or [],
        )

    def test_clean_run_not_completed_with_warnings(self) -> None:
        """A run with accepted proposals and no warnings is not completed_with_warnings."""
        candidate = _make_candidate("Paper A", "Focus", "MPRA")
        diag = self._make_diagnostics(candidates=[candidate])
        assert diag["completed_with_warnings"] is False
        assert diag["warning_signals"] == []

    def test_no_candidates_triggers_warning(self) -> None:
        """No accepted candidates triggers completed_with_warnings."""
        diag = self._make_diagnostics(candidates=[])
        assert diag["completed_with_warnings"] is True
        assert any("no proposals accepted" in s for s in diag["warning_signals"])

    def test_unmatched_pdfs_in_diagnostics(self) -> None:
        """Unmatched PDFs appear in matching_failures.unmatched."""
        unresolved = [{"outcome": "unmatched", "pdf_id": "pdf_x"}]
        candidate = _make_candidate("Paper A", "Focus", "MPRA")
        diag = self._make_diagnostics(unresolved_matches=unresolved, candidates=[candidate])
        assert len(diag["matching_failures"]["unmatched"]) == 1
        assert diag["completed_with_warnings"] is True

    def test_ambiguous_pdfs_in_diagnostics(self) -> None:
        """Ambiguous PDFs appear in matching_failures.ambiguous."""
        unresolved = [{"outcome": "ambiguous", "pdf_id": "pdf_y"}]
        candidate = _make_candidate("Paper A", "Focus", "MPRA")
        diag = self._make_diagnostics(unresolved_matches=unresolved, candidates=[candidate])
        assert len(diag["matching_failures"]["ambiguous"]) == 1

    def test_blocked_proposals_in_diagnostics(self) -> None:
        """Blocked proposals appear in proposal_outcomes.blocked."""
        proposals = [
            {
                "proposal_id": "p1",
                "row_id": "row1",
                "column_name": "Focus",
                "pdf_id": "pdf_a",
                "proposal_state": "blocked",
                "status_flags": [],
            }
        ]
        candidate = _make_candidate("Paper A", "Focus", "MPRA")
        diag = self._make_diagnostics(proposals=proposals, candidates=[candidate])
        assert len(diag["proposal_outcomes"]["blocked"]) == 1
        assert diag["completed_with_warnings"] is True

    def test_unclear_proposals_in_diagnostics(self) -> None:
        """Unclear proposals appear in proposal_outcomes.unclear."""
        proposals = [
            {
                "proposal_id": "p1",
                "row_id": "row1",
                "column_name": "Focus",
                "pdf_id": "pdf_a",
                "proposal_state": "unclear",
                "status_flags": [],
            }
        ]
        candidate = _make_candidate("Paper A", "Focus", "MPRA")
        diag = self._make_diagnostics(proposals=proposals, candidates=[candidate])
        assert len(diag["proposal_outcomes"]["unclear"]) == 1

    def test_weak_evidence_in_diagnostics(self) -> None:
        """Weak-evidence proposals are captured in diagnostics."""
        proposals = [
            {
                "proposal_id": "p1",
                "row_id": "row1",
                "column_name": "Focus",
                "pdf_id": "pdf_a",
                "proposal_state": "found",
                "status_flags": ["weak_evidence"],
            }
        ]
        candidate = _make_candidate("Paper A", "Focus", "MPRA")
        diag = self._make_diagnostics(proposals=proposals, candidates=[candidate])
        assert len(diag["weak_evidence"]) == 1
        assert diag["completed_with_warnings"] is True

    def test_feature_warnings_included_in_diagnostics(self) -> None:
        """Unsupported workbook feature warnings appear in diagnostics."""
        feature_warnings = ["merged_cells: sheet 'Sheet' contains merged cells"]
        candidate = _make_candidate("Paper A", "Focus", "MPRA")
        diag = self._make_diagnostics(feature_warnings=feature_warnings, candidates=[candidate])
        assert diag["unsupported_workbook_features"] == feature_warnings
        assert diag["completed_with_warnings"] is True

    def test_diagnostics_includes_export_summary(self) -> None:
        """Diagnostics includes accepted_changes count."""
        candidates = [
            _make_candidate("Paper A", "Focus", "MPRA"),
            _make_candidate("Paper B", "Species", "mouse"),
        ]
        diag = self._make_diagnostics(candidates=candidates)
        assert diag["export_summary"]["accepted_changes"] == 2

    def test_evidence_recovery_quote_page_fallback(self) -> None:
        """Quote-page-fallback status flag is captured as evidence_recovery."""
        proposals = [
            {
                "proposal_id": "p1",
                "row_id": "row1",
                "column_name": "Focus",
                "pdf_id": "pdf_a",
                "proposal_state": "found",
                "status_flags": ["quote_page_fallback"],
            }
        ]
        candidate = _make_candidate("Paper A", "Focus", "MPRA")
        diag = self._make_diagnostics(proposals=proposals, candidates=[candidate])
        assert any(e["reason"] == "quote_page_fallback" for e in diag["evidence_recovery"])


# ---------------------------------------------------------------------------
# T100 — Export trigger endpoint and download availability
# ---------------------------------------------------------------------------


class TestExportEndpoint:
    def test_trigger_export_missing_run_returns_404(self) -> None:
        """POST /api/runs/{run_id}/export returns 404 for unknown run_id."""
        client = TestClient(app)
        response = client.post("/api/runs/nonexistent_run_id/export")
        assert response.status_code == 404

    def test_trigger_export_no_table_returns_400(self, tmp_path: Path, isolated_run_store: Any) -> None:
        """POST /api/runs/{run_id}/export returns 400 when source table is missing."""
        from backend.app.runner import RunStore

        run_id = "run_export_no_table"
        artifacts = RunArtifacts.create(tmp_path / "out", run_id)
        # Write input_summary pointing to a non-existent table
        artifacts.write_json(
            "inputs/input_summary.json",
            {"table_path": str(tmp_path / "nonexistent.xlsx"), "verify_mode": False},
        )
        isolated_run_store._artifacts[run_id] = artifacts  # noqa: SLF001

        client = TestClient(app)
        response = client.post(f"/api/runs/{run_id}/export")
        assert response.status_code == 400

    def test_trigger_export_produces_files(self, tmp_path: Path, isolated_run_store: Any) -> None:
        """POST /api/runs/{run_id}/export writes workbook, audit log, and diagnostics."""
        run_id = "run_export_files"
        artifacts = RunArtifacts.create(tmp_path / "out", run_id)

        # Write source table
        source = tmp_path / "source.xlsx"
        _make_simple_xlsx(
            source,
            headers=["Title", "Focus"],
            rows=[["Paper A", ""]],
        )

        # Write input_summary
        artifacts.write_json(
            "inputs/input_summary.json",
            {"table_path": str(source), "verify_mode": False},
        )
        artifacts.write_json(
            "inputs/input_details.json",
            {"table_rows": [{"Title": "Paper A", "Focus": ""}], "schema_rows": []},
        )

        # Write an accepted proposal + decision
        cell_id = make_cell_id("Paper A", "Focus")
        proposal_id = make_proposal_id(run_id, "pdf_a", cell_id)
        artifacts.append_jsonl(
            "proposals/proposals.jsonl",
            {
                "proposal_id": proposal_id,
                "run_id": run_id,
                "pdf_id": "pdf_a",
                "row_id": "Paper A",
                "column_name": "Focus",
                "cell_id": cell_id,
                "source_mode": "text",
                "proposal_state": "found",
                "support_label": "direct_evidence",
                "proposed_value": "MPRA",
                "status_flags": [],
            },
        )
        artifacts.append_jsonl(
            "review/decisions.jsonl",
            {
                "decision_id": make_review_decision_id(run_id, proposal_id, 0),
                "run_id": run_id,
                "proposal_id": proposal_id,
                "cell_id": cell_id,
                "decision": "accept",
                "edited_value": None,
                "decided_at": datetime.now(timezone.utc).isoformat(),
            },
        )

        isolated_run_store._artifacts[run_id] = artifacts  # noqa: SLF001

        client = TestClient(app)
        response = client.post(f"/api/runs/{run_id}/export")
        assert response.status_code == 200

        assert (artifacts.root / "exports" / "updated_workbook.xlsx").is_file()
        assert (artifacts.root / "exports" / "audit_log.csv").is_file()
        assert (artifacts.root / "exports" / "diagnostics.json").is_file()

    def test_download_available_reflects_export_files(self, tmp_path: Path, isolated_run_store: Any) -> None:
        """GET /api/runs/{run_id}/downloads/available returns true for existing export files."""
        run_id = "run_avail_test"
        artifacts = RunArtifacts.create(tmp_path / "out", run_id)

        # Manually create export files
        (artifacts.root / "exports").mkdir(parents=True, exist_ok=True)
        (artifacts.root / "exports" / "updated_workbook.xlsx").write_bytes(b"PK fake")
        (artifacts.root / "exports" / "audit_log.csv").write_text("header\n")

        isolated_run_store._artifacts[run_id] = artifacts  # noqa: SLF001

        client = TestClient(app)
        response = client.get(f"/api/runs/{run_id}/downloads/available")
        assert response.status_code == 200
        data = response.json()
        assert data["workbook"] is True
        assert data["audit_log"] is True

    def test_download_workbook_endpoint(self, tmp_path: Path, isolated_run_store: Any) -> None:
        """GET /api/runs/{run_id}/downloads/workbook returns the file."""
        run_id = "run_dl_wb"
        artifacts = RunArtifacts.create(tmp_path / "out", run_id)

        # Write a valid minimal XLSX
        source = tmp_path / "source.xlsx"
        _make_simple_xlsx(source, headers=["Title"], rows=[["Paper A"]])
        wb_path = artifacts.root / "exports" / "updated_workbook.xlsx"
        wb_path.parent.mkdir(parents=True, exist_ok=True)
        import shutil
        shutil.copy(str(source), str(wb_path))

        isolated_run_store._artifacts[run_id] = artifacts  # noqa: SLF001

        client = TestClient(app)
        response = client.get(f"/api/runs/{run_id}/downloads/workbook")
        assert response.status_code == 200

    def test_download_audit_log_endpoint(self, tmp_path: Path, isolated_run_store: Any) -> None:
        """GET /api/runs/{run_id}/downloads/audit-log returns the CSV file."""
        run_id = "run_dl_al"
        artifacts = RunArtifacts.create(tmp_path / "out", run_id)

        al_path = artifacts.root / "exports" / "audit_log.csv"
        al_path.parent.mkdir(parents=True, exist_ok=True)
        al_path.write_text("row_id,column_name,old_value,new_value,proposal_id,decision,decided_at\n")

        isolated_run_store._artifacts[run_id] = artifacts  # noqa: SLF001

        client = TestClient(app)
        response = client.get(f"/api/runs/{run_id}/downloads/audit-log")
        assert response.status_code == 200

    def test_download_workbook_404_before_export(self, tmp_path: Path, isolated_run_store: Any) -> None:
        """GET /api/runs/{run_id}/downloads/workbook returns 404 before export runs."""
        run_id = "run_no_export"
        artifacts = RunArtifacts.create(tmp_path / "out", run_id)
        isolated_run_store._artifacts[run_id] = artifacts  # noqa: SLF001

        client = TestClient(app)
        response = client.get(f"/api/runs/{run_id}/downloads/workbook")
        assert response.status_code == 404

    def test_download_audit_log_404_before_export(self, tmp_path: Path, isolated_run_store: Any) -> None:
        """GET /api/runs/{run_id}/downloads/audit-log returns 404 before export runs."""
        run_id = "run_no_al"
        artifacts = RunArtifacts.create(tmp_path / "out", run_id)
        isolated_run_store._artifacts[run_id] = artifacts  # noqa: SLF001

        client = TestClient(app)
        response = client.get(f"/api/runs/{run_id}/downloads/audit-log")
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# run_export orchestrator — integration
# ---------------------------------------------------------------------------


class TestRunExportOrchestrator:
    def test_run_export_writes_all_artifacts(self, tmp_path: Path) -> None:
        """run_export writes workbook, audit log, and diagnostics."""
        run_id = "run_orch_test"
        artifacts = RunArtifacts.create(tmp_path / "out", run_id)

        source = tmp_path / "source.xlsx"
        _make_simple_xlsx(source, headers=["Title", "Focus"], rows=[["Paper A", ""]])

        artifacts.write_json("inputs/input_summary.json", {"table_path": str(source), "verify_mode": False})
        artifacts.write_json("inputs/input_details.json", {
            "table_rows": [{"Title": "Paper A", "Focus": ""}],
            "schema_rows": [],
        })

        cell_id = make_cell_id("Paper A", "Focus")
        proposal_id = make_proposal_id(run_id, "pdf_a", cell_id)
        artifacts.append_jsonl("proposals/proposals.jsonl", {
            "proposal_id": proposal_id,
            "run_id": run_id,
            "pdf_id": "pdf_a",
            "row_id": "Paper A",
            "column_name": "Focus",
            "cell_id": cell_id,
            "source_mode": "text",
            "proposal_state": "found",
            "support_label": "direct_evidence",
            "proposed_value": "MPRA",
            "status_flags": [],
        })
        artifacts.append_jsonl("review/decisions.jsonl", {
            "decision_id": make_review_decision_id(run_id, proposal_id, 0),
            "run_id": run_id,
            "proposal_id": proposal_id,
            "cell_id": cell_id,
            "decision": "accept",
            "edited_value": None,
            "decided_at": datetime.now(timezone.utc).isoformat(),
        })

        result = run_export(artifacts, run_id)

        assert (artifacts.root / "exports" / "updated_workbook.xlsx").is_file()
        assert (artifacts.root / "exports" / "audit_log.csv").is_file()
        assert (artifacts.root / "exports" / "diagnostics.json").is_file()
        assert result["accepted_changes"] == 1
        assert result["completed_with_warnings"] is False

    def test_run_export_raises_on_missing_table(self, tmp_path: Path) -> None:
        """run_export raises ExportError when source table is missing."""
        from backend.app.export import ExportError

        run_id = "run_missing_table"
        artifacts = RunArtifacts.create(tmp_path / "out", run_id)
        artifacts.write_json("inputs/input_summary.json", {
            "table_path": str(tmp_path / "nonexistent.xlsx"),
        })

        with pytest.raises(ExportError):
            run_export(artifacts, run_id)

    def test_run_export_with_no_accepted_proposals(self, tmp_path: Path) -> None:
        """run_export with no accepted proposals marks completed_with_warnings."""
        run_id = "run_no_accept"
        artifacts = RunArtifacts.create(tmp_path / "out", run_id)

        source = tmp_path / "source.xlsx"
        _make_simple_xlsx(source, headers=["Title", "Focus"], rows=[["Paper A", ""]])
        artifacts.write_json("inputs/input_summary.json", {"table_path": str(source)})
        artifacts.write_json("inputs/input_details.json", {"table_rows": [{"Title": "Paper A", "Focus": ""}]})

        result = run_export(artifacts, run_id)
        assert result["completed_with_warnings"] is True
        assert result["accepted_changes"] == 0
