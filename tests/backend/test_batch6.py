from __future__ import annotations

import json
import os
import time
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from pypdf import PdfWriter

from app.main import app


client = TestClient(app)


def _wait_for_terminal(run_id: str) -> dict:
    for _ in range(500):
        payload = client.get(f"/api/runs/{run_id}").json()
        if payload["status"] in {"completed", "completed_with_warnings", "failed"}:
            return payload
        time.sleep(0.05)
    raise AssertionError("run did not reach terminal state")


def _make_pdf(path: Path, title: str, author: str, year: int) -> None:
    writer = PdfWriter()
    writer.add_blank_page(width=612, height=792)
    writer.add_metadata({"/Title": title, "/Author": author, "/CreationDate": f"D:{year}0101000000"})
    with path.open("wb") as handle:
        writer.write(handle)


def _start_run(cfg: Path) -> tuple[str, dict]:
    created = client.post("/api/runs", json={"config_path": str(cfg)}).json()
    run = _wait_for_terminal(created["run_id"])
    assert run["status"] in {"completed", "completed_with_warnings"}
    return created["run_id"], run


def test_batch6_export_and_diagnostics(tmp_path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.append(["Title", "Authors", "Publication Year", "Material"])
    ws.append(["Alpha Study", "Alice Smith", 2020, ""]) 
    ws.freeze_panes = "A2"
    ws.merge_cells("A3:B3")
    ws["D2"] = "=1+1"
    ws.auto_filter.ref = "A1:D2"
    table = tmp_path / "table.xlsx"
    workbook.save(table)

    schema = tmp_path / "schema.csv"
    schema.write_text("column_name,description\nMaterial,Primary material\n", encoding="utf-8")

    pdf_dir = tmp_path / "pdfs"
    pdf_dir.mkdir()
    _make_pdf(pdf_dir / "alpha.pdf", "Alpha Study", "Alice Smith", 2020)

    cfg = tmp_path / "config.json"
    cfg.write_text(
        json.dumps(
            {
                "paths": {
                    "table_path": str(table),
                    "schema_path": str(schema),
                    "pdf_dir": str(pdf_dir),
                    "output_dir": str(tmp_path / "artifacts"),
                },
                "review": {"verify_mode": True},
                "export": {"highlight_color": "FFFF00"},
            }
        ),
        encoding="utf-8",
    )

    run_id, run = _start_run(cfg)
    proposals = client.get(f"/api/runs/{run_id}/proposals").json()["items"]
    target = proposals[0]

    decision = client.post(
        f"/api/runs/{run_id}/review/decisions",
        json={"proposal_id": target["proposal_id"], "decision": "accept_edited", "edited_value": "Graphene"},
    )
    assert decision.status_code == 200

    exported = client.post(f"/api/runs/{run_id}/export")
    assert exported.status_code == 200

    manifest = client.get(f"/api/runs/{run_id}/downloads").json()["downloads"]
    assert manifest["updated_workbook"]["ready"] is True
    assert manifest["audit_log"]["ready"] is True

    run_dir = Path(run["artifact_dir"])
    updated = load_workbook(run_dir / "exports" / "updated.xlsx")
    updated_ws = updated.active
    assert updated_ws["D2"].value == "Graphene"
    assert updated_ws["D2"].fill.start_color.rgb in {"00FFFF00", "FFFF00"}

    audit_rows = [
        json.loads(line)
        for line in (run_dir / "exports" / "audit_log.jsonl").read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    assert audit_rows
    assert audit_rows[0]["reviewer_decision"] == "accept_edited"
    assert audit_rows[0]["decision_timestamp"]

    diagnostics = json.loads((run_dir / "logs" / "diagnostics.json").read_text(encoding="utf-8"))
    assert "frozen_panes" in diagnostics["unsupported_workbook_features"]

    refreshed = client.get(f"/api/runs/{run_id}").json()
    assert refreshed["status"] == "completed_with_warnings"


def test_batch6_hermetic_e2e_flows(tmp_path: Path) -> None:
    table = tmp_path / "table.csv"
    table.write_text(
        "\n".join(
            [
                "Title,Authors,Publication Year,Material,Figure Metric",
                "Alpha Study,Alice Smith,2020,,existing",
                "Gamma Study,Gina West,2021,,baseline",
            ]
        ),
        encoding="utf-8",
    )
    schema = tmp_path / "schema.csv"
    schema.write_text(
        "\n".join(
            [
                "column_name,description",
                "Material,Text-derived material",
                "Figure Metric,Value from figure chart",
            ]
        ),
        encoding="utf-8",
    )
    pdf_dir = tmp_path / "pdfs"
    pdf_dir.mkdir()
    _make_pdf(pdf_dir / "alpha_1.pdf", "Alpha Study", "Alice Smith", 2020)
    _make_pdf(pdf_dir / "alpha_2.pdf", "Alpha Study", "Alice Smith", 2020)
    _make_pdf(pdf_dir / "gamma.pdf", "Gamma Study", "Gina West", 2021)
    _make_pdf(pdf_dir / "unmatched.pdf", "Different Study", "Unknown", 2010)

    cfg = tmp_path / "config.json"
    cfg.write_text(
        json.dumps(
            {
                "paths": {
                    "table_path": str(table),
                    "schema_path": str(schema),
                    "pdf_dir": str(pdf_dir),
                    "output_dir": str(tmp_path / "artifacts"),
                },
                "review": {"verify_mode": True},
            }
        ),
        encoding="utf-8",
    )

    run_id, run = _start_run(cfg)
    listing = client.get(f"/api/runs/{run_id}/proposals").json()
    assert listing["counters"]["total"] > 0
    assert any(item["match_outcome"] == "duplicate_row_conflict" for item in listing["items"])
    assert any(item["match_outcome"] == "unmatched" for item in listing["items"])
    assert any(item["support_label"] == "figure_based_evidence" for item in listing["items"])

    weak_item = next(
        item
        for item in listing["items"]
        if "weak_evidence" in item.get("warning_categories", []) and item.get("match_outcome") == "matched"
    )
    detail = client.get(f"/api/runs/{run_id}/proposals/{weak_item['proposal_id']}").json()
    assert "weak_evidence" in detail["warning_status_flags"]

    verify_target = next(
        item
        for item in listing["items"]
        if item["column_name"] == "Figure Metric" and item["row_id"] == "row_1"
    )
    decision = client.post(
        f"/api/runs/{run_id}/review/decisions",
        json={"proposal_id": verify_target["proposal_id"], "decision": "accept_edited", "edited_value": "manual-verified-value"},
    )
    assert decision.status_code == 200

    export_res = client.post(f"/api/runs/{run_id}/export")
    assert export_res.status_code == 200

    run_dir = Path(run["artifact_dir"])
    audit_rows = [
        json.loads(line)
        for line in (run_dir / "exports" / "audit_log.jsonl").read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    assert len(audit_rows) == 1
    assert audit_rows[0]["proposal_id"] == verify_target["proposal_id"]


def test_batch6_live_smoke_opt_in(tmp_path: Path) -> None:
    if os.getenv("PTA_LIVE_SMOKE") != "1":
        pytest.skip("Set PTA_LIVE_SMOKE=1 to run live LM Studio smoke coverage")

    table = tmp_path / "table.csv"
    table.write_text(
        "Title,Authors,Publication Year,Material\nAlpha Study,Alice Smith,2020,\n",
        encoding="utf-8",
    )
    schema = tmp_path / "schema.csv"
    schema.write_text("column_name,description\nMaterial,Primary material\n", encoding="utf-8")
    pdf_dir = tmp_path / "pdfs"
    pdf_dir.mkdir()
    _make_pdf(pdf_dir / "alpha.pdf", "Alpha Study", "Alice Smith", 2020)

    cfg = tmp_path / "config.json"
    cfg.write_text(
        json.dumps(
            {
                "paths": {
                    "table_path": str(table),
                    "schema_path": str(schema),
                    "pdf_dir": str(pdf_dir),
                    "output_dir": str(tmp_path / "artifacts"),
                },
                "provider": {"enable_live_calls": True, "base_url": "http://localhost:1234/v1", "model_name": "local-model"},
            }
        ),
        encoding="utf-8",
    )

    _, run = _start_run(cfg)
    assert run["status"] in {"completed", "completed_with_warnings"}


def test_batch6_performance_smoke(tmp_path: Path) -> None:
    schema = tmp_path / "schema.csv"
    schema.write_text("column_name,description\nMaterial,Primary material\n", encoding="utf-8")

    def _run_with_count(prefix: str, count: int) -> float:
        table = tmp_path / f"{prefix}_table.csv"
        rows = ["Title,Authors,Publication Year,Material"]
        pdf_dir = tmp_path / f"{prefix}_pdfs"
        pdf_dir.mkdir()
        for idx in range(count):
            title = f"Study {idx}"
            author = f"Author {idx}"
            year = 2020 + (idx % 4)
            rows.append(f"{title},{author},{year},")
            _make_pdf(pdf_dir / f"{prefix}_{idx}.pdf", title, author, year)
        table.write_text("\n".join(rows), encoding="utf-8")

        cfg = tmp_path / f"{prefix}_config.json"
        cfg.write_text(
            json.dumps(
                {
                    "paths": {
                        "table_path": str(table),
                        "schema_path": str(schema),
                        "pdf_dir": str(pdf_dir),
                        "output_dir": str(tmp_path / f"{prefix}_artifacts"),
                    }
                }
            ),
            encoding="utf-8",
        )
        start = time.perf_counter()
        _start_run(cfg)
        return time.perf_counter() - start

    small = _run_with_count("small", 2)
    medium = _run_with_count("medium", 8)

    assert small < 8.0
    assert medium < 20.0
