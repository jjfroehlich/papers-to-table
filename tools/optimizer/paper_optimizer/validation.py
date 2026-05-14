from __future__ import annotations

import csv
import shutil
from pathlib import Path
from typing import Any

from .benchmarks import Benchmarks
from .contracts import Candidate, LaunchResult
from .utils import read_json


class PreflightError(ValueError):
    pass


def _is_non_empty_string(value: Any) -> bool:
    return isinstance(value, str) and bool(value.strip())


def _existing_path(path_str: str | None) -> bool:
    return bool(path_str) and Path(path_str).exists()


def _command_head_exists(token: str, working_dir: Path | None = None) -> bool:
    if not token or "{" in token:
        return True
    token_path = Path(token)
    if token_path.is_absolute():
        return token_path.exists()
    if any(sep in token for sep in ("/", "\\")):
        if working_dir is None:
            return token_path.exists()
        return (working_dir / token_path).exists()
    return shutil.which(token) is not None


def _script_token_exists(token: str, working_dir: Path | None = None) -> bool:
    if not token or "{" in token:
        return True
    if not any(token.endswith(ext) for ext in (".py", ".exe", ".bat", ".cmd", ".ps1", ".sh")):
        return True
    token_path = Path(token)
    if token_path.is_absolute():
        return token_path.exists()
    if working_dir is not None:
        return (working_dir / token_path).exists()
    return token_path.exists()


def _validate_command_tokens(
    *,
    section_name: str,
    command: list[str],
    errors: list[str],
    working_dir: Path | None,
) -> None:
    if not command:
        errors.append(f"{section_name} command is empty")
        return
    if not _command_head_exists(command[0], working_dir=working_dir):
        errors.append(f"{section_name} command executable is not available: {command[0]}")
    for token in command[1:]:
        if not _script_token_exists(token, working_dir=working_dir):
            errors.append(f"{section_name} command references a missing script or executable: {token}")


def _candidate_prompt_ids(config: dict[str, Any]) -> list[str]:
    prompt_ids: list[str] = []
    baseline = config.get("baseline_candidate", {})
    if _is_non_empty_string(baseline.get("prompt_bundle_id")):
        prompt_ids.append(str(baseline["prompt_bundle_id"]))

    for row in config.get("compare_candidates", []) or []:
        if isinstance(row, dict) and _is_non_empty_string(row.get("prompt_bundle_id")):
            prompt_ids.append(str(row["prompt_bundle_id"]))

    search_space = config.get("search_space", {})
    if isinstance(search_space, dict):
        for prompt_id in search_space.get("prompt_bundle_ids", []):
            if _is_non_empty_string(prompt_id):
                prompt_ids.append(str(prompt_id))

    seen: set[str] = set()
    ordered: list[str] = []
    for prompt_id in prompt_ids:
        if prompt_id in seen:
            continue
        seen.add(prompt_id)
        ordered.append(prompt_id)
    return ordered


def _candidate_prompt_bundle_roots(main_working_dir: Path) -> list[Path]:
    return [
        main_working_dir / "backend" / "app" / "prompt_bundles",
        main_working_dir / "backend" / "src" / "backend" / "app" / "prompt_bundles",
        main_working_dir / "src" / "backend" / "app" / "prompt_bundles",
    ]


def _mapping_has_target(mapping: dict[str, str] | list[str] | None, target_name: str) -> bool:
    if mapping is None:
        return False
    if isinstance(mapping, list):
        return target_name in mapping
    return target_name in mapping


def _looks_like_fixture_path(path_str: str | None) -> bool:
    if not _is_non_empty_string(path_str):
        return False
    normalized = str(path_str).replace("\\", "/").lower()
    return "tests/fixtures" in normalized or "configs/benchmarks" in normalized


def _eval_arg_value(eval_args: list[str], flag: str) -> str | None:
    for index, token in enumerate(eval_args):
        if token != flag:
            continue
        if index + 1 < len(eval_args):
            value = eval_args[index + 1]
            if _is_non_empty_string(value):
                return str(value)
    return None


def _validate_judge_contract(*, benchmark_id: str, eval_args: list[str], required_judges: list[str], errors: list[str]) -> None:
    judge_a = _eval_arg_value(eval_args, "--judge-model")
    judge_b = _eval_arg_value(eval_args, "--judge-model-b")
    judge_api_base_b = _eval_arg_value(eval_args, "--judge-api-base-b")
    if "--judge-model-b" in eval_args and not judge_b:
        errors.append(f"benchmarks.manifests.{benchmark_id}.eval_args declares --judge-model-b without a model id")
    if judge_api_base_b and not judge_b:
        errors.append(f"benchmarks.manifests.{benchmark_id}.eval_args declares --judge-api-base-b without --judge-model-b")
    if judge_api_base_b and not judge_a:
        errors.append(f"benchmarks.manifests.{benchmark_id}.eval_args declares --judge-api-base-b without judge_a")
    if "judge_a" in required_judges and not judge_a:
        errors.append(f"benchmarks.manifests.{benchmark_id} requires judge_a but eval_args is missing --judge-model")
    if "judge_b" in required_judges and not judge_b:
        errors.append(f"benchmarks.manifests.{benchmark_id} requires judge_b but eval_args is missing --judge-model-b")
    if judge_b and not judge_a:
        errors.append(f"benchmarks.manifests.{benchmark_id}.eval_args cannot configure judge_b without judge_a")


def _read_tabular_header(path: Path, *, sheet_name: str | None = None) -> list[str]:
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        last_decode_error: UnicodeDecodeError | None = None
        for encoding in ("utf-8-sig", "cp1252"):
            try:
                with path.open("r", encoding=encoding, newline="") as handle:
                    return list(csv.DictReader(handle).fieldnames or [])
            except UnicodeDecodeError as exc:
                last_decode_error = exc
        if last_decode_error is not None:
            raise last_decode_error
        return []
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            selected_sheet_name = sheet_name or workbook.sheetnames[0]
            worksheet = workbook[selected_sheet_name]
            first_row = next(worksheet.iter_rows(values_only=True), ())
            return [str(value).strip() if value is not None else "" for value in first_row]
        finally:
            workbook.close()
    return []


def _read_tabular_shape(path: Path, *, sheet_name: str | None = None) -> tuple[list[str], int]:
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        last_decode_error: UnicodeDecodeError | None = None
        for encoding in ("utf-8-sig", "cp1252"):
            try:
                with path.open("r", encoding=encoding, newline="") as handle:
                    reader = csv.DictReader(handle)
                    header = list(reader.fieldnames or [])
                    return header, sum(1 for _row in reader)
            except UnicodeDecodeError as exc:
                last_decode_error = exc
        if last_decode_error is not None:
            raise last_decode_error
        return [], 0
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            selected_sheet_name = sheet_name or workbook.sheetnames[0]
            worksheet = workbook[selected_sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            first_row = next(rows, ())
            header = [str(value).strip() if value is not None else "" for value in first_row]
            return header, sum(1 for _row in rows)
        finally:
            workbook.close()
    return [], 0


def _schema_columns(path: Path) -> set[str]:
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        last_decode_error: UnicodeDecodeError | None = None
        for encoding in ("utf-8-sig", "cp1252"):
            try:
                with path.open("r", encoding=encoding, newline="") as handle:
                    reader = csv.DictReader(handle)
                    if reader.fieldnames and "column_name" in reader.fieldnames:
                        return {
                            str(row.get("column_name", "")).strip()
                            for row in reader
                            if str(row.get("column_name", "")).strip()
                        }
                    return {name for name in (reader.fieldnames or []) if name}
            except UnicodeDecodeError as exc:
                last_decode_error = exc
        if last_decode_error is not None:
            raise last_decode_error
        return set()
    if suffix == ".json":
        payload = read_json(path)
        columns = payload.get("columns", []) if isinstance(payload, dict) else []
        if isinstance(columns, dict):
            return {str(name).strip() for name in columns if str(name).strip()}
        if isinstance(columns, list):
            names = set()
            for column in columns:
                if isinstance(column, dict):
                    name = column.get("name") or column.get("column_name")
                    if name:
                        names.add(str(name).strip())
                elif column:
                    names.add(str(column).strip())
            return names
    return set()


def _validate_gold_contract(*, benchmark_id: str, manifest: Any, errors: list[str]) -> None:
    if not manifest.gold_path or not Path(manifest.gold_path).exists():
        return
    gold_path = Path(manifest.gold_path)
    try:
        gold_header = _read_tabular_header(gold_path, sheet_name=manifest.gold_sheet)
    except Exception as exc:
        errors.append(f"benchmarks.manifests.{benchmark_id}.gold_path could not be inspected: {exc}")
        return

    gold_header_set = {name for name in gold_header if name}
    for required in ("row_id", "row_index"):
        if required not in gold_header_set:
            errors.append(
                f"benchmarks.manifests.{benchmark_id}.gold_path is missing required stable join column: {required}"
            )

    schema_path = manifest.schema_path or manifest.eval_schema_path
    if not schema_path or not Path(schema_path).exists():
        return
    schema_columns = _schema_columns(Path(schema_path))
    if not schema_columns:
        return
    ignored = {"row_id", "row_index", "column_name", "gold_value", "cell_id"}
    gold_columns = {
        name
        for name in gold_header
        if name and name not in ignored and not name.endswith("__cell_id")
    }
    unknown = sorted(gold_columns - schema_columns)
    if unknown:
        errors.append(
            f"benchmarks.manifests.{benchmark_id}.gold_path has columns absent from schema: {', '.join(unknown)}"
        )


def _validate_table_gold_alignment(*, benchmark_id: str, manifest: Any, errors: list[str]) -> None:
    if not manifest.gold_path or not manifest.table_path:
        return
    gold_path = Path(manifest.gold_path)
    table_path = Path(manifest.table_path)
    if not gold_path.exists() or not table_path.exists():
        return
    try:
        _gold_header, gold_row_count = _read_tabular_shape(gold_path, sheet_name=manifest.gold_sheet)
        _table_header, table_row_count = _read_tabular_shape(table_path)
    except Exception as exc:
        errors.append(f"benchmarks.manifests.{benchmark_id} table/gold row alignment could not be inspected: {exc}")
        return
    if gold_row_count > 0 and table_row_count == 0:
        errors.append(
            f"benchmarks.manifests.{benchmark_id}.table_path has no paper rows while gold_path has "
            f"{gold_row_count}; eval-mode extraction needs stable template rows so app row_id values join to gold"
        )
    if gold_row_count > 0 and table_row_count not in (0, gold_row_count):
        errors.append(
            f"benchmarks.manifests.{benchmark_id}.table_path row count ({table_row_count}) does not match "
            f"gold_path row count ({gold_row_count}); optimizer eval joins require the same paper rows"
        )


def validate_preflight(
    config: dict[str, Any],
    benchmarks: Benchmarks,
    *,
    require_holdout: bool,
    experiment_dir: Path | None = None,
) -> None:
    errors: list[str] = []

    main_cfg = config["main_app"]
    eval_cfg = config["eval_app"]

    main_working_dir = Path(main_cfg["repo_root"]).resolve() if "repo_root" in main_cfg else None
    eval_working_dir = Path(eval_cfg["repo_root"]).resolve() if "repo_root" in eval_cfg else None

    if main_working_dir is not None and not main_working_dir.exists():
        errors.append(f"main_app.repo_root does not exist: {main_working_dir}")
    if eval_working_dir is not None and not eval_working_dir.exists():
        errors.append(f"eval_app.repo_root does not exist: {eval_working_dir}")

    if "base_config_path" in main_cfg and not Path(main_cfg["base_config_path"]).exists():
        errors.append(f"main_app.base_config_path does not exist: {main_cfg['base_config_path']}")

    for section_name, section, working_dir in [
        ("main_app", main_cfg, main_working_dir),
        ("eval_app", eval_cfg, eval_working_dir),
    ]:
        if "command" in section:
            _validate_command_tokens(section_name=section_name, command=list(section["command"]), errors=errors, working_dir=working_dir)
        if "command_prefix" in section:
            _validate_command_tokens(
                section_name=f"{section_name}.command_prefix",
                command=list(section["command_prefix"]),
                errors=errors,
                working_dir=working_dir,
            )

    for benchmark_id, manifest in benchmarks.manifests.items():
        for label, path_str in [
            (f"benchmarks.manifests.{benchmark_id}.table_path", manifest.table_path),
            (f"benchmarks.manifests.{benchmark_id}.pdf_dir", manifest.pdf_dir),
            (f"benchmarks.manifests.{benchmark_id}.gold_path", manifest.gold_path),
        ]:
            if not _existing_path(path_str):
                errors.append(f"{label} does not exist: {path_str}")
        if manifest.schema_path and not Path(manifest.schema_path).exists():
            errors.append(f"benchmarks.manifests.{benchmark_id}.schema_path does not exist: {manifest.schema_path}")
        if manifest.eval_schema_path and not Path(manifest.eval_schema_path).exists():
            errors.append(f"benchmarks.manifests.{benchmark_id}.eval_schema_path does not exist: {manifest.eval_schema_path}")
        if manifest.require_non_fixture_inputs:
            for label, path_str in [
                ("table_path", manifest.table_path),
                ("schema_path", manifest.schema_path),
                ("pdf_dir", manifest.pdf_dir),
                ("gold_path", manifest.gold_path),
            ]:
                if _looks_like_fixture_path(path_str):
                    errors.append(
                        f"benchmarks.manifests.{benchmark_id}.{label} points at fixture assets but the manifest requires real benchmark inputs: {path_str}"
                    )
        _validate_judge_contract(
            benchmark_id=benchmark_id,
            eval_args=list(manifest.eval_args),
            required_judges=list(manifest.required_judges or []),
            errors=errors,
        )
        _validate_gold_contract(benchmark_id=benchmark_id, manifest=manifest, errors=errors)
        _validate_table_gold_alignment(benchmark_id=benchmark_id, manifest=manifest, errors=errors)

    prompt_bundle_root = None
    if main_working_dir is not None:
        prompt_bundle_roots = _candidate_prompt_bundle_roots(main_working_dir)
        prompt_bundle_root = next((root for root in prompt_bundle_roots if root.exists()), None)
        if prompt_bundle_root is None:
            checked_paths = ", ".join(str(root) for root in prompt_bundle_roots)
            errors.append(f"Prompt bundle directory does not exist under main_app.repo_root; checked: {checked_paths}")

    if prompt_bundle_root is not None and prompt_bundle_root.exists():
        for prompt_id in _candidate_prompt_ids(config):
            bundle_dir = prompt_bundle_root / prompt_id
            if not bundle_dir.exists():
                errors.append(f"Prompt bundle does not exist under main_app.repo_root: {bundle_dir}")

    metric_groups = eval_cfg.get("metric_groups", {}) if isinstance(eval_cfg.get("metric_groups"), dict) else {}
    primary_metric = str(config["acceptance"]["primary_metric"])
    if metric_groups:
        if not _mapping_has_target(metric_groups.get("primary"), primary_metric):
            errors.append(
                f"eval_app.metric_groups.primary is missing the configured acceptance.primary_metric target: {primary_metric}"
            )

        guardrails = config.get("acceptance", {}).get("guardrails", {})
        guardrail_mapping = metric_groups.get("guardrail")
        for metric_name in guardrails:
            if metric_name == "runtime_seconds":
                continue
            if not _mapping_has_target(guardrail_mapping, metric_name):
                errors.append(f"eval_app.metric_groups.guardrail is missing configured guardrail target: {metric_name}")

    if experiment_dir is not None:
        results_jsonl = experiment_dir / "results" / "results.jsonl"
        if not results_jsonl.exists():
            errors.append(f"Experiment results were not found: {results_jsonl}")

    if errors:
        raise PreflightError("Preflight validation failed:\n- " + "\n- ".join(errors))


def validate_main_launch_contract(candidate: Candidate, launch: LaunchResult) -> list[str]:
    errors: list[str] = []
    payload = launch.payload if isinstance(launch.payload, dict) else {}

    if payload.get("schema_version") != "main_app_automation.v1":
        errors.append("main-app automation payload schema_version is missing or invalid")

    for key in ["run_id", "status", "mode"]:
        if not _is_non_empty_string(payload.get(key)):
            errors.append(f"main-app automation payload is missing required field: {key}")

    run_summary = payload.get("run_summary")
    if not isinstance(run_summary, dict):
        errors.append("main-app automation payload is missing run_summary")
    else:
        for key in ["prompt_hash", "prompt_bundle_id", "retrieval_mode", "provider_mode"]:
            if not _is_non_empty_string(run_summary.get(key)):
                errors.append(f"main-app automation payload run_summary is missing required field: {key}")
        if run_summary.get("prompt_bundle_id") != candidate.prompt_bundle_id:
            errors.append("main-app automation payload prompt_bundle_id does not match candidate prompt_bundle_id")

    artifacts = payload.get("artifacts")
    if not isinstance(artifacts, dict):
        errors.append("main-app automation payload is missing artifacts")
        artifacts = {}

    for key in ["run_dir", "run_json_path", "config_snapshot_path", "run_summary_path"]:
        path_str = artifacts.get(key)
        if not _is_non_empty_string(path_str):
            errors.append(f"main-app automation payload artifacts is missing required path: {key}")
            continue
        if not Path(path_str).exists():
            errors.append(f"main-app automation payload path does not exist for {key}: {path_str}")

    run_json_path = artifacts.get("run_json_path")
    if _is_non_empty_string(run_json_path) and Path(str(run_json_path)).exists():
        run_payload = read_json(Path(str(run_json_path)))
        for key in ["run_id", "run_mode", "provider_mode", "retrieval_mode", "prompt_bundle_id", "prompt_hash", "provider_text_model_id"]:
            if not _is_non_empty_string(run_payload.get(key)):
                errors.append(f"main-app run.json is missing required provenance field: {key}")
        if run_payload.get("prompt_bundle_id") != candidate.prompt_bundle_id:
            errors.append("main-app run.json prompt_bundle_id does not match candidate prompt_bundle_id")
        if run_payload.get("provider_text_model_id") != candidate.text_model_id:
            errors.append("main-app run.json provider_text_model_id does not match candidate text_model_id")
        if candidate.vision_model_id is not None and run_payload.get("provider_vision_model_id") != candidate.vision_model_id:
            errors.append("main-app run.json provider_vision_model_id does not match candidate vision_model_id")

    return errors


def validate_eval_summary_contract(config: dict[str, Any], launch: LaunchResult, eval_summary: dict[str, Any]) -> list[str]:
    errors: list[str] = []

    if not isinstance(eval_summary, dict):
        return ["eval summary is not a JSON object"]

    if not _is_non_empty_string(launch.summary_path):
        errors.append("eval launch did not record summary_path")
    elif not Path(str(launch.summary_path)).exists():
        errors.append(f"eval summary_path does not exist: {launch.summary_path}")

    for key in ["run_id", "run_dir"]:
        if not _is_non_empty_string(eval_summary.get(key)):
            errors.append(f"eval summary is missing required field: {key}")

    metrics = eval_summary.get("metrics")
    if not isinstance(metrics, dict):
        errors.append("eval summary is missing required metrics object")
        metrics = {}

    metadata = eval_summary.get("metadata")
    if not isinstance(metadata, dict):
        errors.append("eval summary is missing required metadata object")

    grouped_metrics_present = any(
        isinstance(eval_summary.get(group_name), dict)
        for group_name in ["primary_metrics", "guardrail_metrics", "diagnostic_metrics"]
    )

    primary_metric = str(config["acceptance"]["primary_metric"])
    if grouped_metrics_present:
        primary_group = eval_summary.get("primary_metrics", {}) if isinstance(eval_summary.get("primary_metrics"), dict) else {}
        if primary_metric not in primary_group:
            errors.append(f"eval summary primary_metrics is missing required metric: {primary_metric}")

        guardrail_group = eval_summary.get("guardrail_metrics", {}) if isinstance(eval_summary.get("guardrail_metrics"), dict) else {}
        for metric_name in config.get("acceptance", {}).get("guardrails", {}):
            if metric_name == "runtime_seconds":
                continue
            if metric_name not in guardrail_group:
                errors.append(f"eval summary guardrail_metrics is missing required metric: {metric_name}")
    else:
        metric_groups = config["eval_app"].get("metric_groups", {}) if isinstance(config["eval_app"].get("metric_groups"), dict) else {}
        for group_name in ["primary", "guardrail", "diagnostic"]:
            mapping = metric_groups.get(group_name)
            if isinstance(mapping, dict):
                for target_name, source_name in mapping.items():
                    if source_name not in metrics and source_name not in eval_summary:
                        errors.append(
                            f"eval summary is missing source metric '{source_name}' required for metric_groups.{group_name}.{target_name}"
                        )
            elif isinstance(mapping, list):
                for metric_name in mapping:
                    if metric_name not in metrics and metric_name not in eval_summary:
                        errors.append(
                            f"eval summary is missing source metric '{metric_name}' required for metric_groups.{group_name}"
                        )

    return errors
