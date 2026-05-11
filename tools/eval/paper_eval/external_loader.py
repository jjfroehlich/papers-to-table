from __future__ import annotations

import csv
from csv import Error as CsvError
from pathlib import Path
from typing import Any

from paper_eval.contracts import LoadedRun, ProposalRecord, RunMetadata, EvaluatorSchema
from paper_eval.errors import ContractError
from paper_eval.normalize import is_empty_value


_RESERVED_WIDE_COLUMNS = {"row_id", "row_index"}


def load_external_result(path: Path, *, run_id: str | None, schema: EvaluatorSchema) -> LoadedRun:
    if not path.exists():
        raise ContractError(f"External result input does not exist: {path}")
    if not path.is_file():
        raise ContractError(f"External result input is not a file: {path}")

    suffix = path.suffix.casefold()
    if suffix == ".csv":
        rows, fieldnames, sheet_name = _read_csv(path), None, None
        fieldnames = list(rows[0].keys()) if rows else _read_csv_header(path)
    elif suffix in {".xlsx", ".xlsm"}:
        rows, fieldnames, sheet_name = _read_xlsx(path)
    else:
        raise ContractError(f"Unsupported external result file type: {path.suffix}")

    proposals = _rows_to_proposals(
        rows,
        fieldnames,
        run_id=run_id or f"external_{path.stem}",
        schema=schema,
    )
    metadata = RunMetadata(
        run_id=run_id or f"external_{path.stem}",
        run_dir=path.parent,
        run_mode="external_result",
        extras={
            "external_result_path": str(path.resolve()),
            "external_result_sheet": sheet_name,
        },
    )
    return LoadedRun(
        run_dir=path.parent,
        metadata=metadata,
        proposals=proposals,
        contract_warnings=["external_result_table_has_no_run_bundle_evidence"],
    )


def _read_csv(path: Path) -> list[dict[str, Any]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    except CsvError as exc:
        raise ContractError(f"External result CSV could not be parsed at {path}: {exc}") from exc


def _read_csv_header(path: Path) -> list[str]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            return [str(value).strip() for value in next(reader, [])]
    except CsvError as exc:
        raise ContractError(f"External result CSV could not be parsed at {path}: {exc}") from exc


def _read_xlsx(path: Path) -> tuple[list[dict[str, Any]], list[str], str]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ContractError(
            "XLSX external results require openpyxl; install requirements.txt before scoring XLSX files."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = workbook.sheetnames[0]
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ContractError(f"External result workbook '{path}' is empty.")
        fieldnames = [str(value).strip() if value is not None else "" for value in rows[0]]
        return [dict(zip(fieldnames, values)) for values in rows[1:]], fieldnames, sheet_name
    finally:
        workbook.close()


def _rows_to_proposals(
    rows: list[dict[str, Any]],
    fieldnames: list[str],
    *,
    run_id: str,
    schema: EvaluatorSchema,
) -> list[ProposalRecord]:
    if not fieldnames:
        raise ContractError("External result input is empty or missing a header row.")
    fieldname_set = set(fieldnames)
    if {"row_id", "column_name", "proposed_value"}.issubset(fieldname_set):
        return _long_rows_to_proposals(rows, run_id=run_id, schema=schema)
    return _wide_rows_to_proposals(rows, fieldnames, run_id=run_id, schema=schema)


def _long_rows_to_proposals(
    rows: list[dict[str, Any]],
    *,
    run_id: str,
    schema: EvaluatorSchema,
) -> list[ProposalRecord]:
    proposals: list[ProposalRecord] = []
    for row in rows:
        row_id = _required_text(row.get("row_id"), "row_id")
        column_name = _required_text(row.get("column_name"), "column_name")
        value = row.get("proposed_value")
        if is_empty_value(value):
            continue
        proposals.append(_proposal_from_value(run_id, row_id, column_name, value, row.get("row_index"), schema))
    return proposals


def _wide_rows_to_proposals(
    rows: list[dict[str, Any]],
    fieldnames: list[str],
    *,
    run_id: str,
    schema: EvaluatorSchema,
) -> list[ProposalRecord]:
    if "row_id" not in fieldnames:
        raise ContractError("External wide-format result inputs must include a 'row_id' column.")
    data_columns = [
        name
        for name in fieldnames
        if name
        and name not in _RESERVED_WIDE_COLUMNS
        and not name.endswith("__cell_id")
        and schema.should_score_column(name)
    ]
    proposals: list[ProposalRecord] = []
    for row in rows:
        row_id = _required_text(row.get("row_id"), "row_id")
        row_index = row.get("row_index")
        for column_name in data_columns:
            value = row.get(column_name)
            if is_empty_value(value):
                continue
            proposals.append(_proposal_from_value(run_id, row_id, column_name, value, row_index, schema))
    return proposals


def _proposal_from_value(
    run_id: str,
    row_id: str,
    column_name: str,
    value: Any,
    row_index: Any,
    schema: EvaluatorSchema,
) -> ProposalRecord:
    column_schema = schema.column(column_name)
    return ProposalRecord(
        run_id=run_id,
        row_id=row_id,
        column_name=column_name,
        cell_id=f"external::{row_id}::{column_name}",
        proposed_value=value,
        state="external",
        field_type=column_schema.field_type if column_schema else None,
        allowed_values=list(column_schema.allowed_values) if column_schema else [],
        scoring_policy=column_schema.scoring_policy if column_schema else None,
        aliases=dict(column_schema.aliases) if column_schema else {},
        row_index=_optional_int(row_index),
        raw={"external_result": True},
    )


def _required_text(value: Any, field_name: str) -> str:
    if value is None or not str(value).strip():
        raise ContractError(f"External result input is missing required stable join field '{field_name}'.")
    return str(value).strip()


def _optional_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    return int(value)
