from __future__ import annotations

import csv
import hashlib
from csv import Error as CsvError
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from paper_eval.contracts import GoldCell, GoldDataset
from paper_eval.errors import ContractError
from paper_eval.normalize import is_empty_value


SYNTHESIZED_ROW_ID_WARNING = "gold_row_ids_synthesized_from_row_index_and_title"


@dataclass
class _GoldRowsResult:
    cells: list[GoldCell]
    contract_warnings: list[str] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)


def load_gold(
    path: Path,
    *,
    sheet_name: str | None = None,
    allowed_row_indices: set[int] | None = None,
    scored_columns: set[str] | None = None,
    excluded_columns: set[str] | None = None,
) -> GoldDataset:
    if not path.exists():
        raise ContractError(f"Gold input does not exist: {path}")
    if not path.is_file():
        raise ContractError(f"Gold input is not a file: {path}")
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        return _load_csv_gold(
            path,
            allowed_row_indices=allowed_row_indices,
            scored_columns=scored_columns,
            excluded_columns=excluded_columns,
        )
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx_gold(
            path,
            sheet_name=sheet_name,
            allowed_row_indices=allowed_row_indices,
            scored_columns=scored_columns,
            excluded_columns=excluded_columns,
        )
    raise ContractError(f"Unsupported gold file type: {path.suffix}")


def _load_csv_gold(
    path: Path,
    *,
    allowed_row_indices: set[int] | None,
    scored_columns: set[str] | None,
    excluded_columns: set[str] | None,
) -> GoldDataset:
    try:
        rows, fieldnames = _read_csv_dicts(path)
    except CsvError as exc:
        raise ContractError(f"Gold CSV could not be parsed at {path}: {exc}") from exc
    except UnicodeDecodeError as exc:
        raise ContractError(f"Gold CSV could not be decoded at {path}: {exc}") from exc
    if not fieldnames:
        raise ContractError(f"Gold CSV '{path}' is empty or missing a header row.")
    result = _rows_to_gold_cells(
        rows,
        fieldnames,
        sheet_name=None,
        scored_columns=scored_columns,
        excluded_columns=excluded_columns,
    )
    cells = _filter_gold_cells_by_row_index(result.cells, allowed_row_indices=allowed_row_indices)
    return GoldDataset(
        source_path=path,
        sheet_name=None,
        cells=cells,
        contract_warnings=result.contract_warnings,
        metadata=result.metadata,
    )


def _read_csv_dicts(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    last_decode_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle)
                return list(reader), list(reader.fieldnames or [])
        except UnicodeDecodeError as exc:
            last_decode_error = exc
    if last_decode_error is not None:
        raise last_decode_error
    return [], []


def _load_xlsx_gold(
    path: Path,
    *,
    sheet_name: str | None,
    allowed_row_indices: set[int] | None,
    scored_columns: set[str] | None,
    excluded_columns: set[str] | None,
) -> GoldDataset:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ContractError(
            "XLSX gold inputs require openpyxl; install requirements.txt before scoring XLSX files."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        selected_sheet_name = sheet_name or workbook.sheetnames[0]
        if selected_sheet_name not in workbook.sheetnames:
            raise ContractError(
                f"Worksheet '{selected_sheet_name}' was not found in {path.name}. "
                f"Available sheets: {', '.join(workbook.sheetnames)}"
            )

        worksheet = workbook[selected_sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ContractError(f"Gold worksheet '{selected_sheet_name}' is empty.")
        fieldnames = [str(value).strip() if value is not None else "" for value in rows[0]]
        data_rows = [dict(zip(fieldnames, values)) for values in rows[1:]]
        result = _rows_to_gold_cells(
            data_rows,
            fieldnames,
            sheet_name=selected_sheet_name,
            scored_columns=scored_columns,
            excluded_columns=excluded_columns,
        )
        cells = _filter_gold_cells_by_row_index(result.cells, allowed_row_indices=allowed_row_indices)
        return GoldDataset(
            source_path=path,
            sheet_name=selected_sheet_name,
            cells=cells,
            contract_warnings=result.contract_warnings,
            metadata=result.metadata,
        )
    finally:
        workbook.close()


def _rows_to_gold_cells(
    rows: list[dict[str, Any]],
    fieldnames: list[str],
    *,
    sheet_name: str | None,
    scored_columns: set[str] | None,
    excluded_columns: set[str] | None,
) -> _GoldRowsResult:
    fieldname_set = set(fieldnames)
    contract_warnings: list[str] = []
    metadata: dict[str, Any] = {}
    if {"row_id", "column_name", "gold_value"}.issubset(fieldname_set):
        cells = _load_long_form_rows(
            rows,
            sheet_name=sheet_name,
            scored_columns=scored_columns,
            excluded_columns=excluded_columns,
        )
    else:
        if "row_id" not in fieldnames:
            rows, fieldnames = _synthesize_wide_gold_join_fields(rows, fieldnames)
            contract_warnings.append(SYNTHESIZED_ROW_ID_WARNING)
            metadata["gold_row_ids_synthesized"] = True
            metadata["gold_row_id_algorithm"] = "sha256(row_index::Title)[:12]"
        cells = _load_wide_form_rows(
            rows,
            fieldnames,
            sheet_name=sheet_name,
            scored_columns=scored_columns,
            excluded_columns=excluded_columns,
        )
    _validate_unique_gold_join_keys(cells)
    return _GoldRowsResult(cells=cells, contract_warnings=contract_warnings, metadata=metadata)


def _load_long_form_rows(
    rows: list[dict[str, Any]],
    *,
    sheet_name: str | None,
    scored_columns: set[str] | None,
    excluded_columns: set[str] | None,
) -> list[GoldCell]:
    cells: list[GoldCell] = []
    for row in rows:
        row_id = _required_join_value(row.get("row_id"), "row_id")
        column_name = _required_join_value(row.get("column_name"), "column_name")
        if not _column_is_scored(column_name, scored_columns=scored_columns, excluded_columns=excluded_columns):
            continue
        raw_value = row.get("gold_value")
        cells.append(
            GoldCell(
                row_id=row_id,
                column_name=column_name,
                cell_id=_optional_text(row.get("cell_id")),
                row_index=_optional_int(row.get("row_index")),
                raw_value=raw_value,
                is_present=not is_empty_value(raw_value),
                sheet_name=sheet_name,
            )
        )
    return cells


def _load_wide_form_rows(
    rows: list[dict[str, Any]],
    fieldnames: list[str],
    *,
    sheet_name: str | None,
    scored_columns: set[str] | None,
    excluded_columns: set[str] | None,
) -> list[GoldCell]:
    if "row_id" not in fieldnames:
        raise ContractError(
            "Gold wide-format inputs must include a 'row_id' column to support stable joins."
        )
    reserved_columns = {"row_id", "row_index"}
    data_columns = [
        name
        for name in fieldnames
        if name
        and name not in reserved_columns
        and not name.endswith("__cell_id")
        and _column_is_scored(name, scored_columns=scored_columns, excluded_columns=excluded_columns)
    ]
    cells: list[GoldCell] = []
    for row in rows:
        row_id = _required_join_value(row.get("row_id"), "row_id")
        row_index = _optional_int(row.get("row_index"))
        for column_name in data_columns:
            raw_value = row.get(column_name)
            cells.append(
                GoldCell(
                    row_id=row_id,
                    column_name=column_name,
                    cell_id=_optional_text(row.get(f"{column_name}__cell_id")),
                    row_index=row_index,
                    raw_value=raw_value,
                    is_present=not is_empty_value(raw_value),
                    sheet_name=sheet_name,
                )
            )
    return cells


def _synthesize_wide_gold_join_fields(
    rows: list[dict[str, Any]],
    fieldnames: list[str],
) -> tuple[list[dict[str, Any]], list[str]]:
    synthesized_rows: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        row_index = _optional_int(row.get("row_index"))
        if row_index is None:
            row_index = index
        title = _optional_text(row.get("Title")) or ""
        synthesized = dict(row)
        synthesized["row_index"] = row_index
        synthesized["row_id"] = _generate_row_id(row_index=row_index, title=title)
        synthesized_rows.append(synthesized)

    updated_fieldnames = list(fieldnames)
    if "row_index" not in updated_fieldnames:
        updated_fieldnames.insert(0, "row_index")
    updated_fieldnames.insert(0, "row_id")
    return synthesized_rows, updated_fieldnames


def _generate_row_id(*, row_index: int, title: str) -> str:
    digest = hashlib.sha256(f"{row_index}::{title}".encode("utf-8")).hexdigest()[:12]
    return f"row_{digest}"


def _required_join_value(value: Any, field_name: str) -> str:
    text = _optional_text(value)
    if not text:
        raise ContractError(f"Gold input is missing required stable join field '{field_name}'.")
    return text


def _optional_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _optional_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    return int(value)


def _validate_unique_gold_join_keys(cells: list[GoldCell]) -> None:
    seen: dict[tuple[str, str], GoldCell] = {}
    for cell in cells:
        join_key = cell.join_key
        if join_key in seen:
            raise ContractError(
                "Gold input publishes duplicate stable join keys; each scored cell must have a unique "
                f"row_id + column_name pair. Duplicate: row_id='{cell.row_id}', column_name='{cell.column_name}'."
            )
        seen[join_key] = cell


def _column_is_scored(
    column_name: str,
    *,
    scored_columns: set[str] | None,
    excluded_columns: set[str] | None,
) -> bool:
    if scored_columns:
        return column_name in scored_columns
    if excluded_columns:
        return column_name not in excluded_columns
    return True


def _filter_gold_cells_by_row_index(
    cells: list[GoldCell],
    *,
    allowed_row_indices: set[int] | None,
) -> list[GoldCell]:
    if allowed_row_indices is None:
        return cells
    if not allowed_row_indices:
        return []

    missing_row_index = [cell for cell in cells if cell.row_index is None]
    if missing_row_index:
        raise ContractError(
            "Gold input cannot be restricted to matched run rows because one or more gold cells are missing "
            "the required row_index field."
        )

    return [cell for cell in cells if cell.row_index in allowed_row_indices]
