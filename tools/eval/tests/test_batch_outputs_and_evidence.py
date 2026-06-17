import csv
import json
import shutil
import tempfile
import unittest
from pathlib import Path

from paper_eval.cli import main
from paper_eval.evidence import validate_evidence_anchors
from paper_eval.run_loader import load_run

FIXTURE_ROOT = Path(__file__).resolve().parent / "fixtures" / "example_eval"


class EvidenceValidationTests(unittest.TestCase):
    def test_anchor_is_valid_when_quote_is_locatable(self) -> None:
        result = validate_evidence_anchors(
            [
                self._evidence_item(
                    page=1,
                    quote_text="result was positive",
                )
            ],
            page_text_by_page={1: "The result was positive in the trial cohort."},
            page_count=3,
        )

        self.assertEqual(result.outcome, "anchor_valid")
        self.assertTrue(result.anchor_valid)
        self.assertFalse(result.evidence_present_but_unvalidated)

    def test_evidence_present_but_unvalidated_when_text_is_unavailable(self) -> None:
        result = validate_evidence_anchors(
            [
                self._evidence_item(
                    page=1,
                    quote_text="result was positive",
                )
            ],
            page_text_by_page={},
            page_count=3,
        )

        self.assertEqual(result.outcome, "evidence_present_but_unvalidated")
        self.assertFalse(result.anchor_valid)
        self.assertTrue(result.evidence_present_but_unvalidated)

    def test_evidence_is_unvalidated_when_quote_cannot_be_located(self) -> None:
        result = validate_evidence_anchors(
            [
                self._evidence_item(
                    page=2,
                    quote_text="result was positive",
                )
            ],
            page_text_by_page={2: "A different sentence appears here."},
            page_count=3,
        )

        self.assertEqual(result.outcome, "evidence_present_but_unvalidated")
        self.assertFalse(result.anchor_valid)
        self.assertTrue(result.evidence_present_but_unvalidated)

    def test_anchor_is_invalid_when_required_anchor_fields_are_missing(self) -> None:
        result = validate_evidence_anchors(
            [
                self._evidence_item(
                    page=0,
                    quote_text="result was positive",
                )
            ],
            page_text_by_page={},
            page_count=3,
        )

        self.assertEqual(result.outcome, "anchor_invalid")
        self.assertFalse(result.anchor_valid)
        self.assertFalse(result.evidence_present_but_unvalidated)

    def _evidence_item(self, *, page: int, quote_text: str) -> object:
        from paper_eval.contracts import EvidenceItem

        return EvidenceItem(page=page, quote_text=quote_text)


class BatchEvaluationTests(unittest.TestCase):
    def test_run_loader_flattens_metadata_from_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            run_dir = self._create_run_bundle(
                base / "run-a",
                run_payload={
                    "run_id": "run-a",
                    "mode": "batch",
                    "provider": {"model_id": "text-model-a", "vision_model_id": "vision-a", "token": "provider-a"},
                    "parser": {"identity": "parser-a", "version": "1.2.3"},
                    "prompt": {"id": "prompt-17", "hash": "prompt-hash-a"},
                    "schema": {"version": "schema-v2"},
                    "style_profile_mode": "masked_rows",
                    "parser_cache_enabled": True,
                    "parse_cache_hit_count": 4,
                },
                config_payload={"config": {"hash": "cfg-a"}},
                include_page_text=True,
            )

            loaded_run = load_run(run_dir)
            metadata = loaded_run.metadata.flat_metadata()

            self.assertEqual(metadata["mode"], "batch")
            self.assertEqual(metadata["model_id"], "text-model-a")
            self.assertEqual(metadata["vision_model_id"], "vision-a")
            self.assertEqual(metadata["parser_identity"], "parser-a")
            self.assertEqual(metadata["parser_version"], "1.2.3")
            self.assertEqual(metadata["parser_identity_version"], "parser-a@1.2.3")
            self.assertEqual(metadata["style_profile_mode"], "masked_rows")
            self.assertTrue(metadata["parser_cache_enabled"])
            self.assertEqual(metadata["parse_cache_hit_count"], 4)
            self.assertEqual(metadata["prompt_identity"], "prompt-17")
            self.assertEqual(metadata["schema_identity"], "schema-v2")
            self.assertEqual(metadata["config_hash"], "cfg-a")
            self.assertEqual(metadata["run__provider__model_id"], "text-model-a")
            self.assertEqual(metadata["config_snapshot__config__hash"], "cfg-a")

    def test_evaluate_writes_batch_outputs_with_one_row_per_run(self) -> None:
        try:
            import pyarrow.parquet as pq
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:  # pragma: no cover
            self.skipTest(str(exc))

        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            run_a = self._create_run_bundle(
                base / "run-a",
                run_payload={
                    "run_id": "run-a",
                    "mode": "batch",
                    "provider": {"model_id": "text-model-a", "vision_model_id": "vision-a"},
                    "parser": {"identity": "parser-a", "version": "1.2.3"},
                    "prompt": {"id": "prompt-17"},
                    "schema": {"version": "schema-v2"},
                },
                config_payload={"config": {"hash": "cfg-a"}},
                include_page_text=True,
            )
            run_b = self._create_run_bundle(
                base / "run-b",
                run_payload={
                    "run_id": "run-b",
                    "run_mode": "batch",
                    "provider_text_model_id": "text-model-b",
                    "provider_vision_model_id": "vision-b",
                    "parser_identity": "parser-b",
                    "parser_version": "2.0.0",
                    "prompt_hash": "prompt-hash-b",
                    "schema_hash": "schema-hash-b",
                },
                config_payload={"config_hash": "cfg-b"},
                include_page_text=False,
            )
            gold_path = base / "gold.csv"
            gold_path.write_text("row_id,status\nrow-1,yes\n", encoding="utf-8")
            output_dir = base / "out"

            exit_code = main(
                [
                    "evaluate",
                    "--run",
                    str(run_a),
                    "--run",
                    str(run_b),
                    "--gold",
                    str(gold_path),
                    "--out",
                    str(output_dir),
                ]
            )

            self.assertEqual(exit_code, 0)
            compare_dir = output_dir / "compare"
            csv_path = compare_dir / "runs_comparison.csv"
            xlsx_path = compare_dir / "runs_comparison.xlsx"
            parquet_path = compare_dir / "runs_comparison.parquet"
            self.assertTrue(csv_path.exists())
            self.assertTrue(xlsx_path.exists())
            self.assertTrue(parquet_path.exists())

            csv_rows = self._read_csv(csv_path)
            self.assertEqual(len(csv_rows), 2)
            row_a = self._find_row(csv_rows, "run-a")
            row_b = self._find_row(csv_rows, "run-b")
            self.assertEqual(row_a["model_id"], "text-model-a")
            self.assertEqual(row_a["prompt_identity"], "prompt-17")
            self.assertEqual(row_a["anchor_valid_rate"], "1.0")
            self.assertEqual(row_a["correct_and_anchored_rate"], "1.0")
            self.assertEqual(row_b["model_id"], "text-model-b")
            self.assertEqual(row_b["prompt_identity"], "prompt-hash-b")
            self.assertEqual(row_b["anchor_valid_rate"], "0.0")
            self.assertEqual(row_b["evidence_present_but_unvalidated_count"], "1")

            workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
            try:
                worksheet = workbook["runs_comparison"]
                workbook_rows = list(worksheet.iter_rows(values_only=True))
                self.assertEqual(len(workbook_rows), 3)
                self.assertIn("run_id", workbook_rows[0])
            finally:
                workbook.close()

            parquet_rows = pq.read_table(parquet_path).to_pylist()
            self.assertEqual(len(parquet_rows), 2)
            self.assertEqual({row["run_id"] for row in parquet_rows}, {"run-a", "run-b"})

            summary_a = json.loads((output_dir / "per-run" / "run-a" / "run_summary.json").read_text(encoding="utf-8"))
            summary_b = json.loads((output_dir / "per-run" / "run-b" / "run_summary.json").read_text(encoding="utf-8"))
            self.assertEqual(summary_a["metrics"]["anchor_valid_rate"], 1.0)
            self.assertEqual(summary_b["metrics"]["anchor_valid_rate"], 0.0)
            self.assertEqual(summary_b["metrics"]["evidence_present_but_unvalidated_count"], 1)

    def test_batch_evaluation_writes_expected_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            copied_fixture = Path(temp_dir) / "fixture"
            shutil.copytree(FIXTURE_ROOT, copied_fixture)
            output_dir = Path(temp_dir) / "out"

            exit_code = main(
                [
                    "evaluate",
                    "--runs-root",
                    str(copied_fixture / "runs"),
                    "--gold",
                    str(copied_fixture / "gold.csv"),
                    "--schema",
                    str(copied_fixture / "schema.json"),
                    "--out",
                    str(output_dir),
                ]
            )

            self.assertEqual(exit_code, 0)
            csv_rows = self._read_csv(output_dir / "compare" / "runs_comparison.csv")
            self.assertEqual(len(csv_rows), 2)
            row_a = self._find_row(csv_rows, "run-a")
            row_b = self._find_row(csv_rows, "run-b")
            self.assertEqual(row_a["gold_table_hash"], "gold-hash-1")
            self.assertEqual(row_a["masked_table_snapshot_path"], "masked/masked_table.csv")
            self.assertEqual(row_a["structured_accuracy"], "1.0")
            self.assertEqual(row_b["structured_accuracy"], "0.5")
            self.assertEqual(row_b["text_accuracy"], "0.5")

    def test_join_key_failures_are_explicit_and_inspectable(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            run_dir = self._create_run_bundle(
                base / "run-a",
                run_payload={"run_id": "run-a", "provider_text_model_id": "model-a"},
                config_payload={"config_hash": "cfg-a"},
                include_page_text=True,
            )
            gold_path = base / "gold.csv"
            gold_path.write_text(
                "row_id,status,score,score__cell_id\n"
                "row-1,yes,10,cell-score-1\n"
                "row-2,no,20,cell-score-2\n",
                encoding="utf-8",
            )
            proposal_defaults = {
                "run_id": "run-a",
                "proposal_status": "value_proposed",
                "evidence_status": "direct_strong",
                "review_bucket": "review",
                "reason_codes": [],
            }
            proposal_rows = [
                {
                    "row_id": "row-1",
                    "column_name": "status",
                    "cell_id": "cell-status-1",
                    "proposed_value": "yes",
                    "field_type": "boolean",
                },
                {
                    "row_id": "row-1",
                    "column_name": "score",
                    "cell_id": "cell-score-1",
                    "proposed_value": "10",
                    "field_type": "numeric",
                },
                {
                    "row_id": "row-2",
                    "column_name": "score",
                    "cell_id": "wrong-cell-id",
                    "proposed_value": "20",
                    "field_type": "numeric",
                },
                {
                    "row_id": "row-2",
                    "column_name": "status",
                    "cell_id": "cell-status-2a",
                    "proposed_value": "no",
                    "field_type": "boolean",
                },
                {
                    "row_id": "row-2",
                    "column_name": "status",
                    "cell_id": "cell-status-2b",
                    "proposed_value": "no",
                    "field_type": "boolean",
                },
                {
                    "row_id": "row-9",
                    "column_name": "status",
                    "cell_id": "cell-status-9",
                    "proposed_value": "yes",
                    "field_type": "boolean",
                },
                {
                    "row_id": "row-1",
                    "column_name": "Title",
                    "cell_id": "cell-title-1",
                    "proposed_value": "Study title",
                    "field_type": "text",
                },
            ]
            with (run_dir / "proposals" / "proposals.jsonl").open("w", encoding="utf-8") as handle:
                for proposal_row in proposal_rows:
                    handle.write(json.dumps({**proposal_defaults, **proposal_row}) + "\n")

            output_dir = base / "out"
            self.assertEqual(
                main(
                    [
                        "evaluate",
                        "--run",
                        str(run_dir),
                        "--gold",
                        str(gold_path),
                        "--out",
                        str(output_dir),
                    ]
                ),
                0,
            )

            summary = json.loads((output_dir / "per-run" / "run-a" / "run_summary.json").read_text(encoding="utf-8"))
            self.assertEqual(summary["metrics"]["missing_proposal_count"], 0)
            self.assertEqual(summary["metrics"]["duplicate_proposal_join_count"], 1)
            self.assertEqual(summary["metrics"]["cell_id_mismatch_count"], 1)
            self.assertEqual(summary["metrics"]["unmatched_proposal_count"], 1)
            self.assertEqual(summary["metrics"]["excluded_proposal_count"], 1)
            self.assertEqual(summary["metrics"]["join_failure_count"], 3)
            self.assertIn("duplicate_proposals:row-2:status:None", summary["join_diagnostics"])
            self.assertIn("cell_id_mismatch:row-2:score:cell-score-2", summary["join_diagnostics"])
            self.assertIn("unmatched_proposal:row-9:status:cell-status-9", summary["join_diagnostics"])
            self.assertNotIn("excluded_proposal:row-1:Title:cell-title-1", summary["join_diagnostics"])
            self.assertIn(
                "excluded_proposal:row-1:Title:cell-title-1",
                summary["excluded_proposal_diagnostics"],
            )

            scored_records = [
                json.loads(line)
                for line in (output_dir / "per-run" / "run-a" / "scored_cells.jsonl").read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]
            excluded_record = next(record for record in scored_records if record["join_status"] == "excluded_proposal")
            self.assertEqual(excluded_record["column_name"], "Title")
            self.assertIn("proposal_for_excluded_column", excluded_record["diagnostic_flags"])

    def test_compare_command_rebuilds_comparison_outputs_from_per_run_summaries(self) -> None:
        try:
            import pyarrow.parquet as pq  # noqa: F401
            from openpyxl import load_workbook  # noqa: F401
        except ModuleNotFoundError as exc:  # pragma: no cover
            self.skipTest(str(exc))

        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            run_dir = self._create_run_bundle(
                base / "run-a",
                run_payload={"run_id": "run-a", "provider_text_model_id": "model-a"},
                config_payload={"config_hash": "cfg-a"},
                include_page_text=True,
            )
            gold_path = base / "gold.csv"
            gold_path.write_text("row_id,status\nrow-1,yes\n", encoding="utf-8")
            output_dir = base / "out"
            rebuilt_dir = base / "rebuilt-compare"

            self.assertEqual(
                main(
                    [
                        "evaluate",
                        "--run",
                        str(run_dir),
                        "--gold",
                        str(gold_path),
                        "--out",
                        str(output_dir),
                    ]
                ),
                0,
            )

            self.assertEqual(
                main(
                    [
                        "compare",
                        "--summaries",
                        str(output_dir / "per-run"),
                        "--out",
                        str(rebuilt_dir),
                    ]
                ),
                0,
            )

            rebuilt_rows = self._read_csv(rebuilt_dir / "runs_comparison.csv")
            self.assertEqual(len(rebuilt_rows), 1)
            self.assertEqual(rebuilt_rows[0]["run_id"], "run-a")

    def _create_run_bundle(
        self,
        run_dir: Path,
        *,
        run_payload: dict[str, object],
        config_payload: dict[str, object],
        include_page_text: bool,
    ) -> Path:
        (run_dir / "proposals").mkdir(parents=True)
        (run_dir / "inputs").mkdir(parents=True)
        (run_dir / "summaries").mkdir(parents=True)
        run_payload = {"artifact_schema_version": "main_run_bundle", **run_payload}
        (run_dir / "run.json").write_text(json.dumps(run_payload), encoding="utf-8")
        (run_dir / "config.snapshot.json").write_text(json.dumps(config_payload), encoding="utf-8")
        (run_dir / "inputs" / "input_summary.json").write_text(json.dumps({"page_count": 3}), encoding="utf-8")
        (run_dir / "summaries" / "run_summary.json").write_text(json.dumps({"status": "complete"}), encoding="utf-8")
        proposal_rows = [
            {
                "run_id": run_payload["run_id"],
                "row_id": "row-1",
                "column_name": "status",
                "cell_id": "cell-status-1",
                "proposed_value": "yes",
                "field_type": "boolean",
                "proposal_status": "value_proposed",
                "evidence_status": "direct_strong",
                "review_bucket": "review",
                "reason_codes": [],
                "evidence": [{"page": 1, "quote_text": "result was positive"}],
            }
        ]
        with (run_dir / "proposals" / "proposals.jsonl").open("w", encoding="utf-8") as handle:
            for row in proposal_rows:
                handle.write(json.dumps(row))
                handle.write("\n")
        if include_page_text:
            (run_dir / "evidence").mkdir(parents=True)
            (run_dir / "evidence" / "page_text.json").write_text(
                json.dumps({"1": "The result was positive in the trial cohort."}),
                encoding="utf-8",
            )
        return run_dir

    def _read_csv(self, path: Path) -> list[dict[str, str]]:
        with path.open("r", encoding="utf-8", newline="") as handle:
            return list(csv.DictReader(handle))

    def _find_row(self, rows: list[dict[str, str]], run_id: str) -> dict[str, str]:
        for row in rows:
            if row["run_id"] == run_id:
                return row
        raise AssertionError(f"Missing row for {run_id}")


if __name__ == "__main__":
    unittest.main()
