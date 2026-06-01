import csv
import io
import json
import shutil
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path
from unittest import mock

from paper_eval.cli import main
from paper_eval.contracts import JudgeResponse
from paper_eval.evidence import validate_evidence_anchors
from paper_eval.errors import CliUsageError, ContractError
from paper_eval.gold_loader import load_gold
from paper_eval.run_loader import discover_run_directories, load_run
from paper_eval.schema_loader import load_schema

FIXTURE_ROOT = Path(__file__).resolve().parent / "fixtures" / "example_eval"


class LoaderAndCliTests(unittest.TestCase):
    def test_run_loader_requires_required_artifact_files_with_explicit_message(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            run_dir = Path(temp_dir) / "run-a"
            run_dir.mkdir(parents=True)

            with self.assertRaisesRegex(ContractError, "missing required artifact files: run.json, proposals/proposals.jsonl"):
                load_run(run_dir)

    def test_run_loader_requires_stable_join_fields(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            run_dir = Path(temp_dir) / "run-a"
            (run_dir / "proposals").mkdir(parents=True)
            (run_dir / "run.json").write_text(
                json.dumps({"artifact_schema_version": "main_run_bundle", "run_id": "run-a"}),
                encoding="utf-8",
            )
            (run_dir / "proposals" / "proposals.jsonl").write_text(
                json.dumps({"run_id": "run-a", "column_name": "outcome", "cell_id": "cell-1"}) + "\n",
                encoding="utf-8",
            )

            with self.assertRaises(ContractError):
                load_run(run_dir)

    def test_run_loader_requires_eval_mode_provenance_fields(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            run_dir = self._create_run_bundle(Path(temp_dir) / "run-a")
            run_payload = json.loads((run_dir / "run.json").read_text(encoding="utf-8"))
            run_payload.pop("gold_table_hash")
            (run_dir / "run.json").write_text(json.dumps(run_payload), encoding="utf-8")

            with self.assertRaisesRegex(ContractError, "Missing fields for run 'run-a': gold_table_hash"):
                load_run(run_dir)

    def test_run_loader_requires_eval_mode_snapshot_paths_to_exist(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            run_dir = self._create_run_bundle(Path(temp_dir) / "run-a")
            run_payload = json.loads((run_dir / "run.json").read_text(encoding="utf-8"))
            run_payload["masked_table_snapshot_path"] = "masked/missing.csv"
            (run_dir / "run.json").write_text(json.dumps(run_payload), encoding="utf-8")

            with self.assertRaisesRegex(ContractError, "references missing provenance artifact 'masked_table_snapshot_path'"):
                load_run(run_dir)

    def test_run_loader_rejects_unsupported_artifact_schema_version(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            run_dir = self._create_run_bundle(Path(temp_dir) / "run-a")
            run_payload = json.loads((run_dir / "run.json").read_text(encoding="utf-8"))
            run_payload["artifact_schema_version"] = "main_run_bundle.v999"
            (run_dir / "run.json").write_text(json.dumps(run_payload), encoding="utf-8")

            with self.assertRaisesRegex(ContractError, "Unsupported run artifact version"):
                load_run(run_dir)

    def test_run_loader_rejects_missing_artifact_schema_version(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            run_dir = self._create_run_bundle(Path(temp_dir) / "run-a")
            run_payload = json.loads((run_dir / "run.json").read_text(encoding="utf-8"))
            run_payload.pop("artifact_schema_version")
            (run_dir / "run.json").write_text(json.dumps(run_payload), encoding="utf-8")

            with self.assertRaisesRegex(ContractError, "Unsupported run artifact version: <missing>"):
                load_run(run_dir)

    def test_run_loader_rejects_legacy_artifact_schema_version(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            run_dir = self._create_run_bundle(Path(temp_dir) / "run-a")
            run_payload = json.loads((run_dir / "run.json").read_text(encoding="utf-8"))
            run_payload["artifact_schema_version"] = "main_run_bundle" + ".v2"
            (run_dir / "run.json").write_text(json.dumps(run_payload), encoding="utf-8")

            with self.assertRaisesRegex(ContractError, "Unsupported run artifact version"):
                load_run(run_dir)

    def test_run_loader_accepts_main_app_eval_artifacts_shape(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            run_dir = Path(temp_dir) / "run-a"
            (run_dir / "proposals").mkdir(parents=True)
            (run_dir / "inputs").mkdir(parents=True)
            (run_dir / "summaries").mkdir(parents=True)
            (run_dir / "inputs" / "gold_table.xlsx").write_text("gold", encoding="utf-8")
            (run_dir / "inputs" / "masked_working_table.xlsx").write_text("masked", encoding="utf-8")
            run_payload = {
                "artifact_schema_version": "main_run_bundle",
                "run_id": "run-a",
                "run_mode": "eval",
                "provider_text_model_id": "text-model-1",
                "prompt_hash": "prompt-hash",
                "schema_hash": "schema-hash",
                "config_hash": "config-hash",
                "parser_identity": "docling",
                "eval_artifacts": {
                    "gold_table": {
                        "source_reference": "D:/tables/gold.xlsx",
                        "content_hash": "gold-hash",
                        "snapshot_path": "inputs/gold_table.xlsx",
                    },
                    "masked_working_table": {
                        "path": "inputs/masked_working_table.xlsx",
                        "content_hash": "masked-hash",
                    },
                },
            }
            (run_dir / "run.json").write_text(json.dumps(run_payload), encoding="utf-8")
            (run_dir / "proposals" / "proposals.jsonl").write_text(
                json.dumps(
                    {
                        "run_id": "run-a",
                        "row_id": "row-1",
                        "column_name": "notes",
                        "cell_id": "cell-1",
                        "proposed_value": "Some value",
                        "proposal_status": "value_proposed",
                        "evidence_status": "direct_strong",
                        "review_bucket": "review",
                        "reason_codes": [],
                        "field_type": "text",
                    }
                )
                + "\n",
                encoding="utf-8",
            )

            loaded = load_run(run_dir)

            self.assertEqual(loaded.metadata.run_mode, "eval")
            self.assertEqual(loaded.metadata.gold_source_ref, "D:/tables/gold.xlsx")
            self.assertEqual(loaded.metadata.gold_table_hash, "gold-hash")
            self.assertEqual(loaded.metadata.gold_table_snapshot_path, "inputs/gold_table.xlsx")
            self.assertEqual(loaded.metadata.masked_table_hash, "masked-hash")
            self.assertEqual(loaded.metadata.masked_table_snapshot_path, "inputs/masked_working_table.xlsx")

    def test_run_loader_loads_matched_row_indices_from_matching_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            run_dir = self._create_run_bundle(Path(temp_dir) / "run-a", matched_row_indices={1, 3})

            loaded = load_run(run_dir)

            self.assertEqual(loaded.matched_row_indices, {1, 3})

    def test_run_loader_loads_canonical_evidence_dir_and_parsed_page_text_fallback(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            run_dir = self._create_run_bundle(Path(temp_dir) / "run-a")
            (run_dir / "evidence").mkdir(parents=True)
            (run_dir / "parsed" / "pdf-1").mkdir(parents=True)
            (run_dir / "parsed" / "pdf-1" / "parsed_document.json").write_text(
                json.dumps(
                    {
                        "blocks": [
                            {
                                "page_number": 1,
                                "text": "Direct quote from page one.",
                                "normalized_text": "direct quote from page one.",
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )
            (run_dir / "proposals" / "proposals.jsonl").write_text(
                json.dumps(
                    {
                        "run_id": "run-a",
                        "row_id": "row-1",
                        "column_name": "notes",
                        "cell_id": "cell-1",
                        "pdf_id": "pdf-1",
                        "proposed_value": "Some value",
                        "proposal_status": "value_proposed",
                        "evidence_status": "direct_strong",
                        "review_bucket": "review",
                        "reason_codes": [],
                        "field_type": "text",
                        "evidence_ids": ["ev-1"],
                    }
                )
                + "\n",
                encoding="utf-8",
            )
            (run_dir / "evidence" / "ev-1.json").write_text(
                json.dumps(
                    {
                        "evidence_schema_version": "main_evidence",
                        "evidence_id": "ev-1",
                        "pdf_id": "pdf-1",
                        "page_number": 1,
                        "quote_text": "Direct quote from page one.",
                    }
                ),
                encoding="utf-8",
            )

            loaded = load_run(run_dir)

            evidence_item = loaded.proposals[0].evidence_items[0]
            self.assertIn("Direct quote from page one.", evidence_item.raw["source_text"])
            validation = validate_evidence_anchors([evidence_item], page_text_by_page=loaded.page_text_by_page)
            self.assertTrue(validation.anchor_valid)

    def test_run_loader_resolves_top_level_proposal_evidence_ids(self) -> None:
        temp_root = Path.cwd() / ".tmp_top_level_evidence_loader_test"
        shutil.rmtree(temp_root, ignore_errors=True)
        try:
            run_dir = self._create_run_bundle(temp_root / "run-a")
            (run_dir / "evidence").mkdir(parents=True)
            (run_dir / "proposals" / "proposals.jsonl").write_text(
                json.dumps(
                    {
                        "run_id": "run-a",
                        "row_id": "row-1",
                        "column_name": "notes",
                        "cell_id": "cell-1",
                        "pdf_id": "pdf-1",
                        "proposed_value": "Some value",
                        "proposal_status": "value_proposed",
                        "evidence_status": "direct_strong",
                        "review_bucket": "review",
                        "reason_codes": [],
                        "field_type": "text",
                        "primary_evidence_id": "ev-primary",
                        "evidence_ids": ["ev-primary", "ev-support"],
                        "ordered_supporting_evidence_ids": ["ev-support"],
                    }
                )
                + "\n",
                encoding="utf-8",
            )
            for evidence_id, quote in [
                ("ev-primary", "Primary quote."),
                ("ev-support", "Supporting quote."),
            ]:
                (run_dir / "evidence" / f"{evidence_id}.json").write_text(
                    json.dumps(
                        {
                            "evidence_schema_version": "main_evidence",
                            "evidence_id": evidence_id,
                            "pdf_id": "pdf-1",
                            "page_number": 1,
                            "quote_text": quote,
                        }
                    ),
                    encoding="utf-8",
                )

            loaded = load_run(run_dir)

            self.assertEqual([item.evidence_id for item in loaded.proposals[0].evidence_items], ["ev-primary", "ev-support"])
        finally:
            shutil.rmtree(temp_root, ignore_errors=True)

    def test_run_loader_normalizes_nullable_collection_fields(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            run_dir = self._create_run_bundle(Path(temp_dir) / "run-a")
            (run_dir / "proposals" / "proposals.jsonl").write_text(
                json.dumps(
                    {
                        "run_id": "run-a",
                        "row_id": "row-1",
                        "column_name": "notes",
                        "cell_id": "cell-1",
                        "proposed_value": None,
                        "proposal_status": "error",
                        "evidence_status": "not_applicable",
                        "review_bucket": "diagnostic",
                        "reason_codes": ["provider_error"],
                        "field_type": None,
                        "allowed_values": None,
                        "aliases": None,
                    }
                )
                + "\n",
                encoding="utf-8",
            )

            loaded = load_run(run_dir)

            self.assertEqual(loaded.proposals[0].allowed_values, [])
            self.assertEqual(loaded.proposals[0].aliases, {})

    def test_run_loader_invalid_json_has_explicit_message(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            run_dir = Path(temp_dir) / "run-a"
            (run_dir / "proposals").mkdir(parents=True)
            (run_dir / "run.json").write_text("{not json", encoding="utf-8")
            (run_dir / "proposals" / "proposals.jsonl").write_text("{}", encoding="utf-8")

            with self.assertRaisesRegex(ContractError, "Invalid JSON in .*run.json"):
                load_run(run_dir)

    def test_run_loader_invalid_proposals_jsonl_has_explicit_message(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            run_dir = Path(temp_dir) / "run-a"
            (run_dir / "proposals").mkdir(parents=True)
            (run_dir / "run.json").write_text(
                json.dumps({"artifact_schema_version": "main_run_bundle", "run_id": "run-a"}),
                encoding="utf-8",
            )
            (run_dir / "proposals" / "proposals.jsonl").write_text("{bad json\n", encoding="utf-8")

            with self.assertRaisesRegex(ContractError, "Invalid JSON in .*proposals.jsonl line 1"):
                load_run(run_dir)

    def test_discover_run_directories_requires_existing_inputs(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            missing_run = Path(temp_dir) / "missing-run"
            with self.assertRaisesRegex(CliUsageError, "Run path does not exist"):
                discover_run_directories([missing_run], None)

    def test_discover_runs_root_requires_directory(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            runs_root = Path(temp_dir) / "runs.txt"
            runs_root.write_text("not a directory", encoding="utf-8")

            with self.assertRaisesRegex(CliUsageError, "Runs root is not a directory"):
                discover_run_directories([], runs_root)

    def test_gold_loader_marks_present_and_empty_cells(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            gold_path = Path(temp_dir) / "gold.csv"
            gold_path.write_text(
                "row_id,status,notes\n"
                "row-1,yes,\n"
                "row-2,  ,NA\n",
                encoding="utf-8",
            )

            gold = load_gold(gold_path)
            indexed = {(cell.row_id, cell.column_name): cell for cell in gold.cells}

            self.assertTrue(indexed[("row-1", "status")].is_present)
            self.assertFalse(indexed[("row-1", "notes")].is_present)
            self.assertFalse(indexed[("row-2", "status")].is_present)
            self.assertTrue(indexed[("row-2", "notes")].is_present)

    def test_gold_loader_can_restrict_to_matched_row_indices(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            gold_path = Path(temp_dir) / "gold.csv"
            gold_path.write_text(
                "row_id,row_index,status,notes\n"
                "row-1,1,yes,alpha\n"
                "row-2,2,no,beta\n",
                encoding="utf-8",
            )

            gold = load_gold(gold_path, allowed_row_indices={2})

            self.assertEqual({cell.row_id for cell in gold.cells}, {"row-2"})

    def test_gold_loader_synthesizes_legacy_wide_row_ids_with_warning(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            gold_path = Path(temp_dir) / "gold.csv"
            gold_path.write_text(
                "Title,status,notes\n"
                "Paper A,yes,alpha\n"
                "Paper B,no,beta\n",
                encoding="utf-8",
            )

            gold = load_gold(gold_path)

            self.assertEqual({cell.row_index for cell in gold.cells}, {0, 1})
            self.assertEqual({cell.row_id for cell in gold.cells}, {"row_acab937a79e6", "row_8e11caa603ac"})
            self.assertIn("gold_row_ids_synthesized_from_row_index_and_title", gold.contract_warnings)
            self.assertTrue(gold.metadata["gold_row_ids_synthesized"])

    def test_gold_loader_rejects_duplicate_join_keys(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            gold_path = Path(temp_dir) / "gold.csv"
            gold_path.write_text(
                "row_id,column_name,gold_value\n"
                "row-1,status,yes\n"
                "row-1,status,no\n",
                encoding="utf-8",
            )

            with self.assertRaisesRegex(ContractError, "duplicate stable join keys"):
                load_gold(gold_path)

    def test_gold_loader_missing_file_has_explicit_message(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            gold_path = Path(temp_dir) / "missing.csv"

            with self.assertRaisesRegex(ContractError, "Gold input does not exist"):
                load_gold(gold_path)

    def test_gold_loader_empty_csv_has_explicit_message(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            gold_path = Path(temp_dir) / "gold.csv"
            gold_path.write_text("", encoding="utf-8")

            with self.assertRaisesRegex(ContractError, "empty or missing a header row"):
                load_gold(gold_path)

    def test_schema_loader_invalid_json_has_explicit_message(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            schema_path = Path(temp_dir) / "schema.json"
            schema_path.write_text("{bad json", encoding="utf-8")

            with self.assertRaisesRegex(ContractError, "Invalid JSON in schema file"):
                load_schema(schema_path)

    def test_schema_loader_rejects_invalid_columns_shape(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            schema_path = Path(temp_dir) / "schema.json"
            schema_path.write_text(json.dumps({"columns": "bad"}), encoding="utf-8")

            with self.assertRaisesRegex(ContractError, "Schema 'columns' must be either"):
                load_schema(schema_path)

    def test_schema_loader_canonicalizes_field_type_aliases(self) -> None:
        temp_root = Path.cwd() / ".tmp_schema_loader_aliases"
        shutil.rmtree(temp_root, ignore_errors=True)
        try:
            temp_root.mkdir(parents=True)
            schema_path = temp_root / "schema.json"
            schema_path.write_text(
                json.dumps(
                    {
                        "columns": {
                            "count": {"field_type": "number"},
                            "replicates": {"field_type": "integer"},
                            "score": {"field_type": "float"},
                            "status": {"field_type": "bool"},
                            "label": {"field_type": "enum"},
                            "group": {"field_type": "category"},
                            "summary": {"field_type": "string"},
                            "notes_snake": {"field_type": "free_text"},
                            "notes_kebab": {"field_type": "free-text"},
                        }
                    }
                ),
                encoding="utf-8",
            )

            schema = load_schema(schema_path)

            self.assertEqual(schema.column("count").field_type, "numeric")
            self.assertEqual(schema.column("replicates").field_type, "numeric")
            self.assertEqual(schema.column("score").field_type, "numeric")
            self.assertEqual(schema.column("status").field_type, "boolean")
            self.assertEqual(schema.column("label").field_type, "categorical")
            self.assertEqual(schema.column("group").field_type, "categorical")
            self.assertEqual(schema.column("summary").field_type, "text")
            self.assertEqual(schema.column("notes_snake").field_type, "text")
            self.assertEqual(schema.column("notes_kebab").field_type, "text")
        finally:
            shutil.rmtree(temp_root, ignore_errors=True)

    def test_schema_loader_rejects_unknown_field_type(self) -> None:
        temp_root = Path.cwd() / ".tmp_schema_loader_unknown_type"
        shutil.rmtree(temp_root, ignore_errors=True)
        try:
            temp_root.mkdir(parents=True)
            schema_path = temp_root / "schema.json"
            schema_path.write_text(json.dumps({"columns": {"value": {"field_type": "mystery"}}}), encoding="utf-8")

            with self.assertRaisesRegex(ContractError, "Unsupported field_type 'mystery'"):
                load_schema(schema_path)
        finally:
            shutil.rmtree(temp_root, ignore_errors=True)

    def test_schema_loader_excludes_metadata_columns_from_scoring_by_default(self) -> None:
        schema = load_schema(None)

        self.assertEqual(
            set(schema.excluded_columns),
            {"Title", "Authors", "Publication Year", "DOI", "Journal"},
        )
        for column_name in ["Title", "Authors", "Publication Year", "DOI", "Journal"]:
            self.assertFalse(schema.should_score_column(column_name))
        self.assertTrue(schema.should_score_column("Main analysis output"))

    def test_schema_loader_merges_default_metadata_exclusions_with_schema_exclusions(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            schema_path = Path(temp_dir) / "schema.json"
            schema_path.write_text(json.dumps({"excluded_columns": ["Internal Notes"]}), encoding="utf-8")

            schema = load_schema(schema_path)

        self.assertEqual(
            set(schema.excluded_columns),
            {"Title", "Authors", "Publication Year", "DOI", "Journal", "Internal Notes"},
        )

    def test_gold_loader_keeps_metadata_in_file_but_excludes_it_from_scored_cells(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            gold_path = Path(temp_dir) / "gold.csv"
            gold_path.write_text(
                "row_id,row_index,Title,Authors,Publication Year,DOI,Journal,Main analysis output\n"
                "row-1,0,Paper A,A. Author,2024,10.123/example,Nature,cell-type map\n",
                encoding="utf-8",
            )
            schema = load_schema(None)

            gold = load_gold(gold_path, excluded_columns=set(schema.excluded_columns))

        self.assertEqual([(cell.row_id, cell.column_name, cell.raw_value) for cell in gold.cells], [("row-1", "Main analysis output", "cell-type map")])

    def test_gold_xlsx_uses_first_sheet_by_default_and_supports_selection(self) -> None:
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError as exc:  # pragma: no cover
            self.skipTest(str(exc))

        with tempfile.TemporaryDirectory() as temp_dir:
            gold_path = Path(temp_dir) / "gold.xlsx"
            workbook = Workbook()
            first = workbook.active
            first.title = "First"
            first.append(["row_id", "status"])
            first.append(["row-1", "yes"])
            second = workbook.create_sheet("Second")
            second.append(["row_id", "status"])
            second.append(["row-2", "no"])
            workbook.save(gold_path)

            default_gold = load_gold(gold_path)
            explicit_gold = load_gold(gold_path, sheet_name="Second")

            self.assertEqual(default_gold.sheet_name, "First")
            self.assertEqual(default_gold.cells[0].row_id, "row-1")
            self.assertEqual(explicit_gold.sheet_name, "Second")
            self.assertEqual(explicit_gold.cells[0].row_id, "row-2")

    def test_gold_xlsx_invalid_sheet_has_explicit_message(self) -> None:
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError as exc:  # pragma: no cover
            self.skipTest(str(exc))

        with tempfile.TemporaryDirectory() as temp_dir:
            gold_path = Path(temp_dir) / "gold.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "First"
            worksheet.append(["row_id", "status"])
            worksheet.append(["row-1", "yes"])
            workbook.save(gold_path)

            with self.assertRaisesRegex(ContractError, "Worksheet 'Missing' was not found"):
                load_gold(gold_path, sheet_name="Missing")

    def test_cli_scores_single_run_and_outputs_expected_artifacts(self) -> None:
        try:
            from openpyxl import Workbook  # noqa: F401
        except ModuleNotFoundError as exc:  # pragma: no cover
            self.skipTest(str(exc))

        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            run_dir = self._create_run_bundle(base / "run-a")
            gold_path = base / "gold.csv"
            gold_path.write_text(
                "row_id,row_index,status,score,score__cell_id,notes\n"
                "row-1,1,yes,10,cell-score-1,\n"
                "row-2,2,,11,cell-score-2,Text gold\n"
                "row-3,3,no,20,cell-score-3,\n",
                encoding="utf-8",
            )
            schema_path = base / "schema.json"
            schema_path.write_text(
                json.dumps(
                    {
                        "global_numeric_tolerance": {"abs_tol": 0.1},
                        "columns": {
                            "status": {"field_type": "boolean"},
                            "score": {"field_type": "numeric", "numeric_tolerance": {"abs_tol": 0.5}},
                            "notes": {"field_type": "text", "scoring_policy": "judge"},
                        },
                    }
                ),
                encoding="utf-8",
            )
            output_dir = base / "out"

            class FakeJudge:
                def judge(self, judge_request) -> JudgeResponse:
                    self.last_request = judge_request
                    return JudgeResponse(verdict="correct", rationale_label="semantic_match")

            stdout = io.StringIO()
            with mock.patch("paper_eval.cli.build_text_judge", return_value=FakeJudge()):
                with redirect_stdout(stdout):
                    exit_code = main(
                        [
                            "evaluate",
                            "--run",
                            str(run_dir),
                            "--gold",
                            str(gold_path),
                            "--schema",
                            str(schema_path),
                            "--judge-model",
                            "fake-judge-v1",
                            "--out",
                            str(output_dir),
                        ]
                    )

            self.assertEqual(exit_code, 0)
            self.assertIn("Scored run run-a", stdout.getvalue())

            run_output = output_dir / "per-run" / "run-a"
            self.assertTrue((run_output / "scored_cells.jsonl").exists())
            self.assertTrue((run_output / "scored_cells.csv").exists())
            self.assertTrue((run_output / "run_summary.json").exists())
            self.assertTrue((run_output / "run_summary.csv").exists())
            self.assertTrue((run_output / "judge_records.jsonl").exists())

            summary = json.loads((run_output / "run_summary.json").read_text(encoding="utf-8"))
            self.assertEqual(summary["metrics"]["gold_present_cell_count"], 6)
            self.assertEqual(summary["metrics"]["gold_empty_cell_count"], 3)
            self.assertEqual(summary["metrics"]["filled_on_gold_empty_count"], 1)
            self.assertEqual(summary["metrics"]["missing_proposal_count"], 0)
            self.assertEqual(summary["metrics"]["cell_id_mismatch_count"], 1)
            self.assertEqual(summary["metrics"]["unmatched_proposal_count"], 1)
            self.assertEqual(summary["metrics"]["unscored_text_cell_count"], 0)
            self.assertEqual(summary["metrics"]["text_scored_cell_count"], 1)
            self.assertEqual(summary["metrics"]["judge_text_scored_cell_count"], 1)
            self.assertAlmostEqual(summary["metrics"]["text_accuracy"], 1.0)
            self.assertAlmostEqual(summary["metrics"]["structured_accuracy"], 1.0)
            self.assertEqual(summary["metrics"]["structured_deterministic_failure_count"], 0)
            self.assertEqual(summary["metrics"]["structured_adjudication_eligible_count"], 0)

            rows = self._read_csv(run_output / "scored_cells.csv")
            row_1_score = self._find_row(rows, row_id="row-1", column_name="score")
            row_2_status = self._find_row(rows, row_id="row-2", column_name="status")
            row_3_score = self._find_row(rows, row_id="row-3", column_name="score")
            note_row = self._find_row(rows, row_id="row-2", column_name="notes")

            self.assertEqual(row_1_score["was_scored"], "True")
            self.assertEqual(row_1_score["is_correct"], "True")
            self.assertEqual(row_1_score["deterministic_failure_kind"], "")
            self.assertEqual(row_1_score["adjudication_eligible"], "False")
            self.assertIn('"allowed_error": 0.5', row_1_score["diagnostics"])
            self.assertEqual(row_2_status["join_status"], "gold_empty_diagnostic")
            self.assertIn("filled_on_gold_empty", row_2_status["diagnostic_flags"])
            self.assertEqual(row_3_score["join_status"], "cell_id_mismatch")
            self.assertEqual(note_row["was_scored"], "True")
            self.assertEqual(note_row["judge_verdict"], "correct")
            self.assertEqual(note_row["judge_provider"], "lm_studio")
            self.assertEqual(note_row["judge_configured_model_id"], "fake-judge-v1")
            self.assertEqual(note_row["judge_resolved_model_id"], "")
            self.assertEqual(note_row["judge_model_id"], "fake-judge-v1")
            self.assertIn('"rationale_label": "semantic_match"', note_row["diagnostics"])

            judge_records = [
                json.loads(line)
                for line in (run_output / "judge_records.jsonl").read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]
            self.assertEqual(len(judge_records), 1)
            self.assertEqual(judge_records[0]["judge_provider"], "lm_studio")
            self.assertEqual(judge_records[0]["judge_configured_model_id"], "fake-judge-v1")
            self.assertIsNone(judge_records[0]["judge_resolved_model_id"])
            self.assertEqual(judge_records[0]["judge_model_id"], "fake-judge-v1")
            self.assertEqual(judge_records[0]["judge_verdict"], "correct")
            self.assertIsNotNone(judge_records[0]["judge_input_hash"])

    def test_cli_can_enable_text_exact_match_fast_path(self) -> None:
        try:
            from openpyxl import Workbook  # noqa: F401
        except ModuleNotFoundError as exc:  # pragma: no cover
            self.skipTest(str(exc))

        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            run_dir = self._create_run_bundle(base / "run-a")
            gold_path = base / "gold.csv"
            gold_path.write_text(
                "row_id,row_index,status,score,notes\n"
                "row-1,1,yes,10,\n"
                "row-2,2,,11,some   text\n"
                "row-3,3,no,20,\n",
                encoding="utf-8",
            )
            schema_path = base / "schema.json"
            schema_path.write_text(
                json.dumps(
                    {
                        "columns": {
                            "status": {"field_type": "boolean"},
                            "score": {"field_type": "numeric"},
                            "notes": {"field_type": "text", "scoring_policy": "judge"},
                        },
                    }
                ),
                encoding="utf-8",
            )
            output_dir = base / "out"

            class FailingJudge:
                def judge(self, judge_request) -> JudgeResponse:
                    raise AssertionError(f"judge should not be called for {judge_request.column_name}")

            with mock.patch("paper_eval.cli.build_text_judge", return_value=FailingJudge()):
                exit_code = main(
                    [
                        "evaluate",
                        "--run",
                        str(run_dir),
                        "--gold",
                        str(gold_path),
                        "--schema",
                        str(schema_path),
                        "--judge-model",
                        "fake-judge-v1",
                        "--enable-text-exact-match-fast-path",
                        "--out",
                        str(output_dir),
                    ]
                )

            self.assertEqual(exit_code, 0)
            run_output = output_dir / "per-run" / "run-a"
            self.assertFalse((run_output / "judge_records.jsonl").exists())

            summary = json.loads((run_output / "run_summary.json").read_text(encoding="utf-8"))
            self.assertEqual(summary["metrics"]["judge_text_scored_cell_count"], 0)
            self.assertEqual(summary["metrics"]["deterministic_text_scored_cell_count"], 1)

            rows = self._read_csv(run_output / "scored_cells.csv")
            note_row = self._find_row(rows, row_id="row-2", column_name="notes")
            self.assertEqual(note_row["scoring_policy"], "deterministic")
            self.assertEqual(note_row["is_correct"], "True")
            self.assertIn("text_exact_match_fast_path", note_row["diagnostic_flags"])

    def test_cli_restricts_gold_scoring_to_rows_with_matched_pdfs(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            run_dir = self._create_run_bundle(base / "run-a", matched_row_indices={1})
            gold_path = base / "gold.csv"
            gold_path.write_text(
                "row_id,row_index,status,score,notes\n"
                "row-1,1,yes,10,\n"
                "row-2,2,no,20,Text gold\n"
                "row-3,3,yes,30,Other text\n",
                encoding="utf-8",
            )
            output_dir = base / "out"

            self.assertEqual(
                main(
                    [
                        "evaluate",
                        "--run",
                        str(run_dir),
                        "--gold",
                        str(gold_path),
                        "--out",
                        str(output_dir),
                    ]
                ),
                0,
            )

            summary = json.loads((output_dir / "per-run" / "run-a" / "run_summary.json").read_text(encoding="utf-8"))
            self.assertEqual(summary["metrics"]["gold_present_cell_count"], 2)
            self.assertEqual(summary["metrics"]["missing_proposal_count"], 0)
            self.assertEqual(summary["metrics"]["unmatched_proposal_count"], 0)
            rows = self._read_csv(output_dir / "per-run" / "run-a" / "scored_cells.csv")
            self.assertEqual({row["row_id"] for row in rows}, {"row-1"})

    def test_cli_evaluates_fixture_run_with_expected_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            copied_fixture = Path(temp_dir) / "fixture"
            shutil.copytree(FIXTURE_ROOT, copied_fixture)
            output_dir = Path(temp_dir) / "out"

            exit_code = main(
                [
                    "evaluate",
                    "--run",
                    str(copied_fixture / "runs" / "run-a"),
                    "--gold",
                    str(copied_fixture / "gold.csv"),
                    "--schema",
                    str(copied_fixture / "schema.json"),
                    "--out",
                    str(output_dir),
                ]
            )

            self.assertEqual(exit_code, 0)
            run_output = output_dir / "per-run" / "run-a"
            self.assertTrue((run_output / "scored_cells.jsonl").exists())
            self.assertTrue((run_output / "run_summary.json").exists())
            self.assertTrue((output_dir / "compare" / "runs_comparison.csv").exists())

            summary = json.loads((run_output / "run_summary.json").read_text(encoding="utf-8"))
            self.assertEqual(summary["metrics"]["gold_present_cell_count"], 6)
            self.assertAlmostEqual(summary["metrics"]["structured_accuracy"], 1.0)
            self.assertAlmostEqual(summary["metrics"]["text_accuracy"], 1.0)
            self.assertAlmostEqual(summary["metrics"]["anchor_valid_rate"], 2 / 6)

    def test_cli_supports_runs_root_discovery(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            runs_root = base / "runs"
            self._create_run_bundle(runs_root / "run-a")
            gold_path = base / "gold.csv"
            gold_path.write_text("row_id,status\nrow-1,yes\n", encoding="utf-8")
            output_dir = base / "out"

            exit_code = main(
                [
                    "evaluate",
                    "--runs-root",
                    str(runs_root),
                    "--gold",
                    str(gold_path),
                    "--out",
                    str(output_dir),
                ]
            )

            self.assertEqual(exit_code, 0)
            self.assertTrue((output_dir / "per-run" / "run-a" / "run_summary.json").exists())

    def test_cli_scores_external_filled_table(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            gold_path = base / "gold.csv"
            gold_path.write_text(
                "row_id,status,score\nrow-1,yes,10\nrow-2,no,20\n",
                encoding="utf-8",
            )
            external_path = base / "external.csv"
            external_path.write_text(
                "row_id,status,score\nrow-1,yes,10.0\nrow-2,no,19\n",
                encoding="utf-8",
            )
            output_dir = base / "out"

            exit_code = main(
                [
                    "evaluate",
                    "--external-result",
                    str(external_path),
                    "--gold",
                    str(gold_path),
                    "--out",
                    str(output_dir),
                ]
            )

            self.assertEqual(exit_code, 0)
            run_output = output_dir / "per-run" / "external_external"
            self.assertTrue((run_output / "run_summary.json").exists())
            rows = self._read_csv(run_output / "scored_cells.csv")
            self.assertEqual(self._find_row(rows, row_id="row-1", column_name="status")["is_correct"], "True")
            self.assertEqual(self._find_row(rows, row_id="row-2", column_name="score")["is_correct"], "False")

    def test_cli_supports_repeated_run_arguments(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            run_a = self._create_run_bundle(base / "run-a")
            run_b = self._create_run_bundle(base / "run-b")
            gold_path = base / "gold.csv"
            gold_path.write_text("row_id,status\nrow-1,yes\n", encoding="utf-8")
            output_dir = base / "out"

            exit_code = main(
                [
                    "evaluate",
                    "--run",
                    str(run_a),
                    "--run",
                    str(run_b),
                    "--gold",
                    str(gold_path),
                    "--out",
                    str(output_dir),
                ]
            )

            self.assertEqual(exit_code, 0)
            self.assertTrue((output_dir / "per-run" / "run-a" / "run_summary.json").exists())
            self.assertTrue((output_dir / "per-run" / "run-b" / "run_summary.json").exists())

    def test_cli_evaluate_json_output_mode_emits_machine_readable_payload(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            run_dir = self._create_run_bundle(base / "run-a")
            gold_path = base / "gold.csv"
            gold_path.write_text("row_id,status\nrow-1,yes\n", encoding="utf-8")
            output_dir = base / "out"

            stdout = io.StringIO()
            with redirect_stdout(stdout):
                exit_code = main(
                    [
                        "evaluate",
                        "--run",
                        str(run_dir),
                        "--gold",
                        str(gold_path),
                        "--out",
                        str(output_dir),
                        "--json-output",
                    ]
                )

            self.assertEqual(exit_code, 0)
            payload = json.loads(stdout.getvalue().strip())
            self.assertEqual(payload["schema_version"], "paper_eval_cli.v1")
            self.assertEqual(payload["command"], "evaluate")
            self.assertTrue(payload["success"])
            self.assertEqual(payload["run_count"], 1)
            self.assertEqual(payload["run_ids"], ["run-a"])
            self.assertTrue((output_dir / "per-run" / "run-a" / "run_summary.json").exists())
            self.assertTrue((output_dir / "compare" / "runs_comparison.csv").exists())

    def test_cli_compare_json_output_mode_emits_machine_readable_payload(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base = Path(temp_dir)
            run_dir = self._create_run_bundle(base / "run-a")
            gold_path = base / "gold.csv"
            gold_path.write_text("row_id,status\nrow-1,yes\n", encoding="utf-8")
            output_dir = base / "out"

            self.assertEqual(
                main(
                    [
                        "evaluate",
                        "--run",
                        str(run_dir),
                        "--gold",
                        str(gold_path),
                        "--out",
                        str(output_dir),
                    ]
                ),
                0,
            )

            rebuilt_dir = base / "rebuilt"
            stdout = io.StringIO()
            with redirect_stdout(stdout):
                exit_code = main(
                    [
                        "compare",
                        "--summaries",
                        str(output_dir / "per-run"),
                        "--out",
                        str(rebuilt_dir),
                        "--json-output",
                    ]
                )

            self.assertEqual(exit_code, 0)
            payload = json.loads(stdout.getvalue().strip())
            self.assertEqual(payload["schema_version"], "paper_eval_cli.v1")
            self.assertEqual(payload["command"], "compare")
            self.assertTrue(payload["success"])
            self.assertGreaterEqual(payload["row_count"], 1)
            self.assertTrue((rebuilt_dir / "runs_comparison.csv").exists())
            self.assertTrue((rebuilt_dir / "runs_comparison.xlsx").exists())
            self.assertTrue((rebuilt_dir / "runs_comparison.parquet").exists())

    def test_compare_requires_existing_summary_input(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            missing = Path(temp_dir) / "missing"
            output_dir = Path(temp_dir) / "out"

            with self.assertRaises(SystemExit) as ctx:
                main(["compare", "--summaries", str(missing), "--out", str(output_dir)])

            self.assertEqual(ctx.exception.code, 2)

    def test_gold_loader_supports_long_form(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            gold_path = Path(temp_dir) / "gold.csv"
            gold_path.write_text(
                "row_id,column_name,cell_id,gold_value\nrow-1,status,cell-1,yes\n",
                encoding="utf-8",
            )
            gold = load_gold(gold_path)
            self.assertEqual(gold.cells[0].cell_id, "cell-1")
            self.assertEqual(gold.cells[0].column_name, "status")

    def _create_run_bundle(self, run_dir: Path, matched_row_indices: set[int] | None = None) -> Path:
        (run_dir / "proposals").mkdir(parents=True)
        (run_dir / "inputs").mkdir(parents=True)
        (run_dir / "summaries").mkdir(parents=True)
        (run_dir / "gold").mkdir(parents=True)
        (run_dir / "masked").mkdir(parents=True)
        (run_dir / "run.json").write_text(
            json.dumps(
                {
                    "artifact_schema_version": "main_run_bundle",
                    "run_id": run_dir.name,
                    "run_mode": "eval",
                    "provider_text_model_id": "model-1",
                    "gold_table_hash": "gold-hash-1",
                    "gold_table_snapshot_path": "gold/gold_snapshot.csv",
                    "masked_table_hash": "masked-hash-1",
                    "masked_table_snapshot_path": "masked/masked_table.csv",
                }
            ),
            encoding="utf-8",
        )
        (run_dir / "config.snapshot.json").write_text(json.dumps({"config_hash": "cfg-1"}), encoding="utf-8")
        (run_dir / "inputs" / "input_summary.json").write_text(json.dumps({"pdf_id": "pdf-1"}), encoding="utf-8")
        (run_dir / "summaries" / "run_summary.json").write_text(json.dumps({"status": "complete"}), encoding="utf-8")
        (run_dir / "gold" / "gold_snapshot.csv").write_text("row_id,status\nrow-1,\n", encoding="utf-8")
        (run_dir / "masked" / "masked_table.csv").write_text("row_id,status\nrow-1,\n", encoding="utf-8")
        proposal_rows = [
            {
                "run_id": run_dir.name,
                "row_id": "row-1",
                "column_name": "status",
                "cell_id": "cell-status-1",
                "proposed_value": "true",
                "field_type": "boolean",
            },
            {
                "run_id": run_dir.name,
                "row_id": "row-1",
                "column_name": "score",
                "cell_id": "cell-score-1",
                "proposed_value": "10.4",
                "field_type": "numeric",
            },
            {
                "run_id": run_dir.name,
                "row_id": "row-2",
                "column_name": "status",
                "cell_id": "cell-status-2",
                "proposed_value": "false",
                "field_type": "boolean",
            },
            {
                "run_id": run_dir.name,
                "row_id": "row-2",
                "column_name": "score",
                "cell_id": "cell-score-2",
                "proposed_value": "11.0",
                "field_type": "numeric",
            },
            {
                "run_id": run_dir.name,
                "row_id": "row-2",
                "column_name": "notes",
                "cell_id": "cell-notes-2",
                "proposed_value": "Some text",
                "field_type": "text",
            },
            {
                "run_id": run_dir.name,
                "row_id": "row-3",
                "column_name": "status",
                "cell_id": "cell-status-3",
                "proposed_value": "no",
                "field_type": "boolean",
            },
            {
                "run_id": run_dir.name,
                "row_id": "row-3",
                "column_name": "score",
                "cell_id": "proposal-cell-mismatch",
                "proposed_value": "20.0",
                "field_type": "numeric",
            },
            {
                "run_id": run_dir.name,
                "row_id": "row-9",
                "column_name": "status",
                "cell_id": "cell-status-9",
                "proposed_value": "yes",
                "field_type": "boolean",
            },
        ]
        with (run_dir / "proposals" / "proposals.jsonl").open("w", encoding="utf-8") as handle:
            for row in proposal_rows:
                row.setdefault("proposal_status", "value_proposed")
                row.setdefault("evidence_status", "direct_strong")
                row.setdefault("review_bucket", "review")
                row.setdefault("reason_codes", [])
                handle.write(json.dumps(row))
                handle.write("\n")
        if matched_row_indices is not None:
            (run_dir / "matching").mkdir(parents=True)
            match_results = [
                {
                    "pdf_id": f"pdf-{row_index}",
                    "pdf_path": f"papers/pdf-{row_index}.pdf",
                    "outcome": "matched",
                    "matched_row_index": row_index,
                    "matched_row_title": f"Title {row_index}",
                    "score": 1.0,
                    "runner_up_score": 0.0,
                    "runner_up_row_index": None,
                    "conflict_pdf_ids": [],
                    "conflict_row_indices": [],
                    "reasoning": "fixture match",
                    "blocked": False,
                    "blocked_reason": None,
                    "matched_at": "2026-01-01T00:00:00+00:00",
                }
                for row_index in sorted(matched_row_indices)
            ]
            (run_dir / "matching" / "match_results.json").write_text(
                json.dumps(match_results),
                encoding="utf-8",
            )
        return run_dir

    def _read_csv(self, path: Path) -> list[dict[str, str]]:
        with path.open("r", encoding="utf-8", newline="") as handle:
            return list(csv.DictReader(handle))

    def _find_row(self, rows: list[dict[str, str]], *, row_id: str, column_name: str) -> dict[str, str]:
        for row in rows:
            if row["row_id"] == row_id and row["column_name"] == column_name:
                return row
        raise AssertionError(f"Row not found for {row_id} / {column_name}")


if __name__ == "__main__":
    unittest.main()
