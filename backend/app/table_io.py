from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .ids import make_cell_id, make_row_id
from .models import CellEligibility, CellStatus, InputSummary, SchemaColumn

REQUIRED_METADATA_COLUMNS = ["Title", "Authors", "Publication Year"]


def load_table(path: str) -> tuple[list[dict[str, str]], list[str], str]:
    table_path = Path(path)
    suffix = table_path.suffix.lower()
    if suffix == ".csv":
        with table_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = [{k: (v or "") for k, v in row.items()} for row in reader]
            return rows, list(reader.fieldnames or []), "csv"
    if suffix == ".xlsx":
        wb = load_workbook(table_path)
        ws = wb.active
        headers = [cell.value or "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows: list[dict[str, str]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: "" if value is None else str(value) for i, value in enumerate(row) if i < len(headers)})
        return rows, headers, "xlsx"
    raise ValueError(f"Unsupported table format: {path}")


def load_schema(table_path: str, schema_path: str | None = None) -> list[SchemaColumn]:
    schema_rows: list[dict[str, Any]]
    if schema_path:
        schema_rows, _, _ = load_table(schema_path)
    else:
        table_rows, headers, file_type = load_table(table_path)
        if file_type != "xlsx":
            raise ValueError("A separate schema file is required when the table is CSV")
        wb = load_workbook(table_path)
        if "Schema" not in wb.sheetnames:
            raise ValueError("Workbook is missing a Schema sheet and no separate schema file was provided")
        ws = wb["Schema"]
        headers = [cell.value or "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        schema_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            schema_rows.append({headers[i]: "" if value is None else str(value) for i, value in enumerate(row) if i < len(headers)})
    if not schema_rows:
        raise ValueError("Schema file is empty")
    if "column_name" not in schema_rows[0] or "description" not in schema_rows[0]:
        raise ValueError("Schema must include column_name and description columns")
    columns = [SchemaColumn.model_validate(row) for row in schema_rows if row.get("column_name")]
    return columns


def validate_metadata_columns(headers: list[str]) -> None:
    missing = [name for name in REQUIRED_METADATA_COLUMNS if name not in headers]
    if missing:
        raise ValueError(f"Source table is missing required metadata columns: {', '.join(missing)}")


def classify_cells(
    rows: list[dict[str, str]],
    schema: list[SchemaColumn],
    placeholder_values: list[str],
    verify_mode: bool,
) -> tuple[list[dict[str, Any]], list[CellEligibility]]:
    normalized_rows: list[dict[str, Any]] = []
    eligibility: list[CellEligibility] = []
    placeholder_set = {value.strip().lower() for value in placeholder_values}

    for index, row in enumerate(rows):
        row_id = make_row_id(index, row.get("Title", f"row-{index}"))
        enriched = dict(row)
        enriched["row_id"] = row_id
        enriched["row_index"] = index
        normalized_rows.append(enriched)
        for column in schema:
            current_value = str(row.get(column.column_name, "") or "")
            stripped = current_value.strip()
            status = CellStatus.EMPTY
            eligible = True
            verify_target = False
            reason = ""
            normalized_placeholder = current_value.strip().lower()
            if current_value.lower() in placeholder_set or normalized_placeholder in placeholder_set:
                status = CellStatus.PLACEHOLDER if current_value != "" else CellStatus.EMPTY
                reason = "placeholder_treated_as_empty" if current_value != "" else "missing_value"
            elif stripped:
                status = CellStatus.FILLED
                if verify_mode:
                    verify_target = True
                    reason = "verify_mode_enabled"
                else:
                    eligible = False
                    reason = "filled_cell_verify_disabled"
            else:
                status = CellStatus.EMPTY
                reason = "missing_value"
            eligibility.append(
                CellEligibility(
                    row_id=row_id,
                    column_name=column.column_name,
                    cell_id=make_cell_id(row_id, column.column_name),
                    current_value=current_value,
                    status=status,
                    eligible=eligible or verify_target,
                    verify_target=verify_target,
                    reason=reason,
                )
            )
    return normalized_rows, eligibility


def build_input_summary(
    table_path: str,
    schema_path: str | None,
    pdf_dir: str,
    rows: list[dict[str, Any]],
    schema: list[SchemaColumn],
    verify_mode: bool,
) -> InputSummary:
    pdf_count = len(list(Path(pdf_dir).glob("*.pdf")))
    return InputSummary(
        table_path=table_path,
        schema_path=schema_path,
        pdf_dir=pdf_dir,
        row_count=len(rows),
        pdf_count=pdf_count,
        target_columns=[column.column_name for column in schema],
        verify_mode=verify_mode,
    )
