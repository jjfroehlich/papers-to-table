from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from ..artifacts import ArtifactStore
from ..models import ReviewDecisionType, WarningCategory


class ExportService:
    def __init__(self, store: ArtifactStore) -> None:
        self.store = store

    @staticmethod
    def _latest_decisions(decisions: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
        latest: dict[str, dict[str, Any]] = {}
        for row in decisions:
            proposal_id = row.get("proposal_id")
            decided_at = row.get("decided_at", "")
            prev = latest.get(proposal_id)
            if prev is None or decided_at >= prev.get("decided_at", ""):
                latest[proposal_id] = row
        return latest

    def _detect_unsupported_features(self, table_path: Path) -> list[str]:
        if table_path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            return []
        wb = load_workbook(table_path)
        warnings: list[str] = []
        defined_names = list(wb.defined_names.keys()) if hasattr(wb.defined_names, "keys") else list(wb.defined_names)
        if defined_names:
            warnings.append("named_ranges")
        if getattr(wb, "_pivots", None):
            warnings.append("pivot_tables")
        if wb.vba_archive is not None:
            warnings.append("macros")
        for ws in wb.worksheets:
            if ws.freeze_panes:
                warnings.append("frozen_panes")
            if ws.merged_cells.ranges:
                warnings.append("merged_cells")
            if ws.auto_filter and ws.auto_filter.ref:
                warnings.append("filters")
            if ws.conditional_formatting and len(ws.conditional_formatting) > 0:
                warnings.append("conditional_formatting")
            if ws.row_dimensions and any(dim.hidden for dim in ws.row_dimensions.values()):
                warnings.append("hidden_rows")
            if ws.column_dimensions and any(dim.hidden for dim in ws.column_dimensions.values()):
                warnings.append("hidden_columns")
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 50)):
                if any(cell.data_type == "f" for cell in row):
                    warnings.append("formulas")
                    break
        return sorted(set(warnings))

    def _load_source_workbook(self, table_path: Path) -> Workbook:
        if table_path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            return load_workbook(table_path)
        frame = pd.read_csv(table_path)
        wb = Workbook()
        ws = wb.active
        ws.append(list(frame.columns))
        for row in frame.itertuples(index=False, name=None):
            ws.append(list(row))
        return wb

    def build_exports(self, run_dir: Path, config: dict[str, Any]) -> dict[str, Any]:
        decisions = self.store.read_jsonl(run_dir / "review" / "decisions.jsonl")
        proposals = self.store.read_jsonl(run_dir / "proposals" / "proposals.jsonl")
        latest = self._latest_decisions(decisions)
        proposal_by_id = {item["proposal_id"]: item for item in proposals}

        table_path = Path(config["paths"]["table_path"])
        workbook = self._load_source_workbook(table_path)
        ws = workbook.active
        headers = [ws.cell(row=1, column=idx).value for idx in range(1, ws.max_column + 1)]
        col_idx = {str(name): idx + 1 for idx, name in enumerate(headers) if name is not None}

        highlight_color = config.get("export", {}).get("highlight_color", "FFF59D")
        fill = PatternFill(fill_type="solid", start_color=highlight_color, end_color=highlight_color)

        changed_cells = 0
        audit_rows: list[dict[str, Any]] = []
        for proposal_id, decision in latest.items():
            if decision.get("decision") not in {ReviewDecisionType.ACCEPT.value, ReviewDecisionType.ACCEPT_EDITED.value}:
                continue
            proposal = proposal_by_id.get(proposal_id)
            if proposal is None:
                continue
            row_id = str(proposal.get("row_id", ""))
            if not row_id.startswith("row_"):
                continue
            column_name = str(proposal.get("column_name", ""))
            if column_name not in col_idx:
                continue
            try:
                row_index = int(row_id.split("_", 1)[1])
            except ValueError:
                continue
            excel_row = row_index + 2
            excel_col = col_idx[column_name]
            cell = ws.cell(row=excel_row, column=excel_col)
            old_value = cell.value
            new_value = decision.get("edited_value") if decision.get("decision") == ReviewDecisionType.ACCEPT_EDITED.value else proposal.get("proposed_value")
            if new_value is None:
                continue
            cell.value = new_value
            cell.fill = fill
            changed_cells += 1
            audit_rows.append(
                {
                    "row_id": row_id,
                    "column_name": column_name,
                    "old_value": old_value,
                    "new_value": new_value,
                    "proposal_source": proposal.get("source_mode"),
                    "reviewer_decision": decision.get("decision"),
                    "decision_timestamp": decision.get("decided_at"),
                    "proposal_id": proposal_id,
                    "cell_id": proposal.get("cell_id"),
                }
            )

        export_path = run_dir / "exports" / "updated.xlsx"
        export_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(export_path)

        audit_path = run_dir / "exports" / "audit_log.jsonl"
        lines = "\n".join(json.dumps(item, ensure_ascii=False) for item in audit_rows)
        self.store.atomic_write(audit_path, f"{lines}\n" if lines else "")

        unsupported = self._detect_unsupported_features(table_path)
        return {
            "changed_cells_exported": changed_cells,
            "audit_entries": len(audit_rows),
            "unsupported_features": unsupported,
            "unsupported_warning": bool(unsupported),
        }

    def write_diagnostics(self, run_dir: Path, export_info: dict[str, Any]) -> dict[str, Any]:
        extraction = self.store.read_json(run_dir / "proposals" / "diagnostics.json") if (run_dir / "proposals" / "diagnostics.json").exists() else {"items": []}
        matching = self.store.read_json(run_dir / "matching" / "summary.json") if (run_dir / "matching" / "summary.json").exists() else {"results": []}
        status_index = self.store.read_json(run_dir / "review" / "status_index.json") if (run_dir / "review" / "status_index.json").exists() else {"run_warning_categories": []}

        matching_results = matching.get("results", [])
        extraction_diags = extraction.get("items", [])
        payload = {
            "matching_failures": {
                "unmatched": [m for m in matching_results if m.get("match_outcome") == "unmatched"],
                "ambiguous": [m for m in matching_results if m.get("match_outcome") == "ambiguous"],
                "duplicate_row_conflict": [m for m in matching_results if m.get("match_outcome") == "duplicate_row_conflict"],
            },
            "proposal_outcomes": {
                "blocked": [d for d in extraction_diags if d.get("status") == "blocked"],
                "unclear": [d for d in extraction_diags if d.get("status") == "unclear"],
                "skipped": [d for d in extraction_diags if d.get("status") == "skipped"],
                "error": [d for d in extraction_diags if d.get("status") == "error"],
            },
            "evidence_quality": {
                "weak_evidence_count": sum(1 for item in status_index.get("proposal_status", []) if WarningCategory.WEAK_EVIDENCE.value in item.get("warning_categories", [])),
                "quote_page_no_highlight_count": sum(1 for item in status_index.get("proposal_status", []) if WarningCategory.QUOTE_PAGE_NO_HIGHLIGHT.value in item.get("warning_categories", [])),
                "quote_page_fallback_count": sum(1 for d in extraction_diags if d.get("quote_page_fallback") is True),
                "figure_based_count": sum(1 for d in extraction_diags if d.get("figure_based") is True),
            },
            "unsupported_workbook_features": export_info.get("unsupported_features", []),
            "run_warning_categories": status_index.get("run_warning_categories", []),
            "completed_with_warnings": WarningCategory.COMPLETED_WITH_WARNINGS.value in status_index.get("run_warning_categories", []),
        }
        self.store.write_json(run_dir / "logs" / "diagnostics.json", payload)
        return payload
