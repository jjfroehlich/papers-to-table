"""Batch 6: Export pipeline — content-only XLSX export, audit log, diagnostics.

Tasks covered: T096 (XLSX export with changed-cell highlighting),
T097 (unsupported-feature warnings), T098 (audit log),
T099 (diagnostics JSON + completed-with-warnings).
"""
from __future__ import annotations

import pathlib
from datetime import datetime, timezone
from typing import Any, Optional

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .artifacts import read_json, write_json
from .extraction import load_proposals
from .ingest import load_table
from .matching import load_unmatched, load_ambiguous, load_conflicts
from .review import get_export_candidates, get_latest_decision
from .schemas import ProposalState, SupportLabel, WarningCategory

# Yellow fill for changed cells (T096)
HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# ---------------------------------------------------------------------------
# T097 — Unsupported-feature detection
# ---------------------------------------------------------------------------

_UNSUPPORTED_LABELS: dict[str, str] = {
    "formulas": "Cell formulas will not be preserved in the exported workbook",
    "conditional_formatting": "Conditional formatting will not be preserved in the exported workbook",
    "named_ranges": "Named ranges will not be preserved in the exported workbook",
    "charts": "Charts will not be preserved in the exported workbook",
    "auto_filter": "Auto-filter will not be preserved in the exported workbook",
    "frozen_panes": "Frozen panes will not be preserved in the exported workbook",
    "merged_cells": "Merged cells will not be preserved in the exported workbook",
    "cell_comments": "Cell comments will not be preserved in the exported workbook",
    "hidden_rows": "Hidden rows will not be preserved in the exported workbook",
    "hidden_columns": "Hidden columns will not be preserved in the exported workbook",
    "macros": "Macros are not supported and will be stripped from the exported workbook",
}


def detect_unsupported_features(wb: openpyxl.Workbook) -> list[str]:
    """Inspect an openpyxl workbook for features outside the MVP fidelity boundary.

    Returns a list of human-readable warning strings for each detected feature.
    The exported workbook is content-only — formulas, filters, frozen panes,
    hidden rows/columns, merged cells, conditional formatting, comments, named
    ranges, charts, shapes, and macros are all outside the MVP guarantee.
    """
    found: list[str] = []

    for ws in wb.worksheets:
        # Formulas
        if not _flag_found(found, "formulas"):
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        found.append(_UNSUPPORTED_LABELS["formulas"])
                        break
                if _flag_found(found, "formulas"):
                    break

        # Conditional formatting
        if not _flag_found(found, "conditional_formatting"):
            if ws.conditional_formatting:
                found.append(_UNSUPPORTED_LABELS["conditional_formatting"])

        # Auto-filter
        if not _flag_found(found, "auto_filter"):
            if ws.auto_filter and ws.auto_filter.ref:
                found.append(_UNSUPPORTED_LABELS["auto_filter"])

        # Frozen panes
        if not _flag_found(found, "frozen_panes"):
            if ws.freeze_panes:
                found.append(_UNSUPPORTED_LABELS["frozen_panes"])

        # Merged cells
        if not _flag_found(found, "merged_cells"):
            if ws.merged_cells.ranges:
                found.append(_UNSUPPORTED_LABELS["merged_cells"])

        # Comments
        if not _flag_found(found, "cell_comments"):
            if ws._comments:  # type: ignore[attr-defined]
                found.append(_UNSUPPORTED_LABELS["cell_comments"])

        # Hidden rows
        if not _flag_found(found, "hidden_rows"):
            for row_dim in ws.row_dimensions.values():
                if row_dim.hidden:
                    found.append(_UNSUPPORTED_LABELS["hidden_rows"])
                    break

        # Hidden columns
        if not _flag_found(found, "hidden_columns"):
            for col_dim in ws.column_dimensions.values():
                if col_dim.hidden:
                    found.append(_UNSUPPORTED_LABELS["hidden_columns"])
                    break

        # Charts
        if not _flag_found(found, "charts"):
            if getattr(ws, "_charts", None):
                found.append(_UNSUPPORTED_LABELS["charts"])

    # Named ranges (workbook-level)
    if not _flag_found(found, "named_ranges"):
        try:
            if wb.defined_names and list(wb.defined_names):
                found.append(_UNSUPPORTED_LABELS["named_ranges"])
        except Exception:
            pass

    # Macros (.xlsm vba)
    if not _flag_found(found, "macros"):
        if getattr(wb, "vba_archive", None):
            found.append(_UNSUPPORTED_LABELS["macros"])

    return found


def _flag_found(found: list[str], key: str) -> bool:
    """Return True if the label for 'key' is already in the found list."""
    label = _UNSUPPORTED_LABELS.get(key, "")
    return label in found


# ---------------------------------------------------------------------------
# T096 — Content-only XLSX export with changed-cell highlighting
# ---------------------------------------------------------------------------

def generate_xlsx_export(
    run_dir: pathlib.Path,
    candidates: list[dict],
    source_table_path: str,
) -> pathlib.Path:
    """Generate a content-only XLSX with only accepted changes applied.

    Rules:
    - Content-only: values are copied from the source workbook without
      formulas, formatting, filters, frozen panes, merged cells,
      conditional formatting, comments, named ranges, charts, or macros.
    - Accepted-only: only proposals with an explicit accepted or
      accepted_with_edit decision are written. Unreviewed, confirmed_no_data,
      and rejected are not written.
    - Changed cells are highlighted with a yellow fill.
    - The output is always an XLSX file even when the input was CSV.

    Returns the path to the written XLSX file.
    """
    # Build a map: (row_index, column_name) → export_value from candidates
    # To do this, we need to recover row_index from row_id.  We load the
    # source table, recompute row_id for each row, and build the mapping.
    from .ids import generate_row_id

    source_path = pathlib.Path(source_table_path)
    df = load_table(str(source_path))

    row_id_to_index: dict[str, int] = {}
    for i in range(len(df)):
        title = str(df.iloc[i].get("Title", "") if "Title" in df.columns else "")
        rid = generate_row_id(i, title)
        row_id_to_index[rid] = i

    # Map: (row_index, column_name) → export_value
    changes: dict[tuple[int, str], str] = {}
    for cand in candidates:
        row_id = cand.get("row_id", "")
        row_index = row_id_to_index.get(row_id)
        if row_index is None:
            continue
        col = cand.get("column_name", "")
        export_value = cand.get("export_value") or cand.get("proposed_value") or ""
        changes[(row_index, col)] = export_value

    # Read source as raw xlsx (or create from csv if needed)
    if source_path.suffix.lower() in (".xlsx", ".xls"):
        try:
            source_wb = openpyxl.load_workbook(str(source_path), data_only=True)
        except Exception:
            source_wb = None
    else:
        source_wb = None

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    assert out_ws is not None

    if source_wb is not None:
        src_ws = source_wb.active
        assert src_ws is not None
        # Copy all cell values (not formulas/formatting) from source
        for row in src_ws.iter_rows():
            for cell in row:
                val = cell.value
                # Strip formulas; export as empty string
                if isinstance(val, str) and val.startswith("="):
                    val = ""
                out_ws.cell(row=cell.row, column=cell.column, value=val)
    else:
        # Source is CSV — write header + rows from DataFrame
        header = list(df.columns)
        for col_idx, col_name in enumerate(header, start=1):
            out_ws.cell(row=1, column=col_idx, value=col_name)
        for row_idx, (_, row_series) in enumerate(df.iterrows(), start=2):
            for col_idx, col_name in enumerate(header, start=1):
                out_ws.cell(row=row_idx, column=col_idx, value=row_series[col_name])

    # Determine column positions from the header row (row 1)
    col_name_to_col_idx: dict[str, int] = {}
    for cell in out_ws[1]:  # type: ignore[index]
        if cell.value is not None:
            col_name_to_col_idx[str(cell.value)] = cell.column

    # Apply accepted changes and highlight
    for (row_index, col_name), export_value in changes.items():
        # row_index is 0-based; xlsx row 1 = header, row 2 = first data row
        xlsx_row = row_index + 2
        col_idx = col_name_to_col_idx.get(col_name)
        if col_idx is None:
            continue
        cell = out_ws.cell(row=xlsx_row, column=col_idx)
        cell.value = export_value
        cell.fill = HIGHLIGHT_FILL

    # Write to exports/
    exports_dir = run_dir / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_path = exports_dir / f"workbook_{ts}.xlsx"
    out_wb.save(str(out_path))
    return out_path


# ---------------------------------------------------------------------------
# T098 — Audit log generation
# ---------------------------------------------------------------------------

def generate_audit_log(
    run_dir: pathlib.Path,
    candidates: list[dict],
) -> pathlib.Path:
    """Generate a JSON audit log for all accepted changes.

    Each record includes: row_id, column_name, old_value, new_value,
    proposal_source, reviewer_decision, decision_timestamp,
    review_decision_id.
    """
    # Build proposal_id → proposal for old-value lookup
    proposals = load_proposals(run_dir)
    proposal_map = {p.proposal_id: p for p in proposals}

    records: list[dict] = []
    for cand in candidates:
        proposal_id = cand.get("proposal_id", "")
        proposal = proposal_map.get(proposal_id)
        old_value: Optional[str] = None
        if proposal and proposal.is_verify_mode:
            old_value = proposal.existing_value
        elif proposal:
            # In normal mode, old value is what was in the cell before (unknown at export time;
            # record None to be honest rather than guessing)
            old_value = None

        record = {
            "row_id": cand.get("row_id"),
            "column_name": cand.get("column_name"),
            "cell_id": cand.get("cell_id"),
            "old_value": old_value,
            "new_value": cand.get("export_value") or cand.get("proposed_value"),
            "proposal_source": {
                "proposal_id": proposal_id,
                "pdf_id": cand.get("pdf_id"),
            },
            "reviewer_decision": cand.get("decision"),
            "edited_value": cand.get("edited_value"),
            "review_decision_id": cand.get("review_decision_id"),
            "decision_timestamp": cand.get("decided_at"),
        }
        records.append(record)

    exports_dir = run_dir / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    audit_path = exports_dir / f"audit_log_{ts}.json"
    write_json(audit_path, {"generated_at": ts, "entries": records})
    return audit_path


# ---------------------------------------------------------------------------
# T099 — Diagnostics JSON
# ---------------------------------------------------------------------------

def generate_diagnostics(
    run_dir: pathlib.Path,
    run_data: dict,
    unsupported_feature_warnings: list[str],
) -> pathlib.Path:
    """Generate a diagnostics JSON file for the run.

    Collects:
    - run-level warnings from run.json
    - matching failures (unmatched, ambiguous, blocked)
    - proposal-level outcomes (unclear, skipped, error, weak evidence)
    - unsupported workbook feature warnings
    - completed_with_warnings summary
    """
    # Run-level warnings from run.json
    run_warnings = run_data.get("warnings", [])

    # Matching outcomes
    try:
        unmatched = load_unmatched(run_dir)
    except Exception:
        unmatched = []
    try:
        ambiguous = load_ambiguous(run_dir)
    except Exception:
        ambiguous = []
    try:
        conflicts = load_conflicts(run_dir)
    except Exception:
        conflicts = []

    # Proposal-level diagnostics
    proposals = load_proposals(run_dir)
    unclear_proposals: list[dict] = []
    skipped_proposals: list[dict] = []
    error_proposals: list[dict] = []
    weak_evidence_proposals: list[dict] = []
    blocked_proposals: list[dict] = []

    for p in proposals:
        if p.state == ProposalState.unclear:
            unclear_proposals.append({"proposal_id": p.proposal_id, "cell_id": p.cell_id, "column_name": p.column_name})
        elif p.state == ProposalState.skipped:
            skipped_proposals.append({"proposal_id": p.proposal_id, "cell_id": p.cell_id, "column_name": p.column_name})
        elif p.state == ProposalState.error:
            error_proposals.append({"proposal_id": p.proposal_id, "cell_id": p.cell_id, "column_name": p.column_name})
        elif p.state == ProposalState.blocked:
            blocked_proposals.append({"proposal_id": p.proposal_id, "cell_id": p.cell_id, "column_name": p.column_name})
        if p.support == SupportLabel.weak_evidence:
            weak_evidence_proposals.append({"proposal_id": p.proposal_id, "cell_id": p.cell_id, "column_name": p.column_name})

    status = run_data.get("status", "")
    completed_with_warnings = status == "completed_with_warnings"

    diag: dict[str, Any] = {
        "run_id": run_data.get("run_id"),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "run_status": status,
        "completed_with_warnings": completed_with_warnings,
        "run_warnings": run_warnings,
        "matching": {
            "unmatched_count": len(unmatched),
            "unmatched": unmatched,
            "ambiguous_count": len(ambiguous),
            "ambiguous": ambiguous,
            "duplicate_row_conflicts_count": len(conflicts),
            "duplicate_row_conflicts": conflicts,
        },
        "proposals": {
            "total": len(proposals),
            "blocked_count": len(blocked_proposals),
            "blocked": blocked_proposals,
            "unclear_count": len(unclear_proposals),
            "unclear": unclear_proposals,
            "skipped_count": len(skipped_proposals),
            "skipped": skipped_proposals,
            "error_count": len(error_proposals),
            "error": error_proposals,
            "weak_evidence_count": len(weak_evidence_proposals),
            "weak_evidence": weak_evidence_proposals,
        },
        "unsupported_workbook_features": {
            "warnings_count": len(unsupported_feature_warnings),
            "warnings": unsupported_feature_warnings,
            "fidelity_boundary": (
                "Export is content-only. Formulas, filters, frozen panes, "
                "hidden rows/columns, merged cells, conditional formatting, "
                "comments, named ranges, charts, shapes, and macros are not "
                "preserved in the exported workbook."
            ),
        },
    }

    exports_dir = run_dir / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    diag_path = exports_dir / f"diagnostics_{ts}.json"
    write_json(diag_path, diag)
    return diag_path


# ---------------------------------------------------------------------------
# run_export — orchestrate T096-T099 in one call
# ---------------------------------------------------------------------------

def run_export(
    run_dir: pathlib.Path,
    output_dir: str,
    run_id: str,
) -> dict:
    """Orchestrate the full export pipeline for a completed run.

    Steps:
    1. Read run.json and config snapshot.
    2. Resolve source workbook/table path.
    3. Detect unsupported workbook features (T097).
    4. Get accepted export candidates (T079/accepted-only).
    5. Generate XLSX export (T096).
    6. Generate audit log (T098).
    7. Generate diagnostics JSON (T099).
    8. Return a summary dict with paths and warnings.

    Raises ValueError if the run does not exist or is not in a completed state.
    """
    run_json_path = run_dir / "run.json"
    if not run_json_path.exists():
        raise ValueError(f"run.json not found for run {run_id}")

    run_data = read_json(run_json_path)
    status = run_data.get("status", "")
    if status not in ("completed", "completed_with_warnings"):
        raise ValueError(
            f"Run {run_id} is in status '{status}'; export is only available for "
            "completed or completed_with_warnings runs."
        )

    # Resolve source table path
    config_snap_path = run_dir / "config.snapshot.json"
    table_path: Optional[str] = None
    if config_snap_path.exists():
        config_snap = read_json(config_snap_path)
        table_path = config_snap.get("table_path")
    if not table_path:
        input_summary_path = run_dir / "inputs" / "input_summary.json"
        if input_summary_path.exists():
            input_summary = read_json(input_summary_path)
            table_path = input_summary.get("table_path")
    if not table_path:
        raise ValueError("Cannot resolve source table path for export.")

    # Detect unsupported features (best-effort on XLSX only)
    unsupported_warnings: list[str] = []
    source_path = pathlib.Path(table_path)
    if source_path.suffix.lower() in (".xlsx", ".xls") and source_path.exists():
        try:
            # Load without data_only so formulas are visible as strings
            wb_for_detection = openpyxl.load_workbook(str(source_path), data_only=False)
            unsupported_warnings = detect_unsupported_features(wb_for_detection)
        except Exception:
            pass

    # Get accepted-only export candidates
    candidates = get_export_candidates(run_dir)

    # Generate XLSX export (T096)
    xlsx_path = generate_xlsx_export(run_dir, candidates, table_path)

    # Generate audit log (T098)
    audit_path = generate_audit_log(run_dir, candidates)

    # Generate diagnostics (T099)
    diag_path = generate_diagnostics(run_dir, run_data, unsupported_warnings)

    result: dict = {
        "run_id": run_id,
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "accepted_changes_count": len(candidates),
        "workbook_path": str(xlsx_path),
        "audit_log_path": str(audit_path),
        "diagnostics_path": str(diag_path),
        "unsupported_feature_warnings": unsupported_warnings,
        "unsupported_feature_warnings_count": len(unsupported_warnings),
        "fidelity_boundary": (
            "Content-only export. Formulas, filters, frozen panes, hidden rows/columns, "
            "merged cells, conditional formatting, comments, named ranges, charts, shapes, "
            "and macros are not preserved."
        ),
    }
    return result
