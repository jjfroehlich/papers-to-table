from __future__ import annotations

import csv
from pathlib import Path
from typing import Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from .models import ProposalRecord, ReviewDecisionRecord, ReviewDecisionType

UNSUPPORTED_FEATURE_WARNINGS = [
    "Workbook export is content-only; formulas, charts, comments, merged cells, filters, hidden rows/columns, and conditional formatting are not preserved.",
]
DECISION_TIMESTAMP_NOT_RECORDED = "not_recorded"


def export_reviewed_changes(
    table_path: str,
    rows: list[dict],
    proposals: list[ProposalRecord],
    export_dir: Path,
    highlight_hex: str,
    decision_lookup: Mapping[str, ReviewDecisionRecord] | None = None,
) -> tuple[str, str, int, list[str]]:
    export_dir.mkdir(parents=True, exist_ok=True)
    accepted = [proposal for proposal in proposals if proposal.review_decision in {ReviewDecisionType.ACCEPT, ReviewDecisionType.ACCEPT_EDIT}]
    accepted_map = {proposal.cell_id: proposal.reviewed_value or proposal.proposed_value or "" for proposal in accepted}
    workbook_path = export_dir / "updated_workbook.xlsx"
    audit_path = export_dir / "audit_log.csv"
    fill = PatternFill(fill_type="solid", fgColor=highlight_hex)

    source_path = Path(table_path)
    if source_path.suffix.lower() == ".xlsx":
        source_wb = load_workbook(source_path)
        source_ws = source_wb.active
        headers = [cell.value or "" for cell in next(source_ws.iter_rows(min_row=1, max_row=1))]
        wb = Workbook()
        ws = wb.active
        ws.title = source_ws.title
        ws.append(headers)
        changed = 0
        row_lookup = {row["row_id"]: row for row in rows}
        for row in rows:
            values = []
            for col_index, header in enumerate(headers, start=1):
                cell_id = f"cell-{row['row_id']}-{header.strip().lower().replace(' ', '-')}"
                value = accepted_map.get(cell_id, row.get(header, ""))
                values.append(value)
            ws.append(values)
            excel_row = ws.max_row
            for col_index, header in enumerate(headers, start=1):
                cell_id = f"cell-{row['row_id']}-{header.strip().lower().replace(' ', '-')}"
                if cell_id in accepted_map:
                    ws.cell(excel_row, col_index).fill = fill
                    changed += 1
        wb.save(workbook_path)
    else:
        headers = list(rows[0].keys())
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        changed = 0
        for row in rows:
            values = []
            for col_index, header in enumerate(headers, start=1):
                cell_id = f"cell-{row['row_id']}-{header.strip().lower().replace(' ', '-')}"
                value = accepted_map.get(cell_id, row.get(header, ""))
                values.append(value)
            ws.append(values)
            for col_index, header in enumerate(headers, start=1):
                cell_id = f"cell-{row['row_id']}-{header.strip().lower().replace(' ', '-')}"
                if cell_id in accepted_map:
                    ws.cell(ws.max_row, col_index).fill = fill
                    changed += 1
        wb.save(workbook_path)
    with audit_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["row_identifier", "column_identifier", "old_value", "new_value", "proposal_source", "reviewer_decision", "decision_timestamp"])
        writer.writeheader()
        for proposal in accepted:
            decision = decision_lookup.get(proposal.proposal_id) if decision_lookup else None
            writer.writerow(
                {
                    "row_identifier": proposal.row_id,
                    "column_identifier": proposal.column_name,
                    "old_value": proposal.current_value,
                    "new_value": proposal.reviewed_value or proposal.proposed_value or "",
                    "proposal_source": proposal.source_mode,
                    "reviewer_decision": proposal.review_decision.value,
                    "decision_timestamp": decision.decided_at.isoformat() if decision else DECISION_TIMESTAMP_NOT_RECORDED,
                }
            )
    return workbook_path.name, audit_path.name, changed, UNSUPPORTED_FEATURE_WARNINGS
