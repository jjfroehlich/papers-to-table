from __future__ import annotations

import csv
import io
import json
import pathlib
import shutil
from typing import Optional

import openpyxl
import pandas as pd

TRIVIAL_PLACEHOLDERS = {"n/a", "na", "tbd", "tba", "unknown", "-", "--", "none", "?"}
REQUIRED_METADATA_COLS = {"Title", "Authors", "Publication Year"}


def load_table(path: str) -> pd.DataFrame:
    """Load table from CSV (BOM-safe) or XLSX, normalize missing values to empty string."""
    p = pathlib.Path(path)
    if p.suffix.lower() in (".xlsx", ".xls"):
        df = pd.read_excel(path, dtype=str)
        for col in df.columns:
            df[col] = df[col].fillna("")
        return df
    else:
        with open(path, "r", encoding="utf-8-sig") as f:
            content = f.read()
        df = pd.read_csv(io.StringIO(content), dtype=str)
        df = df.fillna("")
        return df


def _normalize_allowed_values(value: object) -> Optional[list[str]]:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none"}:
        return None
    try:
        parsed = json.loads(text)
        if isinstance(parsed, list):
            values = [str(item).strip() for item in parsed if str(item).strip()]
            return values or None
    except Exception:
        pass

    separator = "|" if "|" in text else ";" if ";" in text else ","
    values = [item.strip() for item in text.split(separator) if item.strip()]
    return values or None


def _schema_row_to_dict(row: dict) -> dict:
    field_type = str(row.get("field_type", "") or "").strip() or None
    return {
        "column_name": str(row.get("column_name", "") or "").strip(),
        "description": str(row.get("description", "") or "").strip(),
        "field_type": field_type,
        "allowed_values": _normalize_allowed_values(row.get("allowed_values")),
    }


def load_schema(schema_path: Optional[str], table_path: str) -> list[dict]:
    """Load schema from CSV or embedded XLSX sheet."""
    if schema_path:
        p = pathlib.Path(schema_path)
        if p.suffix.lower() == ".csv":
            with open(schema_path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                return [_schema_row_to_dict(row) for row in reader]
    tp = pathlib.Path(table_path)
    if tp.suffix.lower() in (".xlsx", ".xls"):
        try:
            df = pd.read_excel(table_path, sheet_name="Schema", dtype=str)
            if "column_name" in df.columns:
                return [_schema_row_to_dict(r.to_dict()) for _, r in df.iterrows()]
        except Exception:
            pass
    return []


def validate_metadata_columns(df: pd.DataFrame) -> list[str]:
    """Validate that Title, Authors, Publication Year exist. Returns error list."""
    errors = []
    for col in REQUIRED_METADATA_COLS:
        if col not in df.columns:
            errors.append(f"Required metadata column missing: '{col}'")
    return errors


def validate_schema_columns(schema: list[dict]) -> list[str]:
    """Validate schema has column_name and description fields."""
    errors = []
    allowed_field_types = {"text", "number", "categorical", "boolean"}
    for i, col in enumerate(schema):
        if not col.get("column_name"):
            errors.append(f"Schema row {i}: missing column_name")
        if not col.get("description"):
            errors.append(
                f"Schema row {i}: missing description for column '{col.get('column_name', '')}'"
            )
        field_type = col.get("field_type")
        field_type_value = getattr(field_type, "value", field_type)
        if field_type and field_type not in allowed_field_types:
            errors.append(
                f"Schema row {i}: invalid field_type '{field_type}' for column "
                f"'{col.get('column_name', '')}'"
            )
        allowed_values = col.get("allowed_values")
        if allowed_values and field_type_value != "categorical":
            errors.append(
                f"Schema row {i}: allowed_values require field_type='categorical' for column "
                f"'{col.get('column_name', '')}'"
            )
    return errors


def is_trivial_placeholder(value: str) -> bool:
    return value.strip().lower() in TRIVIAL_PLACEHOLDERS


def classify_cell_eligibility(
    value: str,
    verify_mode: bool = False,
    eval_mode: bool = False,
) -> str:
    """Returns 'eligible', 'already_filled', 'placeholder', or 'ineligible'."""
    if not value or value.strip() == "":
        return "eligible"
    if is_trivial_placeholder(value):
        return "placeholder"
    if verify_mode or eval_mode:
        return "eligible"
    return "already_filled"


def get_eligible_cells(
    df: pd.DataFrame,
    schema: list[dict],
    verify_mode: bool = False,
    eval_mode: bool = False,
) -> list[dict]:
    """Return list of eligible cells: {row_id, row_index, column_name, current_value, eligibility}."""
    from .ids import generate_row_id

    target_cols = set(get_target_columns(df, schema))

    eligible = []
    for row_idx, row in df.iterrows():
        title = str(row.get("Title", ""))
        row_id = generate_row_id(int(row_idx), title)
        for col_name in target_cols:
            if col_name not in df.columns:
                continue
            value = str(row.get(col_name, ""))
            eligibility = classify_cell_eligibility(value, verify_mode, eval_mode)
            if eligibility in ("eligible", "placeholder") or (
                (verify_mode or eval_mode) and eligibility == "already_filled"
            ):
                eligible.append(
                    {
                        "row_id": row_id,
                        "row_index": int(row_idx),
                        "column_name": col_name,
                        "current_value": value,
                        "eligibility": eligibility,
                    }
                )
    return eligible


def get_target_columns(df: pd.DataFrame, schema: list[dict]) -> list[str]:
    schema_cols = {col["column_name"] for col in schema}
    return [column for column in df.columns if column in schema_cols and column not in REQUIRED_METADATA_COLS]


def create_masked_working_dataframe(
    df: pd.DataFrame,
    schema: list[dict],
) -> tuple[pd.DataFrame, dict]:
    masked_df = df.copy(deep=True)
    target_columns = get_target_columns(masked_df, schema)
    masked_non_empty_count = 0

    for column_name in target_columns:
        column_values = masked_df[column_name].fillna("")
        masked_non_empty_count += sum(1 for value in column_values if str(value).strip())
        masked_df[column_name] = ""

    return masked_df, {
        "target_columns": target_columns,
        "target_cell_count": len(masked_df.index) * len(target_columns),
        "masked_non_empty_cell_count": masked_non_empty_count,
    }


def persist_table_snapshot(source_path: str, destination_path: str) -> None:
    destination = pathlib.Path(destination_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, destination)


def persist_masked_working_copy(
    source_path: str,
    destination_path: str,
    schema: list[dict],
    masked_df: pd.DataFrame,
) -> None:
    source = pathlib.Path(source_path)
    destination = pathlib.Path(destination_path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    if source.suffix.lower() in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        workbook = openpyxl.load_workbook(source)
        worksheet = workbook.worksheets[0]
        header_map = {
            str(cell.value).strip(): index
            for index, cell in enumerate(worksheet[1], start=1)
            if cell.value is not None
        }
        for column_name in get_target_columns(masked_df, schema):
            column_index = header_map.get(column_name)
            if column_index is None:
                continue
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).value = ""
        workbook.save(destination)
        return

    if source.suffix.lower() == ".xls":
        masked_df.to_excel(destination, index=False)
        return

    masked_df.to_csv(destination, index=False)
