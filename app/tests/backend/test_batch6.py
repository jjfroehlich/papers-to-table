"""Batch 6 tests: export integrity, content-only fidelity, changed-cell
highlighting, accepted-only export, unsupported-feature warnings,
audit-log completeness, diagnostics, and completed-with-warnings semantics.

Tasks covered: T096, T097, T098, T099, T101.
"""
from __future__ import annotations

import io
import pathlib
from datetime import datetime, timezone

import openpyxl
import pytest
from fastapi.testclient import TestClient

from backend.app.artifacts import (
    get_run_dir,
    init_run_bundle,
    read_json,
    write_json,
)
from backend.app.export import (
    detect_unsupported_features,
    generate_audit_log,
    generate_diagnostics,
    generate_xlsx_export,
    run_export,
)
from backend.app.extraction import (
    EvidenceRecord,
    ProposalRecord,
    persist_evidence,
    persist_proposal,
)
from backend.app.ids import (
    generate_cell_id,
    generate_evidence_id,
    generate_proposal_id,
    generate_row_id,
    generate_run_id,
)
from backend.app.main import app
from backend.app.review import (
    get_export_candidates,
    record_review_decision,
)
from backend.app.schemas import (
    EvidenceSourceType,
    ProposalState,
    ReviewDecision,
    ReviewResolutionReason,
    RunStatus,
    SupportLabel,
    WarningCategory,
)

client = TestClient(app)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_run(tmp_path: pathlib.Path, status: str = RunStatus.completed.value) -> tuple[pathlib.Path, str]:
    run_id = generate_run_id()
    output_dir = str(tmp_path)
    run_dir = init_run_bundle(output_dir, run_id)
    run_data = {
        "run_id": run_id,
        "status": status,
        "output_dir": output_dir,
        "verify_mode": False,
        "total_rows": 2,
        "eligible_cells": 2,
        "proposals_generated": 0,
        "proposals_reviewed": 0,
        "warnings": [],
    }
    write_json(run_dir / "run.json", run_data)
    return run_dir, run_id


def _make_xlsx_table(tmp_path: pathlib.Path, rows: list[dict]) -> pathlib.Path:
    """Write a simple XLSX workbook and return its path."""
    if not rows:
        rows = [{"Title": "T1", "Authors": "A1", "Publication Year": "2020"}]
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    headers = list(rows[0].keys())
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=1, column=ci, value=h)
    for ri, row in enumerate(rows, start=2):
        for ci, h in enumerate(headers, start=1):
            ws.cell(row=ri, column=ci, value=row.get(h, ""))
    path = tmp_path / "table.xlsx"
    wb.save(str(path))
    return path


def _make_xlsx_table_with_inline_descriptions(
    tmp_path: pathlib.Path,
    descriptions: dict[str, str],
    rows: list[dict],
) -> pathlib.Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    headers = list(rows[0].keys())
    for ci, header in enumerate(headers, start=1):
        ws.cell(row=1, column=ci, value=header)
        ws.cell(row=2, column=ci, value=descriptions.get(header, ""))
    for ri, row in enumerate(rows, start=3):
        for ci, header in enumerate(headers, start=1):
            ws.cell(row=ri, column=ci, value=row.get(header, ""))
    path = tmp_path / "table_inline.xlsx"
    wb.save(str(path))
    return path


def _make_proposal_and_accept(
    run_dir: pathlib.Path,
    run_id: str,
    row_index: int,
    title: str,
    column_name: str,
    proposed_value: str,
    decision: ReviewDecision = ReviewDecision.accepted,
    edited_value: str | None = None,
) -> tuple[ProposalRecord, dict]:
    """Create a proposal, record a decision, and return (proposal, decision_record)."""
    row_id = generate_row_id(row_index, title)
    cell_id = generate_cell_id(row_id, column_name)
    proposal_id = generate_proposal_id(run_id, cell_id)
    prop = ProposalRecord(
        proposal_id=proposal_id,
        run_id=run_id,
        pdf_id="pdf_test",
        row_id=row_id,
        column_name=column_name,
        cell_id=cell_id,
        state=ProposalState.found,
        support=SupportLabel.direct_evidence,
        proposed_value=proposed_value,
        evidence_ids=[],
        warning_flags=[],
        created_at=datetime.now(timezone.utc).isoformat(),
    )
    persist_proposal(run_dir, prop)
    dec = record_review_decision(
        run_dir=run_dir,
        run_id=run_id,
        proposal_id=proposal_id,
        cell_id=cell_id,
        decision=decision,
        resolution_reason=ReviewResolutionReason.accepted_as_proposed,
        edited_value=edited_value,
    )
    return prop, dec


# ---------------------------------------------------------------------------
# T097 — detect_unsupported_features
# ---------------------------------------------------------------------------

class TestDetectUnsupportedFeatures:
    def test_clean_workbook_has_no_warnings(self, tmp_path: pathlib.Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Title"
        ws["B1"] = "Value"
        ws["A2"] = "Test"
        ws["B2"] = "42"
        warnings = detect_unsupported_features(wb)
        assert warnings == []

    def test_formula_detected(self, tmp_path: pathlib.Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "=SUM(B1:B10)"
        warnings = detect_unsupported_features(wb)
        assert any("formula" in w.lower() for w in warnings)

    def test_formula_detected_only_once(self, tmp_path: pathlib.Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "=SUM(B1:B10)"
        ws["A2"] = "=A1*2"
        warnings = detect_unsupported_features(wb)
        formula_warnings = [w for w in warnings if "formula" in w.lower()]
        assert len(formula_warnings) == 1

    def test_frozen_panes_detected(self, tmp_path: pathlib.Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.freeze_panes = "A2"
        warnings = detect_unsupported_features(wb)
        assert any("frozen" in w.lower() for w in warnings)

    def test_merged_cells_detected(self, tmp_path: pathlib.Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.merge_cells("A1:B1")
        warnings = detect_unsupported_features(wb)
        assert any("merged" in w.lower() for w in warnings)

    def test_hidden_row_detected(self, tmp_path: pathlib.Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.row_dimensions[2].hidden = True
        warnings = detect_unsupported_features(wb)
        assert any("hidden row" in w.lower() for w in warnings)

    def test_hidden_column_detected(self, tmp_path: pathlib.Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.column_dimensions["B"].hidden = True
        warnings = detect_unsupported_features(wb)
        assert any("hidden column" in w.lower() for w in warnings)


# ---------------------------------------------------------------------------
# T096 — generate_xlsx_export (content-only, accepted-only, highlighting)
# ---------------------------------------------------------------------------

class TestGenerateXlsxExport:
    def test_writes_xlsx_to_exports_dir(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        rows = [
            {"Title": "Paper One", "Authors": "Smith", "Publication Year": "2020", "Method": ""},
        ]
        table_path = _make_xlsx_table(tmp_path, rows)
        row_index = 0
        prop, _ = _make_proposal_and_accept(run_dir, run_id, row_index, "Paper One", "Method", "PCR")
        candidates = get_export_candidates(run_dir)
        out = generate_xlsx_export(run_dir, candidates, str(table_path))
        assert out.exists()
        assert out.suffix == ".xlsx"
        assert (run_dir / "exports").is_dir()

    def test_accepted_change_written_to_cell(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        rows = [{"Title": "Paper One", "Authors": "A", "Publication Year": "2020", "Method": ""}]
        table_path = _make_xlsx_table(tmp_path, rows)
        _make_proposal_and_accept(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        candidates = get_export_candidates(run_dir)
        out = generate_xlsx_export(run_dir, candidates, str(table_path))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws is not None
        # Header row is 1, first data row is 2
        # Find column index of "Method"
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        method_col = headers.index("Method") + 1
        assert ws.cell(row=2, column=method_col).value == "PCR"

    def test_inline_description_row_preserved_and_data_row_updated(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        table_path = _make_xlsx_table_with_inline_descriptions(
            tmp_path,
            descriptions={
                "Title": "Exact title of the publication",
                "Authors": "Full author list of the paper",
                "Publication Year": "4-digit year of publication",
                "Species": "Species of origin of the biological system being assayed",
            },
            rows=[
                {"Title": "Paper A", "Authors": "Smith", "Publication Year": "2024", "Species": "human"},
                {"Title": "Paper B", "Authors": "Jones", "Publication Year": "2023", "Species": "mouse"},
            ],
        )
        _make_proposal_and_accept(run_dir, run_id, 0, "Paper A", "Species", "rat")
        candidates = get_export_candidates(run_dir)

        out = generate_xlsx_export(run_dir, candidates, str(table_path))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws is not None

        assert ws["D2"].value == "Species of origin of the biological system being assayed"
        assert ws["D3"].value == "rat"
        assert ws["D4"].value == "mouse"
        assert ws["D3"].fill.fill_type == "solid"

    def test_accepted_with_edit_uses_edited_value(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        rows = [{"Title": "Paper One", "Authors": "A", "Publication Year": "2020", "Method": ""}]
        table_path = _make_xlsx_table(tmp_path, rows)
        _make_proposal_and_accept(
            run_dir, run_id, 0, "Paper One", "Method", "PCR",
            decision=ReviewDecision.accepted_with_edit,
            edited_value="RT-PCR",
        )
        candidates = get_export_candidates(run_dir)
        out = generate_xlsx_export(run_dir, candidates, str(table_path))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws is not None
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        method_col = headers.index("Method") + 1
        assert ws.cell(row=2, column=method_col).value == "RT-PCR"

    def test_rejected_proposal_not_written(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        rows = [{"Title": "Paper One", "Authors": "A", "Publication Year": "2020", "Method": ""}]
        table_path = _make_xlsx_table(tmp_path, rows)
        row_id = generate_row_id(0, "Paper One")
        cell_id = generate_cell_id(row_id, "Method")
        proposal_id = generate_proposal_id(run_id, cell_id)
        prop = ProposalRecord(
            proposal_id=proposal_id, run_id=run_id, pdf_id="pdf_test",
            row_id=row_id, column_name="Method", cell_id=cell_id,
            state=ProposalState.found, support=SupportLabel.direct_evidence,
            proposed_value="PCR", evidence_ids=[], warning_flags=[],
            created_at=datetime.now(timezone.utc).isoformat(),
        )
        persist_proposal(run_dir, prop)
        record_review_decision(
            run_dir=run_dir, run_id=run_id, proposal_id=proposal_id,
            cell_id=cell_id, decision=ReviewDecision.rejected,
            resolution_reason=ReviewResolutionReason.rejected_incorrect,
        )
        candidates = get_export_candidates(run_dir)
        assert len(candidates) == 0
        out = generate_xlsx_export(run_dir, candidates, str(table_path))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws is not None
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        method_col = headers.index("Method") + 1
        # Should remain empty (original value — openpyxl may return None or "")
        val = ws.cell(row=2, column=method_col).value
        assert val is None or val == ""

    def test_unreviewed_proposal_not_written(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        rows = [{"Title": "Paper One", "Authors": "A", "Publication Year": "2020", "Method": ""}]
        table_path = _make_xlsx_table(tmp_path, rows)
        # Create proposal but do NOT record a decision
        row_id = generate_row_id(0, "Paper One")
        cell_id = generate_cell_id(row_id, "Method")
        proposal_id = generate_proposal_id(run_id, cell_id)
        prop = ProposalRecord(
            proposal_id=proposal_id, run_id=run_id, pdf_id="pdf_test",
            row_id=row_id, column_name="Method", cell_id=cell_id,
            state=ProposalState.found, support=SupportLabel.direct_evidence,
            proposed_value="PCR", evidence_ids=[], warning_flags=[],
            created_at=datetime.now(timezone.utc).isoformat(),
        )
        persist_proposal(run_dir, prop)
        candidates = get_export_candidates(run_dir)
        assert len(candidates) == 0
        out = generate_xlsx_export(run_dir, candidates, str(table_path))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws is not None
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        method_col = headers.index("Method") + 1
        # Should remain empty (openpyxl may return None or "")
        val = ws.cell(row=2, column=method_col).value
        assert val is None or val == ""

    def test_changed_cell_has_highlight(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        rows = [{"Title": "Paper One", "Authors": "A", "Publication Year": "2020", "Method": ""}]
        table_path = _make_xlsx_table(tmp_path, rows)
        _make_proposal_and_accept(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        candidates = get_export_candidates(run_dir)
        out = generate_xlsx_export(run_dir, candidates, str(table_path))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws is not None
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        method_col = headers.index("Method") + 1
        cell = ws.cell(row=2, column=method_col)
        # Check that fill is applied (yellow = FFFF00)
        assert cell.fill is not None
        assert cell.fill.fgColor is not None
        assert "FF00" in cell.fill.fgColor.rgb or cell.fill.patternType == "solid"

    def test_unchanged_cell_has_no_highlight(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        rows = [{"Title": "Paper One", "Authors": "A", "Publication Year": "2020", "Method": ""}]
        table_path = _make_xlsx_table(tmp_path, rows)
        # No proposals accepted
        candidates = get_export_candidates(run_dir)
        out = generate_xlsx_export(run_dir, candidates, str(table_path))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws is not None
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        method_col = headers.index("Method") + 1
        cell = ws.cell(row=2, column=method_col)
        # Unchanged cell should not have yellow fill
        is_yellow = (
            cell.fill is not None
            and cell.fill.patternType == "solid"
            and cell.fill.fgColor is not None
            and "FF00" in cell.fill.fgColor.rgb
        )
        assert not is_yellow

    def test_formula_stripped_in_export(self, tmp_path: pathlib.Path):
        """Formulas in source workbook must not appear in export (content-only)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Title"
        ws["B1"] = "Authors"
        ws["C1"] = "Publication Year"
        ws["D1"] = "Method"
        ws["A2"] = "Paper One"
        ws["B2"] = "A"
        ws["C2"] = "2020"
        ws["D2"] = "=SUM(1,2)"  # formula
        table_path = tmp_path / "table_formula.xlsx"
        wb.save(str(table_path))

        run_dir, run_id = _make_run(tmp_path / "run_formula")
        candidates: list[dict] = []
        out = generate_xlsx_export(run_dir, candidates, str(table_path))
        out_wb = openpyxl.load_workbook(str(out), data_only=True)
        out_ws = out_wb.active
        assert out_ws is not None
        headers = [out_ws.cell(row=1, column=c).value for c in range(1, out_ws.max_column + 1)]
        method_col = headers.index("Method") + 1
        val = out_ws.cell(row=2, column=method_col).value
        # The export must not contain a formula string
        assert not (isinstance(val, str) and val.startswith("="))

    def test_csv_source_creates_xlsx_export(self, tmp_path: pathlib.Path):
        """Export from a CSV source must still produce an XLSX file."""
        import csv
        csv_path = tmp_path / "table.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["Title", "Authors", "Publication Year", "Method"])
            writer.writeheader()
            writer.writerow({"Title": "Paper One", "Authors": "A", "Publication Year": "2020", "Method": ""})

        run_dir, run_id = _make_run(tmp_path / "run_csv")
        _make_proposal_and_accept(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        candidates = get_export_candidates(run_dir)
        out = generate_xlsx_export(run_dir, candidates, str(csv_path))
        assert out.suffix == ".xlsx"
        assert out.exists()
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws is not None
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Method" in headers


# ---------------------------------------------------------------------------
# T098 — generate_audit_log
# ---------------------------------------------------------------------------

class TestGenerateAuditLog:
    def test_audit_log_written_to_exports(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        _make_proposal_and_accept(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        candidates = get_export_candidates(run_dir)
        path = generate_audit_log(run_dir, candidates)
        assert path.exists()
        assert path.suffix == ".json"

    def test_audit_log_has_required_fields(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        _make_proposal_and_accept(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        candidates = get_export_candidates(run_dir)
        path = generate_audit_log(run_dir, candidates)
        data = read_json(path)
        assert "entries" in data
        assert "generated_at" in data
        assert len(data["entries"]) == 1
        entry = data["entries"][0]
        for field in ["row_id", "column_name", "cell_id", "new_value",
                      "proposal_source", "reviewer_decision", "review_decision_id",
                      "decision_timestamp"]:
            assert field in entry, f"Missing field: {field}"

    def test_audit_log_records_correct_value(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        _make_proposal_and_accept(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        candidates = get_export_candidates(run_dir)
        path = generate_audit_log(run_dir, candidates)
        data = read_json(path)
        entry = data["entries"][0]
        assert entry["new_value"] == "PCR"
        assert entry["column_name"] == "Method"

    def test_audit_log_edited_value_for_accepted_with_edit(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        _make_proposal_and_accept(
            run_dir, run_id, 0, "Paper One", "Method", "PCR",
            decision=ReviewDecision.accepted_with_edit,
            edited_value="RT-PCR",
        )
        candidates = get_export_candidates(run_dir)
        path = generate_audit_log(run_dir, candidates)
        data = read_json(path)
        entry = data["entries"][0]
        assert entry["new_value"] == "RT-PCR"
        assert entry["edited_value"] == "RT-PCR"

    def test_audit_log_empty_when_no_accepted(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        candidates: list[dict] = []
        path = generate_audit_log(run_dir, candidates)
        data = read_json(path)
        assert data["entries"] == []

    def test_audit_log_decision_timestamp_from_decision_record(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        _make_proposal_and_accept(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        candidates = get_export_candidates(run_dir)
        path = generate_audit_log(run_dir, candidates)
        data = read_json(path)
        entry = data["entries"][0]
        # Timestamp should be a non-null ISO string from the decision record
        assert entry["decision_timestamp"] is not None
        assert isinstance(entry["decision_timestamp"], str)


# ---------------------------------------------------------------------------
# T099 — generate_diagnostics
# ---------------------------------------------------------------------------

class TestGenerateDiagnostics:
    def test_diagnostics_file_written(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        run_data = read_json(run_dir / "run.json")
        path = generate_diagnostics(run_dir, run_data, [])
        assert path.exists()
        assert path.suffix == ".json"

    def test_diagnostics_has_required_sections(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        run_data = read_json(run_dir / "run.json")
        path = generate_diagnostics(run_dir, run_data, [])
        data = read_json(path)
        for section in ["run_id", "run_status", "completed_with_warnings",
                        "run_warnings", "matching", "proposals",
                        "unsupported_workbook_features"]:
            assert section in data, f"Missing section: {section}"

    def test_diagnostics_unsupported_features_included(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        run_data = read_json(run_dir / "run.json")
        warnings = ["Cell formulas will not be preserved in the exported workbook"]
        path = generate_diagnostics(run_dir, run_data, warnings)
        data = read_json(path)
        assert data["unsupported_workbook_features"]["warnings_count"] == 1
        assert warnings[0] in data["unsupported_workbook_features"]["warnings"]

    def test_diagnostics_completed_with_warnings_flag(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path, status=RunStatus.completed_with_warnings.value)
        run_data = read_json(run_dir / "run.json")
        path = generate_diagnostics(run_dir, run_data, [])
        data = read_json(path)
        assert data["completed_with_warnings"] is True

    def test_diagnostics_completed_clean_flag(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path, status=RunStatus.completed.value)
        run_data = read_json(run_dir / "run.json")
        path = generate_diagnostics(run_dir, run_data, [])
        data = read_json(path)
        assert data["completed_with_warnings"] is False

    def test_diagnostics_blocked_proposals_listed(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        row_id = generate_row_id(0, "Paper One")
        cell_id = generate_cell_id(row_id, "Method")
        proposal_id = generate_proposal_id(run_id, cell_id)
        prop = ProposalRecord(
            proposal_id=proposal_id, run_id=run_id, pdf_id="pdf_test",
            row_id=row_id, column_name="Method", cell_id=cell_id,
            state=ProposalState.blocked, support=SupportLabel.blocked,
            proposed_value=None, evidence_ids=[], warning_flags=[],
            created_at=datetime.now(timezone.utc).isoformat(),
        )
        persist_proposal(run_dir, prop)
        run_data = read_json(run_dir / "run.json")
        path = generate_diagnostics(run_dir, run_data, [])
        data = read_json(path)
        assert data["proposals"]["blocked_count"] == 1
        assert any(e["proposal_id"] == proposal_id for e in data["proposals"]["blocked"])

    def test_diagnostics_fidelity_boundary_present(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        run_data = read_json(run_dir / "run.json")
        path = generate_diagnostics(run_dir, run_data, [])
        data = read_json(path)
        boundary = data["unsupported_workbook_features"]["fidelity_boundary"]
        assert isinstance(boundary, str) and len(boundary) > 10


# ---------------------------------------------------------------------------
# T096-T099 — run_export orchestration
# ---------------------------------------------------------------------------

class TestRunExport:
    def test_run_completion_alone_does_not_create_exports(self, tmp_path: pathlib.Path):
        run_dir, _run_id = _make_run(tmp_path)
        assert list((run_dir / "exports").iterdir()) == []

    def test_run_export_returns_paths(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        rows = [{"Title": "Paper One", "Authors": "A", "Publication Year": "2020", "Method": ""}]
        table_path = _make_xlsx_table(tmp_path, rows)
        write_json(run_dir / "config.snapshot.json", {"table_path": str(table_path)})
        _make_proposal_and_accept(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        result = run_export(run_dir, str(tmp_path), run_id)
        assert pathlib.Path(result["workbook_path"]).exists()
        assert pathlib.Path(result["audit_log_path"]).exists()
        assert pathlib.Path(result["diagnostics_path"]).exists()

    def test_run_export_accepted_changes_count(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        rows = [
            {"Title": "Paper One", "Authors": "A", "Publication Year": "2020", "Method": "", "Year": ""},
        ]
        table_path = _make_xlsx_table(tmp_path, rows)
        write_json(run_dir / "config.snapshot.json", {"table_path": str(table_path)})
        _make_proposal_and_accept(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        _make_proposal_and_accept(run_dir, run_id, 0, "Paper One", "Year", "2022")
        result = run_export(run_dir, str(tmp_path), run_id)
        assert result["accepted_changes_count"] == 2

    def test_run_export_fails_for_non_completed_run(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path, status=RunStatus.running.value)
        rows = [{"Title": "T", "Authors": "A", "Publication Year": "2020"}]
        table_path = _make_xlsx_table(tmp_path, rows)
        write_json(run_dir / "config.snapshot.json", {"table_path": str(table_path)})
        with pytest.raises(ValueError, match="status"):
            run_export(run_dir, str(tmp_path), run_id)

    def test_run_export_includes_unsupported_feature_warnings(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        # Create a workbook with a formula
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Title"
        ws["B1"] = "Authors"
        ws["C1"] = "Publication Year"
        ws["D1"] = "Method"
        ws["A2"] = "Paper One"
        ws["B2"] = "A"
        ws["C2"] = "2020"
        ws["D2"] = "=SUM(1,2)"
        table_path = tmp_path / "table_with_formula.xlsx"
        wb.save(str(table_path))
        write_json(run_dir / "config.snapshot.json", {"table_path": str(table_path)})
        result = run_export(run_dir, str(tmp_path), run_id)
        assert result["unsupported_feature_warnings_count"] >= 1
        assert any("formula" in w.lower() for w in result["unsupported_feature_warnings"])

    def test_run_export_fidelity_boundary_in_result(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        rows = [{"Title": "T", "Authors": "A", "Publication Year": "2020"}]
        table_path = _make_xlsx_table(tmp_path, rows)
        write_json(run_dir / "config.snapshot.json", {"table_path": str(table_path)})
        result = run_export(run_dir, str(tmp_path), run_id)
        assert "fidelity_boundary" in result
        assert len(result["fidelity_boundary"]) > 10

    def test_run_export_persists_manual_export_summary_in_run_json(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        rows = [{"Title": "Paper One", "Authors": "A", "Publication Year": "2020", "Method": ""}]
        table_path = _make_xlsx_table(tmp_path, rows)
        write_json(run_dir / "config.snapshot.json", {"table_path": str(table_path)})
        _make_proposal_and_accept(run_dir, run_id, 0, "Paper One", "Method", "PCR")

        result = run_export(run_dir, str(tmp_path), run_id)
        run_json = read_json(run_dir / "run.json")

        assert run_json["last_export"]["accepted_changes_count"] == 1
        assert run_json["last_export"]["workbook_path"] == result["workbook_path"]


# ---------------------------------------------------------------------------
# T100/API — POST /api/runs/{run_id}/export endpoint
# ---------------------------------------------------------------------------

class TestExportEndpoint:
    def test_export_endpoint_404_unknown_run(self, tmp_path: pathlib.Path):
        resp = client.post("/api/runs/nonexistent_run/export?output_dir=" + str(tmp_path))
        assert resp.status_code == 404

    def test_export_endpoint_422_non_completed_run(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path, status=RunStatus.running.value)
        rows = [{"Title": "T", "Authors": "A", "Publication Year": "2020"}]
        table_path = _make_xlsx_table(tmp_path, rows)
        write_json(run_dir / "config.snapshot.json", {"table_path": str(table_path)})
        resp = client.post(f"/api/runs/{run_id}/export?output_dir={tmp_path}")
        assert resp.status_code == 422

    def test_export_endpoint_success(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        rows = [{"Title": "Paper One", "Authors": "A", "Publication Year": "2020", "Method": ""}]
        table_path = _make_xlsx_table(tmp_path, rows)
        write_json(run_dir / "config.snapshot.json", {"table_path": str(table_path)})
        _make_proposal_and_accept(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        resp = client.post(f"/api/runs/{run_id}/export?output_dir={tmp_path}")
        assert resp.status_code == 200
        data = resp.json()
        assert "workbook_path" in data
        assert "audit_log_path" in data
        assert "diagnostics_path" in data
        assert data["accepted_changes_count"] == 1

    def test_download_workbook_requires_manual_export_trigger(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        rows = [{"Title": "Paper One", "Authors": "A", "Publication Year": "2020", "Method": ""}]
        table_path = _make_xlsx_table(tmp_path, rows)
        write_json(run_dir / "config.snapshot.json", {"table_path": str(table_path)})
        _make_proposal_and_accept(run_dir, run_id, 0, "Paper One", "Method", "PCR")

        resp = client.get(f"/api/runs/{run_id}/downloads/workbook?output_dir={tmp_path}")

        assert resp.status_code == 404
        assert "Trigger export" in resp.json()["detail"]
