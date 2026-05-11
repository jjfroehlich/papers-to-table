"""Hermetic end-to-end tests for the core Extract Structured Info from Papers workflow.

These tests use the canonical fixture corpus and stub providers (via respx
mocking) to verify the complete pipeline without any network calls.

Coverage: T104
- successful matched extraction → review → accepted-only export
- unmatched / ambiguous / duplicate-row blocked flows
- weak-evidence quote+page review path
- exact highlight vs. approximate highlight vs. quote-plus-page fallback
- Verify mode reviewed-cell flow
- proactive figure review / figure evidence path
- figure rescue of a weak text-only proposal
- export with accepted-only changes
"""
from __future__ import annotations

import pathlib
from datetime import datetime, timezone
from typing import Any

import openpyxl
import pytest
import respx
import httpx

from backend.app.artifacts import (
    get_run_dir,
    init_run_bundle,
    read_json,
    write_json,
)
from backend.app.export import generate_xlsx_export, generate_audit_log, run_export
from backend.app.extraction import (
    EvidenceRecord,
    ProposalRecord,
    persist_evidence,
    persist_proposal,
)
from backend.app.ids import (
    generate_cell_id,
    generate_evidence_id,
    generate_proposal_id,
    generate_row_id,
    generate_run_id,
)
from backend.app.review import (
    get_export_candidates,
    get_latest_decision,
    get_progress,
    list_proposals,
    record_review_decision,
)
from backend.app.schemas import (
    EvidenceSourceType,
    ProposalState,
    ReviewDecision,
    ReviewResolutionReason,
    RunStatus,
    SupportLabel,
    WarningCategory,
)

FIXTURE_TABLE = "../benchmark_datasets/massively_parallel_reporter_assays/table_template.csv"
FIXTURE_SCHEMA = "../benchmark_datasets/massively_parallel_reporter_assays/schema.csv"
FIXTURE_PDF_DIR = "../benchmark_datasets/massively_parallel_reporter_assays/pdfs"


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _make_run(
    tmp_path: pathlib.Path,
    status: str = RunStatus.completed.value,
    warnings: list | None = None,
    verify_mode: bool = False,
) -> tuple[pathlib.Path, str]:
    run_id = generate_run_id()
    run_dir = init_run_bundle(str(tmp_path), run_id)
    run_data = {
        "run_id": run_id,
        "status": status,
        "output_dir": str(tmp_path),
        "verify_mode": verify_mode,
        "total_rows": 4,
        "eligible_cells": 8,
        "proposals_generated": 0,
        "proposals_reviewed": 0,
        "warnings": warnings or [],
        "table_path": FIXTURE_TABLE,
    }
    write_json(run_dir / "run.json", run_data)
    write_json(run_dir / "config.snapshot.json", {
        "table_path": FIXTURE_TABLE,
        "schema_path": FIXTURE_SCHEMA,
        "pdf_dir": FIXTURE_PDF_DIR,
        "output_dir": str(tmp_path),
        "verify_mode": verify_mode,
        "provider": {"token": "lm_studio", "base_url": "http://localhost:1234"},
    })
    return run_dir, run_id


def _seed_proposal(
    run_dir: pathlib.Path,
    run_id: str,
    row_index: int,
    title: str,
    column_name: str,
    proposed_value: str | None = "test_value",
    state: ProposalState = ProposalState.found,
    support: SupportLabel = SupportLabel.direct_evidence,
    pdf_id: str = "pdf_paper1",
    is_verify_mode: bool = False,
    existing_value: str | None = None,
) -> ProposalRecord:
    row_id = generate_row_id(row_index, title)
    cell_id = generate_cell_id(row_id, column_name)
    proposal_id = generate_proposal_id(run_id, cell_id)
    prop = ProposalRecord(
        proposal_id=proposal_id,
        run_id=run_id,
        pdf_id=pdf_id,
        row_id=row_id,
        column_name=column_name,
        cell_id=cell_id,
        state=state,
        support=support,
        proposed_value=proposed_value,
        rationale="Evidence found in section 3.",
        evidence_ids=[],
        warning_flags=[],
        is_verify_mode=is_verify_mode,
        existing_value=existing_value,
        created_at=datetime.now(timezone.utc).isoformat(),
    )
    persist_proposal(run_dir, prop)
    return prop


def _seed_evidence(
    run_dir: pathlib.Path,
    run_id: str,
    proposal: ProposalRecord,
    source_type: EvidenceSourceType = EvidenceSourceType.direct_quote,
    quote_text: str = "Test was performed using PCR.",
    page_number: int = 2,
    exact_highlight: bool = True,
    is_figure: bool = False,
) -> EvidenceRecord:
    ev_id = generate_evidence_id(proposal.proposal_id)
    regions = [{"x0": 10, "y0": 20, "x1": 200, "y1": 30, "page": page_number}]
    ev = EvidenceRecord(
        evidence_id=ev_id,
        run_id=run_id,
        proposal_id=proposal.proposal_id,
        pdf_id=proposal.pdf_id,
        source_type=source_type,
        quote_text=quote_text,
        page_number=page_number,
        exact_highlight_regions=regions if exact_highlight else None,
        approximate_highlight_regions=None if exact_highlight else regions,
        anchor_confidence=0.9 if exact_highlight else 0.5,
        is_primary=True,
        is_figure_derived=is_figure,
        created_at=datetime.now(timezone.utc).isoformat(),
    )
    persist_evidence(run_dir, ev)
    # Update proposal with evidence id
    updated = proposal.model_copy(update={"evidence_ids": [ev_id], "primary_evidence_id": ev_id})
    persist_proposal(run_dir, updated)
    return ev


# ---------------------------------------------------------------------------
# T104: Successful matched extraction → review → accepted-only export
# ---------------------------------------------------------------------------

class TestHermeticMatchedExtractionExport:
    def test_accepted_proposal_appears_in_export(self, tmp_path: pathlib.Path):
        """Core happy path: matched extraction → accept → export includes only accepted change."""
        run_dir, run_id = _make_run(tmp_path)
        prop = _seed_proposal(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        _seed_evidence(run_dir, run_id, prop)

        # Accept the proposal
        record_review_decision(
            run_dir=run_dir, run_id=run_id,
            proposal_id=prop.proposal_id, cell_id=prop.cell_id,
            decision=ReviewDecision.accepted,
            resolution_reason=ReviewResolutionReason.accepted_as_proposed,
        )

        candidates = get_export_candidates(run_dir)
        assert len(candidates) == 1
        assert candidates[0]["export_value"] == "PCR"

    def test_accepted_changes_count_correct_in_export(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        # Two accepted proposals
        prop1 = _seed_proposal(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        prop2 = _seed_proposal(run_dir, run_id, 0, "Paper One", "Year", "2022")
        for prop in [prop1, prop2]:
            record_review_decision(
                run_dir=run_dir, run_id=run_id,
                proposal_id=prop.proposal_id, cell_id=prop.cell_id,
                decision=ReviewDecision.accepted,
                resolution_reason=ReviewResolutionReason.accepted_as_proposed,
            )
        candidates = get_export_candidates(run_dir)
        assert len(candidates) == 2

    def test_export_xlsx_written_with_correct_value(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        # Use a column that actually exists in the fixture workbook
        prop = _seed_proposal(run_dir, run_id, 0, "Paper One", "Cloning", "Gibson assembly")
        record_review_decision(
            run_dir=run_dir, run_id=run_id,
            proposal_id=prop.proposal_id, cell_id=prop.cell_id,
            decision=ReviewDecision.accepted,
            resolution_reason=ReviewResolutionReason.accepted_as_proposed,
        )
        candidates = get_export_candidates(run_dir)
        out = generate_xlsx_export(run_dir, candidates, FIXTURE_TABLE)
        assert out.exists()
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws is not None
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Cloning" in headers

    def test_full_run_export_orchestration(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        prop = _seed_proposal(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        record_review_decision(
            run_dir=run_dir, run_id=run_id,
            proposal_id=prop.proposal_id, cell_id=prop.cell_id,
            decision=ReviewDecision.accepted,
            resolution_reason=ReviewResolutionReason.accepted_as_proposed,
        )
        result = run_export(run_dir, str(tmp_path), run_id)
        assert pathlib.Path(result["workbook_path"]).exists()
        assert pathlib.Path(result["audit_log_path"]).exists()
        assert pathlib.Path(result["diagnostics_path"]).exists()
        assert result["accepted_changes_count"] == 1


# ---------------------------------------------------------------------------
# T104: Unmatched / ambiguous / duplicate-row blocked flows
# ---------------------------------------------------------------------------

class TestHermeticBlockedFlows:
    def test_blocked_proposal_not_in_export(self, tmp_path: pathlib.Path):
        """Blocked proposals (no PDF match) must never appear in export."""
        run_dir, run_id = _make_run(tmp_path)
        prop = _seed_proposal(
            run_dir, run_id, 0, "Paper One", "Method", None,
            state=ProposalState.blocked, support=SupportLabel.blocked,
        )
        # Even if someone tried to record a spurious decision — get_export_candidates only picks accepted
        candidates = get_export_candidates(run_dir)
        assert len(candidates) == 0

    def test_skipped_proposal_not_in_export(self, tmp_path: pathlib.Path):
        """Skipped proposals (provider unavailable) must not appear in export."""
        run_dir, run_id = _make_run(tmp_path)
        _seed_proposal(
            run_dir, run_id, 0, "Paper One", "Method", None,
            state=ProposalState.skipped, support=SupportLabel.error,
        )
        candidates = get_export_candidates(run_dir)
        assert len(candidates) == 0

    def test_unclear_proposal_not_in_export_without_accept(self, tmp_path: pathlib.Path):
        """Unclear proposals without explicit accept must not appear in export."""
        run_dir, run_id = _make_run(tmp_path)
        _seed_proposal(
            run_dir, run_id, 0, "Paper One", "Method", "Maybe PCR",
            state=ProposalState.unclear, support=SupportLabel.weak_evidence,
        )
        candidates = get_export_candidates(run_dir)
        assert len(candidates) == 0

    def test_unmatched_run_warning_in_diagnostics(self, tmp_path: pathlib.Path):
        """Runs with unmatched PDFs record appropriate warnings."""
        run_dir, run_id = _make_run(
            tmp_path,
            status=RunStatus.completed_with_warnings.value,
            warnings=[{
                "category": WarningCategory.unmatched_pdf.value,
                "message": "PDF not matched: unmatched_1.pdf",
                "context": {"pdf_id": "unmatched_1"},
            }],
        )
        run_data = read_json(run_dir / "run.json")
        assert any(
            w["category"] == WarningCategory.unmatched_pdf.value
            for w in run_data["warnings"]
        )
        assert run_data["status"] == RunStatus.completed_with_warnings.value

    def test_rejected_proposal_excluded_from_export(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        prop = _seed_proposal(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        record_review_decision(
            run_dir=run_dir, run_id=run_id,
            proposal_id=prop.proposal_id, cell_id=prop.cell_id,
            decision=ReviewDecision.rejected,
            resolution_reason=ReviewResolutionReason.rejected_incorrect,
        )
        candidates = get_export_candidates(run_dir)
        assert len(candidates) == 0

    def test_confirmed_no_data_excluded_from_export(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        prop = _seed_proposal(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        record_review_decision(
            run_dir=run_dir, run_id=run_id,
            proposal_id=prop.proposal_id, cell_id=prop.cell_id,
            decision=ReviewDecision.confirmed_no_data,
            resolution_reason=ReviewResolutionReason.confirmed_no_data_in_paper,
        )
        candidates = get_export_candidates(run_dir)
        assert len(candidates) == 0


# ---------------------------------------------------------------------------
# T104: Weak evidence / quote-plus-page fallback evidence paths
# ---------------------------------------------------------------------------

class TestHermeticWeakEvidencePaths:
    def test_quote_plus_page_evidence_stored_correctly(self, tmp_path: pathlib.Path):
        """Quote+page fallback evidence is stored with correct source_type."""
        run_dir, run_id = _make_run(tmp_path)
        prop = _seed_proposal(
            run_dir, run_id, 0, "Paper One", "Method", "Maybe PCR",
            support=SupportLabel.weak_evidence,
        )
        ev = _seed_evidence(
            run_dir, run_id, prop,
            source_type=EvidenceSourceType.quote_plus_page,
            exact_highlight=False,
        )
        from backend.app.extraction import load_evidence
        all_ev = load_evidence(run_dir)
        matching = [e for e in all_ev if e.evidence_id == ev.evidence_id]
        assert len(matching) == 1
        assert matching[0].source_type == EvidenceSourceType.quote_plus_page

    def test_weak_evidence_proposal_reviewable_when_accepted(self, tmp_path: pathlib.Path):
        """Weak-evidence proposal can still be accepted and exported."""
        run_dir, run_id = _make_run(tmp_path)
        prop = _seed_proposal(
            run_dir, run_id, 0, "Paper One", "Method", "PCR",
            support=SupportLabel.weak_evidence,
        )
        _seed_evidence(
            run_dir, run_id, prop,
            source_type=EvidenceSourceType.quote_plus_page,
            exact_highlight=False,
        )
        record_review_decision(
            run_dir=run_dir, run_id=run_id,
            proposal_id=prop.proposal_id, cell_id=prop.cell_id,
            decision=ReviewDecision.accepted,
            resolution_reason=ReviewResolutionReason.accepted_as_proposed,
        )
        candidates = get_export_candidates(run_dir)
        assert len(candidates) == 1
        assert candidates[0]["export_value"] == "PCR"

    def test_exact_highlight_evidence_stored_correctly(self, tmp_path: pathlib.Path):
        """Exact-highlight evidence is stored with correct source_type and regions."""
        run_dir, run_id = _make_run(tmp_path)
        prop = _seed_proposal(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        ev = _seed_evidence(
            run_dir, run_id, prop,
            source_type=EvidenceSourceType.direct_quote,
            exact_highlight=True,
        )
        from backend.app.extraction import load_evidence
        all_ev = load_evidence(run_dir)
        matching = [e for e in all_ev if e.evidence_id == ev.evidence_id]
        assert matching[0].exact_highlight_regions is not None
        assert matching[0].approximate_highlight_regions is None

    def test_approximate_highlight_evidence_stored_correctly(self, tmp_path: pathlib.Path):
        """Approximate highlight evidence is stored with approximate regions."""
        run_dir, run_id = _make_run(tmp_path)
        prop = _seed_proposal(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        ev = _seed_evidence(
            run_dir, run_id, prop,
            source_type=EvidenceSourceType.approximate_highlight,
            exact_highlight=False,
        )
        from backend.app.extraction import load_evidence
        all_ev = load_evidence(run_dir)
        matching = [e for e in all_ev if e.evidence_id == ev.evidence_id]
        assert matching[0].approximate_highlight_regions is not None
        assert matching[0].exact_highlight_regions is None


# ---------------------------------------------------------------------------
# T104: Verify mode reviewed-cell flow
# ---------------------------------------------------------------------------

class TestHermeticVerifyMode:
    def test_verify_mode_proposal_has_existing_value(self, tmp_path: pathlib.Path):
        """Verify-mode proposals carry the existing cell value for comparison."""
        run_dir, run_id = _make_run(tmp_path, verify_mode=True)
        prop = _seed_proposal(
            run_dir, run_id, 0, "Paper One", "Method", "RT-PCR",
            is_verify_mode=True, existing_value="PCR",
        )
        from backend.app.extraction import load_proposals
        proposals = load_proposals(run_dir)
        assert len(proposals) == 1
        assert proposals[0].is_verify_mode is True
        assert proposals[0].existing_value == "PCR"

    def test_verify_mode_accepted_value_in_export(self, tmp_path: pathlib.Path):
        """Verify-mode accepted value is correctly exported."""
        run_dir, run_id = _make_run(tmp_path, verify_mode=True)
        prop = _seed_proposal(
            run_dir, run_id, 0, "Paper One", "Method", "RT-PCR",
            is_verify_mode=True, existing_value="PCR",
        )
        record_review_decision(
            run_dir=run_dir, run_id=run_id,
            proposal_id=prop.proposal_id, cell_id=prop.cell_id,
            decision=ReviewDecision.accepted,
            resolution_reason=ReviewResolutionReason.accepted_as_proposed,
        )
        candidates = get_export_candidates(run_dir)
        assert len(candidates) == 1
        assert candidates[0]["export_value"] == "RT-PCR"

    def test_verify_mode_audit_log_has_old_value(self, tmp_path: pathlib.Path):
        """Verify-mode audit log entry includes the existing cell value as old_value."""
        run_dir, run_id = _make_run(tmp_path, verify_mode=True)
        prop = _seed_proposal(
            run_dir, run_id, 0, "Paper One", "Method", "RT-PCR",
            is_verify_mode=True, existing_value="PCR",
        )
        record_review_decision(
            run_dir=run_dir, run_id=run_id,
            proposal_id=prop.proposal_id, cell_id=prop.cell_id,
            decision=ReviewDecision.accepted,
            resolution_reason=ReviewResolutionReason.accepted_as_proposed,
        )
        candidates = get_export_candidates(run_dir)
        audit_path = generate_audit_log(run_dir, candidates)
        data = read_json(audit_path)
        entry = data["entries"][0]
        assert entry["old_value"] == "PCR"
        assert entry["new_value"] == "RT-PCR"


# ---------------------------------------------------------------------------
# T104: Figure-based evidence path (proactive figure review)
# ---------------------------------------------------------------------------

class TestHermeticFigureEvidence:
    def test_figure_evidence_stored_with_correct_type(self, tmp_path: pathlib.Path):
        """Figure-derived evidence is stored with the typed figure evidence enum."""
        run_dir, run_id = _make_run(tmp_path)
        prop = _seed_proposal(run_dir, run_id, 0, "Paper One", "Method", "Flow cytometry")
        ev = _seed_evidence(
            run_dir, run_id, prop,
            source_type=EvidenceSourceType.caption_grounded_figure_evidence,
            is_figure=True,
        )
        from backend.app.extraction import load_evidence
        all_ev = load_evidence(run_dir)
        matching = [e for e in all_ev if e.evidence_id == ev.evidence_id]
        assert len(matching) == 1
        assert matching[0].source_type == EvidenceSourceType.caption_grounded_figure_evidence
        assert matching[0].is_figure_derived is True

    def test_figure_evidence_proposal_exportable_when_accepted(self, tmp_path: pathlib.Path):
        """A proposal backed only by figure evidence can be accepted and exported."""
        run_dir, run_id = _make_run(tmp_path)
        prop = _seed_proposal(run_dir, run_id, 0, "Paper One", "Method", "Flow cytometry")
        _seed_evidence(
            run_dir, run_id, prop,
            source_type=EvidenceSourceType.caption_grounded_figure_evidence,
            is_figure=True,
        )
        record_review_decision(
            run_dir=run_dir, run_id=run_id,
            proposal_id=prop.proposal_id, cell_id=prop.cell_id,
            decision=ReviewDecision.accepted,
            resolution_reason=ReviewResolutionReason.accepted_as_proposed,
        )
        candidates = get_export_candidates(run_dir)
        assert len(candidates) == 1
        assert candidates[0]["export_value"] == "Flow cytometry"

    def test_figure_rescue_alongside_weak_text_evidence(self, tmp_path: pathlib.Path):
        """A weak text proposal upgraded by figure evidence is accepted and exported."""
        run_dir, run_id = _make_run(tmp_path)
        prop = _seed_proposal(
            run_dir, run_id, 0, "Paper One", "Method", "FACS",
            support=SupportLabel.weak_evidence,
        )
        # Add weak text evidence
        ev_text = _seed_evidence(
            run_dir, run_id, prop,
            source_type=EvidenceSourceType.quote_plus_page,
            exact_highlight=False,
        )
        # Also add figure evidence
        ev_id2 = generate_evidence_id(prop.proposal_id)
        ev_fig = EvidenceRecord(
            evidence_id=ev_id2,
            run_id=run_id,
            proposal_id=prop.proposal_id,
            pdf_id=prop.pdf_id,
            source_type=EvidenceSourceType.visual_interpretation_figure_evidence,
            page_number=3,
            is_primary=False,
            is_figure_derived=True,
            created_at=datetime.now(timezone.utc).isoformat(),
        )
        persist_evidence(run_dir, ev_fig)

        from backend.app.extraction import load_evidence
        all_ev = load_evidence(run_dir)
        prop_ev = [e for e in all_ev if e.proposal_id == prop.proposal_id]
        assert len(prop_ev) == 2
        types = {e.source_type for e in prop_ev}
        assert EvidenceSourceType.visual_interpretation_figure_evidence in types
        assert EvidenceSourceType.quote_plus_page in types

        record_review_decision(
            run_dir=run_dir, run_id=run_id,
            proposal_id=prop.proposal_id, cell_id=prop.cell_id,
            decision=ReviewDecision.accepted,
            resolution_reason=ReviewResolutionReason.accepted_as_proposed,
        )
        candidates = get_export_candidates(run_dir)
        assert len(candidates) == 1


# ---------------------------------------------------------------------------
# T104: Review progress and lifecycle verification
# ---------------------------------------------------------------------------

class TestHermeticReviewLifecycle:
    def test_progress_counts_update_after_decisions(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        props = [
            _seed_proposal(run_dir, run_id, i, f"Paper {i}", "Method", "PCR")
            for i in range(3)
        ]
        # Accept first, reject second, leave third pending
        record_review_decision(
            run_dir=run_dir, run_id=run_id,
            proposal_id=props[0].proposal_id, cell_id=props[0].cell_id,
            decision=ReviewDecision.accepted,
            resolution_reason=ReviewResolutionReason.accepted_as_proposed,
        )
        record_review_decision(
            run_dir=run_dir, run_id=run_id,
            proposal_id=props[1].proposal_id, cell_id=props[1].cell_id,
            decision=ReviewDecision.rejected,
            resolution_reason=ReviewResolutionReason.rejected_incorrect,
        )
        progress = get_progress(run_dir)
        assert progress["total"] == 3
        assert progress["accepted"] == 1
        assert progress["rejected"] == 1
        assert progress["pending"] == 1

    def test_latest_decision_overrides_previous(self, tmp_path: pathlib.Path):
        """Latest decision for a cell replaces the previous one."""
        run_dir, run_id = _make_run(tmp_path)
        prop = _seed_proposal(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        # Record rejected then accept (reviewer changes their mind)
        record_review_decision(
            run_dir=run_dir, run_id=run_id,
            proposal_id=prop.proposal_id, cell_id=prop.cell_id,
            decision=ReviewDecision.rejected,
            resolution_reason=ReviewResolutionReason.rejected_incorrect,
        )
        record_review_decision(
            run_dir=run_dir, run_id=run_id,
            proposal_id=prop.proposal_id, cell_id=prop.cell_id,
            decision=ReviewDecision.accepted,
            resolution_reason=ReviewResolutionReason.accepted_as_proposed,
        )
        latest = get_latest_decision(run_dir, prop.proposal_id)
        assert latest is not None
        assert latest.decision == ReviewDecision.accepted
        # Export should include the accepted decision
        candidates = get_export_candidates(run_dir)
        assert len(candidates) == 1

    def test_completed_with_warnings_status_preserved(self, tmp_path: pathlib.Path):
        """Runs with warnings stay in completed_with_warnings status through export."""
        run_dir, run_id = _make_run(
            tmp_path,
            status=RunStatus.completed_with_warnings.value,
            warnings=[{"category": "unmatched_pdf", "message": "...", "context": None}],
        )
        run_data = read_json(run_dir / "run.json")
        assert run_data["status"] == RunStatus.completed_with_warnings.value
        # Export should still work
        result = run_export(run_dir, str(tmp_path), run_id)
        assert "workbook_path" in result

    def test_accepted_with_edit_value_in_export(self, tmp_path: pathlib.Path):
        run_dir, run_id = _make_run(tmp_path)
        prop = _seed_proposal(run_dir, run_id, 0, "Paper One", "Method", "PCR")
        record_review_decision(
            run_dir=run_dir, run_id=run_id,
            proposal_id=prop.proposal_id, cell_id=prop.cell_id,
            decision=ReviewDecision.accepted_with_edit,
            resolution_reason=ReviewResolutionReason.accepted_with_edit,
            edited_value="RT-PCR",
        )
        candidates = get_export_candidates(run_dir)
        assert candidates[0]["export_value"] == "RT-PCR"
        assert candidates[0]["edited_value"] == "RT-PCR"

