"""Tests for the staged asyncio runner."""
from __future__ import annotations

import asyncio
import pathlib
from datetime import datetime, timezone
from types import SimpleNamespace

import openpyxl
import pytest
import respx
import httpx
import pandas as pd

from backend.app.artifacts import (
    get_artifact_summary_path,
    get_config_snapshot_path,
    get_input_summary_path,
    get_provider_diagnostics_path,
    get_provider_model_management_path,
    get_provider_mode_path,
    get_provider_probe_path,
    get_run_json_path,
    get_run_stats_path,
    get_reviewer_summary_path,
    get_run_summary_path,
    init_run_bundle,
    read_json,
    write_json,
)
from backend.app.extraction import EvidenceRecord, ProposalRecord, load_proposals, persist_evidence, persist_proposal
from backend.app import runner as runner_module
from backend.app.config import RunConfig
from backend.app.runner import _parse_cache_key, get_initial_run_data, run_pipeline
from backend.app.ids import generate_proposal_id
from backend.app.matching import MatchResult
from backend.app.schemas import EvidenceSourceType, MatchOutcome, ProposalState, RunStatus, SupportLabel

FIXTURE_TABLE = "../benchmark_datasets/massively_parallel_reporter_assays/table_template.csv"
FIXTURE_SCHEMA = "../benchmark_datasets/massively_parallel_reporter_assays/schema.csv"
FIXTURE_PDF_DIR = "../benchmark_datasets/massively_parallel_reporter_assays/pdfs"
_CAPTURED_ROW_CONTEXTS: list[dict] = []


def make_config(tmp_path: pathlib.Path, **kwargs) -> RunConfig:
    data = {
        "table_path": FIXTURE_TABLE,
        "schema_path": FIXTURE_SCHEMA,
        "pdf_dir": FIXTURE_PDF_DIR,
        "output_dir": str(tmp_path / "runs"),
        "verify_mode": False,
        "provider": {
            "token": "lm_studio",
            "base_url": "http://localhost:1234",
            "text_model": {
                "model_id": "qwen/qwen3-30b-a3b-2507",
            },
        },
        **kwargs,
    }
    return RunConfig.model_validate(data)


class TestGetInitialRunData:
    def test_has_required_fields(self, tmp_path):
        config = make_config(tmp_path)
        data = get_initial_run_data("run_test_123", config, "config.json")
        assert data["run_id"] == "run_test_123"
        assert data["status"] == RunStatus.created.value
        assert data["config_path"] == "config.json"
        assert data["verify_mode"] is False
        assert data["eval_mode"] is False
        assert data["run_mode"] == "normal"
        assert data["provider_token"] == "lm_studio"
        assert data["retrieval_mode"] == "hybrid_experimental"
        assert data["prompt_hash"] is not None
        assert data["prompt_bundle_id"] == "default"
        assert data["prompt_manifest_hash"] is not None
        assert data["prompt_bundle_hash"] is not None
        assert data["prompt_bundle_path"] is not None
        assert data["prompt_keys_used"]
        assert data["prompt_files"]
        assert data["artifact_schema_version"] == "main_run_bundle.v2"
        assert data["proposal_schema_version"] == "main_proposal.v2"
        assert data["evidence_schema_version"] == "main_evidence.v2"
        assert data["structured_output_reason"] is None

    def test_timestamps(self, tmp_path):
        config = make_config(tmp_path)
        data = get_initial_run_data("run_test_123", config, None)
        assert data["started_at"] is None
        assert data["completed_at"] is None
        assert data["created_at"] is not None

    def test_defaults_parser_cache_dir_to_pdf_directory(self, tmp_path):
        pdf_dir = tmp_path / "pdfs"
        pdf_dir.mkdir()

        config = make_config(tmp_path, pdf_dir=str(pdf_dir))
        data = get_initial_run_data("run_test_123", config, None)

        assert data["parser_cache_enabled"] is True
        assert data["parser_cache_dir"] == str((pdf_dir / ".extract_structured_parse_cache").resolve())

    def test_records_style_profile_mode_by_run_mode(self, tmp_path):
        normal_config = make_config(tmp_path)
        normal_data = get_initial_run_data("run_normal", normal_config, None)

        eval_config = make_config(tmp_path, eval_mode=True)
        eval_data = get_initial_run_data("run_eval", eval_config, None)

        assert normal_data["style_profile_mode"] == "sample_rows"
        assert normal_data["style_profile_source"] == "filled_cells"
        assert normal_data["style_profile_benchmark_safe"] is False
        assert eval_data["style_profile_mode"] == "masked_rows"
        assert eval_data["style_profile_source"] == "masked_working_copy"
        assert eval_data["style_profile_benchmark_safe"] is True


class TestParseCacheKey:
    def test_changes_when_runtime_fingerprint_changes(self, tmp_path, monkeypatch):
        config = make_config(tmp_path)
        pdf_path = tmp_path / "paper.pdf"
        pdf_path.write_bytes(b"%PDF-1.4\n%stub")

        monkeypatch.setattr(runner_module, "hash_file", lambda path: "same-pdf-hash")
        monkeypatch.setattr(
            runner_module,
            "_parse_runtime_fingerprint",
            lambda config: {"python_version": "3.11.9", "package_versions": {"pypdfium2": "4.30.0"}},
        )
        baseline = _parse_cache_key(config, str(pdf_path))

        monkeypatch.setattr(
            runner_module,
            "_parse_runtime_fingerprint",
            lambda config: {"python_version": "3.12.1", "package_versions": {"pypdfium2": "4.30.0"}},
        )
        updated = _parse_cache_key(config, str(pdf_path))

        assert baseline != updated

    def test_changes_when_parse_contract_version_changes(self, tmp_path, monkeypatch):
        config = make_config(tmp_path)
        pdf_path = tmp_path / "paper.pdf"
        pdf_path.write_bytes(b"%PDF-1.4\n%stub")

        monkeypatch.setattr(runner_module, "hash_file", lambda path: "same-pdf-hash")
        monkeypatch.setattr(
            runner_module,
            "_parse_runtime_fingerprint",
            lambda config: {"python_version": "3.11.9", "package_versions": {"pypdfium2": "4.30.0"}},
        )
        baseline = _parse_cache_key(config, str(pdf_path))

        monkeypatch.setattr(runner_module, "PARSED_DOCUMENT_CONTRACT_VERSION", "parsed_document.v99")
        updated = _parse_cache_key(config, str(pdf_path))

        assert baseline != updated


class TestRunPipeline:
    @pytest.mark.asyncio
    @respx.mock
    async def test_completes_with_valid_inputs(self, tmp_path, monkeypatch):
        """Happy-path: pipeline completes when LM Studio is reachable."""
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        _patch_fast_no_match_pipeline(monkeypatch)
        config = make_config(tmp_path)
        run_id = "run_test_happy"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        assert run_data["status"] in (
            RunStatus.completed.value,
            RunStatus.completed_with_warnings.value,
        )
        assert run_data["completed_at"] is not None

    @pytest.mark.asyncio
    @respx.mock
    async def test_writes_config_snapshot(self, tmp_path, monkeypatch):
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        _patch_fast_no_match_pipeline(monkeypatch)
        config = make_config(tmp_path)
        run_id = "run_snap"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        snap_path = get_config_snapshot_path(output_dir, run_id)
        assert snap_path.exists()
        snap = read_json(snap_path)
        assert snap["provider"]["token"] == "lm_studio"
        assert snap["retrieval"]["mode"] == "hybrid_experimental"
        assert "strategy" not in snap["retrieval"]

    @pytest.mark.asyncio
    @respx.mock
    async def test_writes_run_stats_summary(self, tmp_path, monkeypatch):
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)
        monkeypatch.setattr("backend.app.runner.parse_pdf", _fake_parse_pdf)
        monkeypatch.setattr("backend.app.runner.run_matching", lambda **kwargs: [])
        monkeypatch.setattr("backend.app.runner.persist_match_artifacts", lambda *args, **kwargs: None)
        monkeypatch.setattr("backend.app.runner.run_style_profiles_stage", _fake_style_profiles)
        config = make_config(tmp_path)
        run_id = "run_stats"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        stats_path = get_run_stats_path(output_dir, run_id)
        assert stats_path.exists()
        stats = read_json(stats_path)
        assert stats["retrieval_mode"] == "hybrid_experimental"
        assert stats["per_run"]["run_total_ms"] is not None
        assert "parse" in stats["per_run"]["stage_ms"]
        assert "stage_parsing_ms" in stats["per_run"]["stage_timing_ms"]
        assert "provider_request_counts" in stats["counters"]
        assert "run_stats_path" in read_json(get_run_summary_path(output_dir, run_id))

    @pytest.mark.asyncio
    async def test_run_stats_include_compact_rollups_and_consistency(self, tmp_path, monkeypatch):
        pdf_dir = tmp_path / "pdfs"
        pdf_dir.mkdir()
        (pdf_dir / "paper_1.pdf").write_bytes(b"%PDF-1.4\n%stub")

        config = make_config(tmp_path, pdf_dir=str(pdf_dir))
        run_id = "run_stats_rollups"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        df = pd.DataFrame([
            {
                "Title": "Matched Paper",
                "Authors": "Smith, J.",
                "Publication Year": "2024",
                "assay": "",
            }
        ])
        schema = [{"column_name": "assay", "description": "Assay description"}]
        eligible = [{"row_index": 0, "column_name": "assay", "current_value": "", "eligibility": "eligible"}]
        match_results = [
            MatchResult(
                pdf_id="paper_1",
                pdf_path="paper_1.pdf",
                outcome=MatchOutcome.matched,
                matched_row_index=0,
                matched_row_title="Matched Paper",
                score=0.9,
                runner_up_score=0.1,
                runner_up_row_index=None,
                reasoning="clear match",
                blocked=False,
                matched_at=datetime.now(timezone.utc).isoformat(),
            )
        ]

        monkeypatch.setattr("backend.app.runner.check_readiness", _ready_ok)
        monkeypatch.setattr("backend.app.runner.load_table", lambda path: df)
        monkeypatch.setattr("backend.app.runner.validate_metadata_columns", lambda df: [])
        monkeypatch.setattr("backend.app.runner.load_schema", lambda schema_path, table_path: schema)
        monkeypatch.setattr("backend.app.runner.validate_schema_columns", lambda schema: [])
        monkeypatch.setattr(
            "backend.app.runner.get_eligible_cells",
            lambda df, schema, verify_mode, eval_mode=False: eligible,
        )
        monkeypatch.setattr("backend.app.runner.parse_pdf", _fake_parse_pdf_with_blocks)
        monkeypatch.setattr("backend.app.runner.run_matching", lambda **kwargs: match_results)
        monkeypatch.setattr("backend.app.runner.persist_match_artifacts", lambda *args, **kwargs: None)
        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)
        monkeypatch.setattr("backend.app.runner.run_style_profiles_stage", _fake_style_profiles)
        monkeypatch.setattr("backend.app.runner.run_retrieval_for_cell", _fake_retrieval_result)
        monkeypatch.setattr("backend.app.runner.extract_cell", _fake_extract_cell_with_metrics)

        await run_pipeline(run_id, config, "config.json", output_dir)

        stats = read_json(get_run_stats_path(output_dir, run_id))
        counters = stats["counters"]
        pdf_stats = stats["per_pdf"]["paper_1"]
        cell_stats = stats["per_cell"][0]

        assert counters["pdf_count"] == 1
        assert counters["eligible_cell_count"] == 1
        assert counters["processed_cell_count"] == 1
        assert counters["matched_pdf_count"] == 1
        assert counters["retrieval_calls"] == 1
        assert counters["retrieval_calls_per_pdf"] == {"paper_1": 1}
        assert counters["chunk_build_count_per_pdf"] == {"paper_1": 3}
        assert counters["idf_build_count_per_pdf"] == {"paper_1": 2}
        assert counters["neighbor_chunks_added_count"] == 2
        assert counters["chunk_build_repeated_work_count"] == 2
        assert counters["idf_build_repeated_work_count"] == 1
        assert counters["retrieval_repeated_work_count"] == 3
        assert counters["text_model_call_count"] == 1
        assert counters["vision_model_call_count"] == 1
        assert counters["evidence_item_count"] == 4
        assert counters["direct_quote_count"] == 1
        assert counters["approximate_highlight_count"] == 1
        assert counters["quote_plus_page_count"] == 1
        assert counters["figure_derived_evidence_count"] == 1
        assert counters["needs_more_evidence_count"] == 1
        assert counters["recall_rescue_used_count"] == 1
        assert counters["figure_review_triggered_count"] == 1
        assert counters["cells_per_pdf"] == {"paper_1": 1}
        assert counters["chunk_count_total"] == 6
        assert counters["chunk_count_by_type"]["paragraph"] == 2
        assert counters["chunk_count_by_type"]["caption"] == 1

        policy_summary = stats["retrieval_policy_summary"]
        assert policy_summary["query_modes"] == ["lexical_with_hints"]
        assert policy_summary["scoring_profiles"] == ["bm25_lite"]
        assert policy_summary["heuristic_tags"] == ["count_like"]
        assert policy_summary["hint_terms"] == ["count"]
        assert policy_summary["allowed_chunk_types"] == [
            "abstract",
            "caption",
            "figure",
            "list_item",
            "paragraph",
            "section",
            "table_region",
        ]
        assert policy_summary["include_captions_values"] == [True]
        assert policy_summary["include_tables_values"] == [True]
        assert policy_summary["include_neighbor_window_values"] == [True]
        assert policy_summary["top_k_values"] == [6]

        assert pdf_stats["pdf_cell_count"] == 1
        assert pdf_stats["retrieval_calls"] == 1
        assert pdf_stats["neighbor_chunks_added_count"] == 2
        assert pdf_stats["text_model_call_count"] == 1
        assert pdf_stats["vision_model_call_count"] == 1
        assert pdf_stats["evidence_item_count"] == 4

        assert cell_stats["candidate_chunk_count"] == 4
        assert cell_stats["selected_chunk_count"] == 2
        assert cell_stats["neighbor_chunks_added_count"] == 2
        assert cell_stats["evidence_item_count"] == 4
        assert cell_stats["direct_quote_count"] == 1
        assert cell_stats["approximate_highlight_count"] == 1
        assert cell_stats["quote_plus_page_count"] == 1
        assert cell_stats["figure_derived_evidence_count"] == 1

        assert stats["consistency"]["processed_cells_match_per_cell_records"] is True
        assert stats["consistency"]["retrieval_calls_match_per_pdf_sum"] is True
        assert stats["consistency"]["evidence_items_match_persisted_records"] is True
        assert stats["consistency"]["per_pdf_evidence_items_match_persisted_records"] is True
        assert stats["consistency"]["text_model_calls_match_per_cell_sum"] is True
        assert stats["consistency"]["vision_model_calls_match_per_cell_sum"] is True

    @pytest.mark.asyncio
    @respx.mock
    async def test_writes_artifact_and_provider_diagnostic_summaries(self, tmp_path, monkeypatch):
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)
        monkeypatch.setattr("backend.app.runner.parse_pdf", _fake_parse_pdf)
        monkeypatch.setattr("backend.app.runner.run_matching", lambda **kwargs: [])
        monkeypatch.setattr("backend.app.runner.persist_match_artifacts", lambda *args, **kwargs: None)
        monkeypatch.setattr("backend.app.runner.run_style_profiles_stage", _fake_style_profiles)
        config = make_config(tmp_path)
        run_id = "run_artifact_summary"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        artifact_summary = read_json(get_artifact_summary_path(output_dir, run_id))
        provider_diagnostics = read_json(get_provider_diagnostics_path(output_dir, run_id))
        provider_probe = read_json(get_provider_probe_path(output_dir, run_id))
        provider_model_management = read_json(get_provider_model_management_path(output_dir, run_id))
        run_data = read_json(get_run_json_path(output_dir, run_id))

        assert artifact_summary["files"]["provider_diagnostics"]["present"] is True
        assert artifact_summary["files"]["provider_probe"]["present"] is True
        assert artifact_summary["files"]["provider_model_management"]["present"] is True
        assert artifact_summary["directories"]["exports"]["file_count"] == 0
        assert artifact_summary["directories"]["review"]["file_count"] == 1
        assert artifact_summary["directories"]["diagnostics"]["file_count"] >= 4
        assert "diagnostics/provider_diagnostics.json" in artifact_summary["sections"]["diagnostics"]
        assert provider_diagnostics["attempt_count"] == 0
        assert provider_probe["provider"] == "lm_studio"
        assert provider_model_management["text_model"]["requested_load_context"] == 32000
        assert provider_model_management["text_model"]["reused_loaded_model"] is True
        assert run_data["artifact_summary_path"] == "summaries/artifact_summary.json"
        assert run_data["provider_diagnostics_path"] == "diagnostics/provider_diagnostics.json"
        assert run_data["provider_probe_path"] == "diagnostics/provider_probe.json"
        assert run_data["provider_model_management_path"] == "diagnostics/provider_model_management.json"

    @pytest.mark.asyncio
    @respx.mock
    async def test_writes_input_summary(self, tmp_path, monkeypatch):
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        _patch_fast_no_match_pipeline(monkeypatch)
        config = make_config(tmp_path)
        run_id = "run_inputs"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        input_path = get_input_summary_path(output_dir, run_id)
        assert input_path.exists()
        summary = read_json(input_path)
        assert summary["table_rows"] is not None
        assert summary["table_rows"] > 0
        assert summary["pdf_count"] is not None
        assert summary["run_mode"] == "normal"
        assert summary["retrieval_mode"] == "hybrid_experimental"
        assert summary["prompt_hash"] is not None
        assert summary["prompt_bundle_id"] == "default"
        assert summary["prompt_manifest_hash"] is not None
        assert summary["prompt_bundle_hash"] is not None
        assert summary["prompt_bundle_path"] is not None
        assert summary["prompt_keys_used"]
        assert summary["prompt_files"]
        assert summary["config_hash"] is not None

    @pytest.mark.asyncio
    @respx.mock
    async def test_eval_mode_persists_masked_and_gold_artifacts(self, tmp_path, monkeypatch):
        _CAPTURED_ROW_CONTEXTS.clear()
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)

        workbook_path = tmp_path / "gold.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Title", "Authors", "Publication Year", "assay"])
        sheet.append(["Matched Paper", "Smith, J.", "2024", "STARR-seq"])
        workbook.save(workbook_path)
        pdf_dir = tmp_path / "pdfs"
        pdf_dir.mkdir()
        (pdf_dir / "paper_1.pdf").write_bytes(b"%PDF-1.4\n%stub")
        schema_path = tmp_path / "schema.csv"
        schema_path.write_text("column_name,description\nassay,Assay description\n", encoding="utf-8")

        config = make_config(
            tmp_path,
            table_path=str(workbook_path),
            schema_path=str(schema_path),
            pdf_dir=str(pdf_dir),
            eval_mode=True,
        )
        run_id = "run_eval_artifacts"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        monkeypatch.setattr("backend.app.runner.parse_pdf", _fake_parse_pdf)
        monkeypatch.setattr("backend.app.runner.run_matching", lambda **kwargs: [
            MatchResult(
                pdf_id="paper_1",
                pdf_path="paper_1.pdf",
                outcome=MatchOutcome.matched,
                matched_row_index=0,
                matched_row_title="Matched Paper",
                score=0.9,
                runner_up_score=0.1,
                runner_up_row_index=None,
                reasoning="clear match",
                blocked=False,
                matched_at=datetime.now(timezone.utc).isoformat(),
            )
        ])
        monkeypatch.setattr("backend.app.runner.persist_match_artifacts", lambda *args, **kwargs: None)
        monkeypatch.setattr("backend.app.runner.run_style_profiles_stage", _fake_style_profiles)
        monkeypatch.setattr("backend.app.runner.run_retrieval_for_cell", lambda **kwargs: None)
        monkeypatch.setattr("backend.app.runner.extract_cell", _fake_extract_cell_capture)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        input_summary = read_json(get_input_summary_path(output_dir, run_id))
        proposals = load_proposals(pathlib.Path(output_dir) / run_id)
        gold_workbook_path = pathlib.Path(output_dir) / run_id / "inputs" / "gold_table.xlsx"
        masked_workbook_path = pathlib.Path(output_dir) / run_id / "inputs" / "masked_working_table.xlsx"
        gold_sheet = openpyxl.load_workbook(gold_workbook_path).active
        masked_sheet = openpyxl.load_workbook(masked_workbook_path).active
        gold_header_row = [cell.value for cell in gold_sheet[1]]
        header_row = [cell.value for cell in masked_sheet[1]]
        assay_column = header_row.index("assay") + 1

        assert run_data["run_mode"] == "eval"
        assert run_data["eval_mode"] is True
        assert input_summary["run_mode"] == "eval"
        assert run_data["prompt_hash"] is not None
        assert run_data["config_hash"] is not None
        assert run_data["schema_hash"] is not None
        assert run_data["eval_artifacts"]["gold_table"]["source_reference"] == str(workbook_path)
        assert run_data["eval_artifacts"]["gold_table"]["snapshot_path"] == "inputs/gold_table.xlsx"
        assert run_data["eval_artifacts"]["masked_working_table"]["path"] == "inputs/masked_working_table.xlsx"
        assert gold_header_row[:2] == ["row_id", "row_index"]
        assert header_row[:2] == ["row_id", "row_index"]
        assert proposals[0].run_mode == "eval"
        assert proposals[0].prompt_hash == run_data["prompt_hash"]
        assert proposals[0].gold_table_hash == run_data["eval_artifacts"]["gold_table"]["content_hash"]
        assert proposals[0].masked_working_table_hash == run_data["eval_artifacts"]["masked_working_table"]["content_hash"]

        artifact_summary = read_json(get_artifact_summary_path(output_dir, run_id))
        assert artifact_summary["eval_artifact_parity"]["expected"] is True
        assert artifact_summary["eval_artifact_parity"]["gold_table_snapshot_present"] is True
        assert artifact_summary["eval_artifact_parity"]["masked_working_table_present"] is True
        assert artifact_summary["proposal_metadata_coverage"]["gold_table_snapshot_path_present"] == 1
        assert artifact_summary["proposal_metadata_coverage"]["masked_working_table_path_present"] == 1

        original_sheet = openpyxl.load_workbook(workbook_path).active
        original_header_row = [cell.value for cell in original_sheet[1]]
        original_assay_column = original_header_row.index("assay") + 1
        original_value = original_sheet.cell(row=2, column=original_assay_column).value
        masked_value = masked_sheet.cell(row=2, column=assay_column).value
        assert original_value == "STARR-seq"
        assert masked_value in ("", None)
        assert _CAPTURED_ROW_CONTEXTS[-1]["assay"] == ""

    @pytest.mark.asyncio
    @respx.mock
    async def test_fails_with_unreachable_provider(self, tmp_path):
        respx.get("http://localhost:1234/v1/models").mock(
            side_effect=httpx.ConnectError("refused")
        )
        config = make_config(tmp_path)
        run_id = "run_fail"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        assert run_data["status"] == RunStatus.failed.value
        assert run_data["error_message"] is not None
        assert run_data["current_stage"] is None

    @pytest.mark.asyncio
    async def test_input_summary_written_before_readiness_check(self, tmp_path):
        """Input summary should exist even when readiness fails."""
        config = make_config(tmp_path, pdf_dir="/nonexistent/dir")
        run_id = "run_early_summary"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        assert run_data["status"] == RunStatus.failed.value
        assert run_data["current_stage"] is None

        # Input summary should still exist
        input_path = get_input_summary_path(output_dir, run_id)
        assert input_path.exists()

    @pytest.mark.asyncio
    @respx.mock
    async def test_records_total_rows_and_eligible_cells(self, tmp_path, monkeypatch):
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        _patch_fast_no_match_pipeline(monkeypatch)
        config = make_config(tmp_path)
        run_id = "run_counts"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        assert run_data["total_rows"] > 0
        assert run_data["eligible_cells"] >= 0

    @pytest.mark.asyncio
    async def test_provider_init_failure_does_not_generate_skipped_proposals(self, tmp_path, monkeypatch):
        config = make_config(tmp_path)
        run_id = "run_no_skipped"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        monkeypatch.setattr("backend.app.runner.check_readiness", _ready_ok)
        monkeypatch.setattr("backend.app.runner.initialize_provider", _raise_provider_error)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        proposals_dir = pathlib.Path(output_dir) / run_id / "proposals"

        assert run_data["status"] == RunStatus.failed.value
        assert run_data["error_message"] == "provider offline"
        assert list(proposals_dir.glob("*.json")) == []
        assert not (proposals_dir / "proposals.jsonl").exists()

    @pytest.mark.asyncio
    @respx.mock
    async def test_json_object_fallback_surfaces_provider_degraded_warning(self, tmp_path, monkeypatch):
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        config = make_config(tmp_path)
        run_id = "run_provider_degraded"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider_json_object)
        monkeypatch.setattr("backend.app.runner.parse_pdf", _fake_parse_pdf)
        monkeypatch.setattr("backend.app.runner.run_matching", lambda **kwargs: [])
        monkeypatch.setattr("backend.app.runner.persist_match_artifacts", lambda *args, **kwargs: None)
        monkeypatch.setattr("backend.app.runner.run_style_profiles_stage", _fake_style_profiles)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        warnings = run_data.get("warnings", [])
        degraded = [w for w in warnings if w.get("category") == "provider_degraded"]
        assert len(degraded) >= 1
        assert "json_object" in degraded[0].get("message", "")
        assert degraded[0].get("context", {}).get("structured_output_reason") == "structured_backend_incompatible"
        assert run_data.get("structured_output_mode") == "json_object"
        assert run_data.get("structured_output_reason") == "structured_backend_incompatible"
        assert run_data.get("structured_output_fallback_used") is True
        assert run_data.get("provider_readiness_reason") is None

        provider_mode = read_json(get_provider_mode_path(output_dir, run_id))
        assert provider_mode.get("structured_output_mode") == "json_object"
        assert provider_mode.get("structured_output_reason") == "structured_backend_incompatible"
        assert provider_mode.get("structured_output_fallback_used") is True

        reviewer_summary = read_json(get_reviewer_summary_path(output_dir, run_id))
        assert reviewer_summary.get("structured_output_reason") == "structured_backend_incompatible"

    @pytest.mark.asyncio
    @respx.mock
    async def test_prompt_only_provider_mode_surfaces_provider_degraded_warning(self, tmp_path, monkeypatch):
        respx.get("http://localhost:1234/v1/models").mock(
            return_value=httpx.Response(200, json={"data": [{"id": "qwen/qwen3-30b-a3b-2507"}]})
        )
        config = make_config(tmp_path)
        run_id = "run_provider_prompt_only"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider_prompt_only)
        monkeypatch.setattr("backend.app.runner.parse_pdf", _fake_parse_pdf)
        monkeypatch.setattr("backend.app.runner.run_matching", lambda **kwargs: [])
        monkeypatch.setattr("backend.app.runner.persist_match_artifacts", lambda *args, **kwargs: None)
        monkeypatch.setattr("backend.app.runner.run_style_profiles_stage", _fake_style_profiles)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        warnings = run_data.get("warnings", [])
        degraded = [w for w in warnings if w.get("category") == "provider_degraded"]
        assert run_data["status"] in {
            RunStatus.completed.value,
            RunStatus.completed_with_warnings.value,
        }
        assert len(degraded) >= 1
        assert "prompt-only JSON mode" in degraded[0].get("message", "")
        assert run_data.get("provider_readiness_reason") is None
        assert run_data.get("structured_output_mode") == "none"
        assert run_data.get("structured_output_reason") == "structured_backend_incompatible"
        assert run_data.get("structured_output_fallback_used") is True

        provider_mode = read_json(get_provider_mode_path(output_dir, run_id))
        assert provider_mode.get("structured_output_mode") == "none"
        assert provider_mode.get("structured_output_reason") == "structured_backend_incompatible"
        assert provider_mode.get("readiness_reason") is None
        assert provider_mode.get("structured_output_fallback_used") is True

        reviewer_summary = read_json(get_reviewer_summary_path(output_dir, run_id))
        assert reviewer_summary.get("provider_readiness_reason") is None

    @pytest.mark.asyncio
    async def test_provider_init_model_unavailable_is_not_classified_as_unreachable(self, tmp_path, monkeypatch):
        config = make_config(tmp_path)
        run_id = "run_model_unavailable"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        monkeypatch.setattr("backend.app.runner.check_readiness", _ready_ok)
        monkeypatch.setattr("backend.app.runner.initialize_provider", _raise_model_unavailable_error)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        assert run_data["status"] == RunStatus.failed.value
        assert run_data["provider_readiness_reason"] == "model_unavailable"
        assert run_data.get("structured_output_reason") is None
        categories = [w.get("category") for w in run_data.get("warnings", [])]
        assert "model_unavailable" in categories
        assert "provider_unreachable" not in categories

    @pytest.mark.asyncio
    async def test_only_usable_matched_rows_create_proposal_artifacts(self, tmp_path, monkeypatch):
        config = make_config(tmp_path)
        run_id = "run_matched_only"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        df = pd.DataFrame([
            {
                "Title": "Matched Paper",
                "Authors": "Smith, J.",
                "Publication Year": "2024",
                "assay": "",
            },
            {
                "Title": "Unmatched Paper",
                "Authors": "Doe, A.",
                "Publication Year": "2023",
                "assay": "",
            },
        ])
        schema = [{"column_name": "assay", "description": "Assay description"}]
        eligible = [
            {"row_index": 0, "column_name": "assay", "current_value": "", "eligibility": "eligible"},
            {"row_index": 1, "column_name": "assay", "current_value": "", "eligibility": "eligible"},
        ]
        match_results = [
            MatchResult(
                pdf_id="paper_1",
                pdf_path="paper_1.pdf",
                outcome=MatchOutcome.matched,
                matched_row_index=0,
                matched_row_title="Matched Paper",
                score=0.9,
                runner_up_score=0.2,
                runner_up_row_index=1,
                reasoning="clear match",
                blocked=False,
                matched_at=datetime.now(timezone.utc).isoformat(),
            ),
            MatchResult(
                pdf_id="paper_2",
                pdf_path="paper_2.pdf",
                outcome=MatchOutcome.unmatched,
                matched_row_index=None,
                matched_row_title=None,
                score=0.0,
                runner_up_score=0.0,
                runner_up_row_index=None,
                reasoning="unmatched",
                blocked=True,
                blocked_reason="unmatched",
                matched_at=datetime.now(timezone.utc).isoformat(),
            ),
        ]

        monkeypatch.setattr("backend.app.runner.check_readiness", _ready_ok)
        monkeypatch.setattr("backend.app.runner.load_table", lambda path: df)
        monkeypatch.setattr("backend.app.runner.validate_metadata_columns", lambda df: [])
        monkeypatch.setattr("backend.app.runner.load_schema", lambda schema_path, table_path: schema)
        monkeypatch.setattr("backend.app.runner.validate_schema_columns", lambda schema: [])
        monkeypatch.setattr(
            "backend.app.runner.get_eligible_cells",
            lambda df, schema, verify_mode, eval_mode=False: eligible,
        )
        monkeypatch.setattr("backend.app.runner.parse_pdf", _fake_parse_pdf)
        monkeypatch.setattr("backend.app.runner.run_matching", lambda **kwargs: match_results)
        monkeypatch.setattr("backend.app.runner.persist_match_artifacts", lambda *args, **kwargs: None)
        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)
        monkeypatch.setattr("backend.app.runner.run_style_profiles_stage", _fake_style_profiles)
        monkeypatch.setattr("backend.app.runner.run_retrieval_for_cell", lambda **kwargs: None)
        monkeypatch.setattr("backend.app.runner.extract_cell", _fake_extract_cell)

        await run_pipeline(run_id, config, "config.json", output_dir)

        proposals_dir = pathlib.Path(output_dir) / run_id / "proposals"
        run_data = read_json(get_run_json_path(output_dir, run_id))

        assert run_data["proposals_generated"] == 2
        assert (proposals_dir / "proposals.jsonl").exists()
        assert (proposals_dir / "proposal_index.json").exists()

    @pytest.mark.asyncio
    async def test_parse_and_evidence_fallback_truth_surfaces_in_run_warnings(self, tmp_path, monkeypatch):
        config = make_config(tmp_path, parser={"cache_enabled": False})
        run_id = "run_warning_truth"
        output_dir = str(tmp_path / "runs")
        (tmp_path / "runs").mkdir(exist_ok=True)

        df = pd.DataFrame([
            {
                "Title": "Matched Paper",
                "Authors": "Smith, J.",
                "Publication Year": "2024",
                "assay": "",
            }
        ])
        schema = [{"column_name": "assay", "description": "Assay description"}]
        eligible = [{"row_index": 0, "column_name": "assay", "current_value": "", "eligibility": "eligible"}]
        match_results = [
            MatchResult(
                pdf_id="paper_1",
                pdf_path="paper_1.pdf",
                outcome=MatchOutcome.matched,
                matched_row_index=0,
                matched_row_title="Matched Paper",
                score=0.9,
                runner_up_score=0.2,
                runner_up_row_index=None,
                reasoning="clear match",
                blocked=False,
                matched_at=datetime.now(timezone.utc).isoformat(),
            ),
        ]

        monkeypatch.setattr("backend.app.runner.check_readiness", _ready_ok)
        monkeypatch.setattr("backend.app.runner.load_table", lambda path: df)
        monkeypatch.setattr("backend.app.runner.validate_metadata_columns", lambda df: [])
        monkeypatch.setattr("backend.app.runner.load_schema", lambda schema_path, table_path: schema)
        monkeypatch.setattr("backend.app.runner.validate_schema_columns", lambda schema: [])
        monkeypatch.setattr(
            "backend.app.runner.get_eligible_cells",
            lambda df, schema, verify_mode, eval_mode=False: eligible,
        )
        monkeypatch.setattr("backend.app.runner.parse_pdf", _fake_parse_pdf_with_warnings)
        monkeypatch.setattr("backend.app.runner.run_matching", lambda **kwargs: match_results)
        monkeypatch.setattr("backend.app.runner.persist_match_artifacts", lambda *args, **kwargs: None)
        monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)
        monkeypatch.setattr("backend.app.runner.run_style_profiles_stage", _fake_style_profiles)
        monkeypatch.setattr("backend.app.runner.run_retrieval_for_cell", lambda **kwargs: None)
        monkeypatch.setattr("backend.app.runner.extract_cell", _fake_extract_cell_with_fallback)

        await run_pipeline(run_id, config, "config.json", output_dir)

        run_data = read_json(get_run_json_path(output_dir, run_id))
        warning_messages = [warning["message"] for warning in run_data["warnings"]]

        assert any("Parser fallback used" in message for message in warning_messages)
        assert any("OCR fallback used" in message for message in warning_messages)
        assert any("require evidence fallback review" in message for message in warning_messages)


async def _ready_ok(*args, **kwargs):
    return SimpleNamespace(ok=True, errors=[])


async def _raise_provider_error(*args, **kwargs):
    from backend.app.provider import ProviderError

    raise ProviderError("provider offline")


async def _raise_model_unavailable_error(*args, **kwargs):
    from backend.app.provider import ProviderError

    raise ProviderError("configured model not loaded", reason="model_unavailable")


async def _raise_no_compatible_structured_mode_error(*args, **kwargs):
    from backend.app.provider import ProviderError

    raise ProviderError(
        "Provider 'lm_studio' cannot run extraction for text model 'qwen/qwen3-30b-a3b-2507' without a compatible structured-output mode.",
        reason="no_compatible_structured_mode",
        details={
            "capabilities": {
                "structured_output_mode": "none",
                "structured_output_reason": "structured_backend_incompatible",
                "structured_output_error": "LM Studio rejected structured-output grammar/regex constraints: Failed to process regex",
            },
            "model_management": _fake_model_management_report(load_requested=True, reused=False),
        },
    )


async def _fake_initialize_provider(*args, **kwargs):
    return _fake_provider_object(), SimpleNamespace(
        mode="live_local",
        locality="local",
        readiness_error=None,
        readiness_reason=None,
        structured_output_mode="json_schema",
        structured_output_reason=None,
        structured_output_fallback_used=False,
        vision_structured_output_mode=None,
        vision_structured_output_reason=None,
        capabilities=None,
        model_management=_fake_model_management_report(),
        model_dump=lambda: {
            "mode": "live_local",
            "locality": "local",
            "readiness_error": None,
            "readiness_reason": None,
            "structured_output_mode": "json_schema",
            "structured_output_reason": None,
            "structured_output_fallback_used": False,
            "vision_structured_output_mode": None,
            "vision_structured_output_reason": None,
            "model_management": _fake_model_management_report(),
        },
    )


async def _fake_initialize_provider_json_object(*args, **kwargs):
    return _fake_provider_object(), SimpleNamespace(
        mode="live_local",
        locality="local",
        readiness_error=None,
        readiness_reason=None,
        structured_output_mode="json_object",
        structured_output_reason="structured_backend_incompatible",
        structured_output_fallback_used=True,
        vision_structured_output_mode=None,
        vision_structured_output_reason=None,
        model_management=_fake_model_management_report(load_requested=True, reused=False),
        capabilities=SimpleNamespace(
            structured_output_mode="json_object",
            structured_output_reason="structured_backend_incompatible",
            structured_output_error="LM Studio rejected structured-output grammar/regex constraints: Failed to process regex",
        ),
        model_dump=lambda: {
            "mode": "live_local",
            "locality": "local",
            "readiness_error": None,
            "readiness_reason": None,
            "structured_output_mode": "json_object",
            "structured_output_reason": "structured_backend_incompatible",
            "structured_output_fallback_used": True,
            "vision_structured_output_mode": None,
            "vision_structured_output_reason": None,
            "model_management": _fake_model_management_report(load_requested=True, reused=False),
            "capabilities": {
                "structured_output_mode": "json_object",
                "structured_output_reason": "structured_backend_incompatible",
                "structured_output_error": "LM Studio rejected structured-output grammar/regex constraints: Failed to process regex",
            },
        },
    )


async def _fake_initialize_provider_prompt_only(*args, **kwargs):
    return _fake_provider_object(), SimpleNamespace(
        mode="live_local",
        locality="local",
        readiness_error=None,
        readiness_reason=None,
        structured_output_mode="none",
        structured_output_reason="structured_backend_incompatible",
        structured_output_fallback_used=True,
        vision_structured_output_mode=None,
        vision_structured_output_reason=None,
        model_management=_fake_model_management_report(load_requested=True, reused=False),
        capabilities=SimpleNamespace(
            structured_output_mode="none",
            structured_output_reason="structured_backend_incompatible",
            structured_output_error="LM Studio rejected structured-output grammar/regex constraints: Failed to process regex",
        ),
        model_dump=lambda: {
            "mode": "live_local",
            "locality": "local",
            "readiness_error": None,
            "readiness_reason": None,
            "structured_output_mode": "none",
            "structured_output_reason": "structured_backend_incompatible",
            "structured_output_fallback_used": True,
            "vision_structured_output_mode": None,
            "vision_structured_output_reason": None,
            "model_management": _fake_model_management_report(load_requested=True, reused=False),
            "capabilities": {
                "structured_output_mode": "none",
                "structured_output_reason": "structured_backend_incompatible",
                "structured_output_error": "LM Studio rejected structured-output grammar/regex constraints: Failed to process regex",
            },
        },
    )


def _fake_provider_object():
    return SimpleNamespace(
        get_model_management_report=lambda: _fake_model_management_report(),
    )


def _fake_model_management_report(*, load_requested: bool = False, reused: bool = True):
    return {
        "provider": "lm_studio",
        "base_url": "http://localhost:1234",
        "text_model": {
            "model_id": "qwen/qwen3-30b-a3b-2507",
            "working_context_budget": 25000,
            "configured_load_context": 32000,
            "requested_load_context": 32000,
            "load_context_is_derived": False,
            "load_requested": load_requested,
            "reused_loaded_model": reused,
            "loaded_instance_id": "qwen/qwen3-30b-a3b-2507",
            "loaded_instance_context_length": 32000,
            "actual_load_config": {"context_length": 32000},
            "load_time_seconds": 4.2 if load_requested else None,
            "status": "loaded_via_api" if load_requested else "reused_loaded_instance",
            "failure": None,
        },
        "vision_model": None,
        "recorded_at": datetime.now(timezone.utc).isoformat(),
    }


async def _fake_style_profiles(**kwargs):
    return {}


def _patch_fast_no_match_pipeline(monkeypatch) -> None:
    monkeypatch.setattr("backend.app.runner.initialize_provider", _fake_initialize_provider)
    monkeypatch.setattr("backend.app.runner.parse_pdf", _fake_parse_pdf)
    monkeypatch.setattr("backend.app.runner.run_matching", lambda **kwargs: [])
    monkeypatch.setattr("backend.app.runner.persist_match_artifacts", lambda *args, **kwargs: None)
    monkeypatch.setattr("backend.app.runner.run_style_profiles_stage", _fake_style_profiles)


def _fake_parse_pdf(**kwargs):
    pdf_id = kwargs["pdf_id"]
    doc = SimpleNamespace(model_dump=lambda: {"pdf_id": pdf_id, "blocks": []})
    diagnostics = SimpleNamespace(
        fallback_used=False,
        actual_parser_used="docling",
        configured_parser="docling",
        ocr_used=False,
        ocr_reason=None,
        parse_warnings=[],
        major_extraction_gaps=[],
    )
    return doc, diagnostics, []


def _fake_parse_pdf_with_blocks(**kwargs):
    pdf_id = kwargs["pdf_id"]
    doc = SimpleNamespace(
        model_dump=lambda: {
            "pdf_id": pdf_id,
            "blocks": [
                {"block_id": "b1", "block_type": "paragraph", "text": "intro", "page_number": 1, "reading_order": 1},
                {"block_id": "b2", "block_type": "paragraph", "text": "methods", "page_number": 1, "reading_order": 2},
                {"block_id": "b3", "block_type": "caption", "text": "figure caption", "page_number": 1, "reading_order": 3},
                {"block_id": "b4", "block_type": "section_heading", "text": "Results", "page_number": 1, "reading_order": 4},
                {"block_id": "b5", "block_type": "list_item", "text": "item 1", "page_number": 1, "reading_order": 5},
                {"block_id": "b6", "block_type": "section_heading", "text": "Discussion", "page_number": 1, "reading_order": 6},
            ],
        }
    )
    diagnostics = SimpleNamespace(
        fallback_used=False,
        actual_parser_used="docling",
        configured_parser="docling",
        ocr_used=False,
        ocr_reason=None,
        parse_warnings=[],
        major_extraction_gaps=[],
    )
    return doc, diagnostics, []


def _fake_parse_pdf_with_warnings(**kwargs):
    pdf_id = kwargs["pdf_id"]
    doc = SimpleNamespace(model_dump=lambda: {"pdf_id": pdf_id, "blocks": []})
    diagnostics = SimpleNamespace(
        fallback_used=True,
        actual_parser_used="pypdfium2",
        configured_parser="docling",
        ocr_used=True,
        ocr_reason="low text extraction",
        parse_warnings=["Low text extraction (scanned PDF)."],
        major_extraction_gaps=["Sparse extracted text"],
    )
    return doc, diagnostics, []


async def _fake_extract_cell(**kwargs):
    proposal_id = generate_proposal_id(kwargs["run_id"], kwargs["cell_id"])
    artifact_context = kwargs.get("artifact_context") or {}
    proposal = ProposalRecord(
        proposal_id=proposal_id,
        run_id=kwargs["run_id"],
        pdf_id=kwargs["pdf_id"],
        row_id=kwargs["row_id"],
        column_name=kwargs["column_name"],
        cell_id=kwargs["cell_id"],
        state=ProposalState.unclear,
        support=SupportLabel.weak_evidence,
        proposed_value="candidate",
        rationale="- extracted",
        evidence_ids=[],
        warning_flags=[],
        needs_more_evidence=False,
        is_verify_mode=False,
        run_mode=artifact_context.get("run_mode", "normal"),
        prompt_version=artifact_context.get("prompt_version"),
        prompt_hash=artifact_context.get("prompt_hash"),
        schema_hash=artifact_context.get("schema_hash"),
        schema_version=artifact_context.get("schema_version"),
        config_hash=artifact_context.get("config_hash"),
        config_snapshot_path=artifact_context.get("config_snapshot_path"),
        parser_identity=artifact_context.get("parser_identity"),
        parser_version=artifact_context.get("parser_version"),
        text_model_id=kwargs.get("text_model_id"),
        vision_model_id=kwargs.get("vision_model_id"),
        gold_table_source_reference=artifact_context.get("gold_table_source_reference"),
        gold_table_hash=artifact_context.get("gold_table_hash"),
        gold_table_snapshot_path=artifact_context.get("gold_table_snapshot_path"),
        masked_working_table_path=artifact_context.get("masked_working_table_path"),
        masked_working_table_hash=artifact_context.get("masked_working_table_hash"),
        created_at=datetime.now(timezone.utc).isoformat(),
    )
    persist_proposal(kwargs["run_dir"], proposal)
    return proposal


async def _fake_extract_cell_with_fallback(**kwargs):
    proposal_id = generate_proposal_id(kwargs["run_id"], kwargs["cell_id"])
    proposal = ProposalRecord(
        proposal_id=proposal_id,
        run_id=kwargs["run_id"],
        pdf_id=kwargs["pdf_id"],
        row_id=kwargs["row_id"],
        column_name=kwargs["column_name"],
        cell_id=kwargs["cell_id"],
        state=ProposalState.unclear,
        support=SupportLabel.weak_evidence,
        proposed_value="candidate",
        rationale="- extracted",
        evidence_ids=[],
        warning_flags=["fallback_evidence_used"],
        needs_more_evidence=False,
        is_verify_mode=False,
        created_at=datetime.now(timezone.utc).isoformat(),
    )
    persist_proposal(kwargs["run_dir"], proposal)
    return proposal


def _fake_retrieval_result(**kwargs):
    return SimpleNamespace(
        mode="lexical",
        request_mode="baseline",
        policy={
            "query_mode": "lexical_with_hints",
            "scoring_profile": "bm25_lite",
            "heuristic_tags": ["count_like"],
            "hint_terms": ["count"],
            "allowed_chunk_types": [
                "abstract",
                "caption",
                "figure",
                "list_item",
                "paragraph",
                "section",
                "table_region",
            ],
            "include_captions": True,
            "include_tables": True,
            "include_neighbor_window": True,
            "top_k": 6,
        },
        stats={
            "total_ms": 4.0,
            "chunk_build_ms": 1.25,
            "idf_build_ms": 0.75,
            "chunk_build_count": 3,
            "idf_build_count": 2,
            "cached_index_used": True,
            "candidate_chunk_count": 4,
            "selected_chunk_count": 2,
            "neighbor_chunk_count": 2,
            "chunk_count_total": 6,
            "chunk_count_by_type": {
                "paragraph": 2,
                "caption": 1,
                "section": 2,
                "list_item": 1,
            },
        },
    )


async def _fake_extract_cell_with_metrics(**kwargs):
    proposal_id = generate_proposal_id(kwargs["run_id"], kwargs["cell_id"])
    stats_sink = kwargs.get("stats_sink")
    if stats_sink is not None:
        stats_sink.update(
            {
                "cell_total_ms": 15.0,
                "text_model_ms": 7.0,
                "text_model_calls": 1,
                "evidence_anchoring_ms": 2.0,
                "evidence_anchor_attempts": 1,
                "figure_review_ms": 3.0,
                "figure_review_calls": 1,
                "evidence_recovery_ms": 1.0,
                "recall_rescue_retrieval_ms": 0.5,
                "recall_rescue_retrieval_prep_ms": 0.25,
                "recall_rescue_used": True,
                "whole_document_used": False,
                "needs_more_evidence": True,
                "figure_hits_count": 1,
                "provider_diagnostics": {
                    "attempt_count": 2,
                    "request_kinds": {"text_structured": 1, "vision_structured": 1},
                    "outcomes": {"success": 2},
                },
                "retrieval_diagnostics": {"classification": "reasoning_gap"},
                "figure_review_diagnostics": {"triggered": True, "useful": True, "rescued_value": False},
                "figure_review_triggered": True,
                "figure_review_useful": True,
                "figure_review_rescued": False,
            }
        )

    proposal = ProposalRecord(
        proposal_id=proposal_id,
        run_id=kwargs["run_id"],
        pdf_id=kwargs["pdf_id"],
        row_id=kwargs["row_id"],
        column_name=kwargs["column_name"],
        cell_id=kwargs["cell_id"],
        state=ProposalState.unclear,
        support=SupportLabel.weak_evidence,
        proposed_value="candidate",
        rationale="- extracted",
        primary_evidence_id="ev_direct",
        evidence_ids=["ev_direct", "ev_approx", "ev_fallback", "ev_figure"],
        warning_flags=["fallback_evidence_used"],
        needs_more_evidence=True,
        recall_rescue_used=True,
        whole_document_used=False,
        figure_review_diagnostics={"triggered": True, "useful": True, "rescued_value": False},
        created_at=datetime.now(timezone.utc).isoformat(),
    )
    persist_proposal(kwargs["run_dir"], proposal)

    evidence_specs = [
        ("ev_direct", EvidenceSourceType.direct_quote, False),
        ("ev_approx", EvidenceSourceType.approximate_highlight, False),
        ("ev_fallback", EvidenceSourceType.quote_plus_page, False),
        ("ev_figure", EvidenceSourceType.caption_grounded_figure_evidence, True),
    ]
    for evidence_id, source_type, is_figure_derived in evidence_specs:
        persist_evidence(
            kwargs["run_dir"],
            EvidenceRecord(
                evidence_id=evidence_id,
                run_id=kwargs["run_id"],
                proposal_id=proposal_id,
                pdf_id=kwargs["pdf_id"],
                source_type=source_type,
                quote_text="evidence",
                page_number=1,
                anchor_confidence=0.9,
                is_primary=evidence_id == "ev_direct",
                is_figure_derived=is_figure_derived,
                created_at=datetime.now(timezone.utc).isoformat(),
            ),
        )

    return proposal


async def _fake_extract_cell_capture(**kwargs):
    _CAPTURED_ROW_CONTEXTS.append(dict(kwargs["row_context"]))
    return await _fake_extract_cell(**kwargs)

