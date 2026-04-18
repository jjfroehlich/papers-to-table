"""T106: Performance smoke tests for parsing, retrieval, extraction, and
review loading.

These tests verify that representative small and medium batch operations
complete within acceptable time bounds so obvious regressions in throughput
are caught early.

Performance tests are marked 'performance' and run in normal CI since they
use in-memory or tmp-path data.  They do NOT require external services.
"""
from __future__ import annotations

import pathlib
import time
from datetime import datetime, timezone

import openpyxl
import pytest

from backend.app.artifacts import init_run_bundle, write_json
from backend.app.export import (
    generate_audit_log,
    generate_diagnostics,
    generate_xlsx_export,
    run_export,
)
from backend.app.extraction import (
    EvidenceRecord,
    ProposalRecord,
    persist_evidence,
    persist_proposal,
)
from backend.app.ids import (
    generate_cell_id,
    generate_evidence_id,
    generate_proposal_id,
    generate_row_id,
    generate_run_id,
)
from backend.app.review import (
    get_export_candidates,
    get_progress,
    list_proposals,
    record_review_decision,
)
from backend.app.schemas import (
    EvidenceSourceType,
    ProposalState,
    ReviewDecision,
    ReviewResolutionReason,
    RunStatus,
    SupportLabel,
)

pytestmark = pytest.mark.performance

# ---------------------------------------------------------------------------
# Timing thresholds
# ---------------------------------------------------------------------------
# These are conservative bounds intended to catch major regressions only.
# They are not micro-benchmark targets.

# Small batch: ≤ 10 proposals
SMALL_BATCH_SIZE = 10
SMALL_BATCH_SECONDS = 2.0  # proposal creation + decisions + export in < 2s

# Medium batch: ≤ 50 proposals
MEDIUM_BATCH_SIZE = 50
MEDIUM_BATCH_SECONDS = 10.0  # proposal creation + decisions + export in < 10s

# Load time: listing proposals and computing progress for medium batch
LOAD_SECONDS = 3.0


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _setup_run(
    tmp_path: pathlib.Path,
    n_rows: int,
    columns: list[str],
) -> tuple[pathlib.Path, str, pathlib.Path]:
    """Create a run with n_rows × len(columns) proposals (all accepted) + table."""
    run_id = generate_run_id()
    run_dir = init_run_bundle(str(tmp_path), run_id)
    run_data = {
        "run_id": run_id,
        "status": RunStatus.completed.value,
        "output_dir": str(tmp_path),
        "verify_mode": False,
        "total_rows": n_rows,
        "eligible_cells": n_rows * len(columns),
        "proposals_generated": n_rows * len(columns),
        "proposals_reviewed": 0,
        "warnings": [],
    }
    write_json(run_dir / "run.json", run_data)

    # Write minimal XLSX table
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    all_cols = ["Title", "Authors", "Publication Year"] + columns
    for ci, h in enumerate(all_cols, start=1):
        ws.cell(row=1, column=ci, value=h)
    for ri in range(n_rows):
        ws.cell(row=ri + 2, column=1, value=f"Paper {ri}")
        ws.cell(row=ri + 2, column=2, value="Author")
        ws.cell(row=ri + 2, column=3, value="2020")
        for ci, _ in enumerate(columns, start=4):
            ws.cell(row=ri + 2, column=ci, value="")
    table_path = tmp_path / "table.xlsx"
    wb.save(str(table_path))
    write_json(run_dir / "config.snapshot.json", {"table_path": str(table_path)})

    now = datetime.now(timezone.utc).isoformat()
    for ri in range(n_rows):
        title = f"Paper {ri}"
        row_id = generate_row_id(ri, title)
        for col in columns:
            cell_id = generate_cell_id(row_id, col)
            proposal_id = generate_proposal_id(run_id, cell_id)
            prop = ProposalRecord(
                proposal_id=proposal_id,
                run_id=run_id,
                pdf_id=f"pdf_{ri}",
                row_id=row_id,
                column_name=col,
                cell_id=cell_id,
                state=ProposalState.found,
                support=SupportLabel.direct_evidence,
                proposed_value=f"value_{ri}_{col}",
                evidence_ids=[],
                warning_flags=[],
                created_at=now,
            )
            persist_proposal(run_dir, prop)
            record_review_decision(
                run_dir=run_dir,
                run_id=run_id,
                proposal_id=proposal_id,
                cell_id=cell_id,
                decision=ReviewDecision.accepted,
                resolution_reason=ReviewResolutionReason.accepted_as_proposed,
            )

    return run_dir, run_id, table_path


# ---------------------------------------------------------------------------
# Small batch performance
# ---------------------------------------------------------------------------

class TestSmallBatchPerformance:
    def test_small_batch_export_within_time(self, tmp_path: pathlib.Path):
        """Small batch (10 proposals × 1 column): full export pipeline completes quickly."""
        run_dir, run_id, table_path = _setup_run(tmp_path, n_rows=SMALL_BATCH_SIZE, columns=["Cloning"])

        t0 = time.monotonic()
        result = run_export(run_dir, str(tmp_path), run_id)
        elapsed = time.monotonic() - t0

        assert result["accepted_changes_count"] == SMALL_BATCH_SIZE
        assert pathlib.Path(result["workbook_path"]).exists()
        assert elapsed < SMALL_BATCH_SECONDS, (
            f"Small batch export took {elapsed:.2f}s, expected < {SMALL_BATCH_SECONDS}s"
        )

    def test_small_batch_proposal_load_within_time(self, tmp_path: pathlib.Path):
        """Loading and listing proposals for a small batch is fast."""
        run_dir, run_id, _ = _setup_run(tmp_path, n_rows=SMALL_BATCH_SIZE, columns=["Cloning"])

        t0 = time.monotonic()
        proposals = list_proposals(run_dir)
        progress = get_progress(run_dir)
        elapsed = time.monotonic() - t0

        assert len(proposals) == SMALL_BATCH_SIZE
        assert progress["total"] == SMALL_BATCH_SIZE
        assert elapsed < LOAD_SECONDS, (
            f"Small batch proposal load took {elapsed:.2f}s, expected < {LOAD_SECONDS}s"
        )

    def test_small_batch_xlsx_has_correct_cell_count(self, tmp_path: pathlib.Path):
        """All accepted changes appear in the exported XLSX."""
        run_dir, run_id, table_path = _setup_run(tmp_path, n_rows=SMALL_BATCH_SIZE, columns=["Cloning"])
        candidates = get_export_candidates(run_dir)
        assert len(candidates) == SMALL_BATCH_SIZE
        out = generate_xlsx_export(run_dir, candidates, str(table_path))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws is not None
        # Check each data row has the value set
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        col_idx = headers.index("Cloning") + 1
        changed = 0
        for ri in range(2, SMALL_BATCH_SIZE + 2):
            if ws.cell(row=ri, column=col_idx).value:
                changed += 1
        assert changed == SMALL_BATCH_SIZE


# ---------------------------------------------------------------------------
# Medium batch performance
# ---------------------------------------------------------------------------

class TestMediumBatchPerformance:
    def test_medium_batch_export_within_time(self, tmp_path: pathlib.Path):
        """Medium batch (50 proposals × 1 column): export completes in reasonable time."""
        run_dir, run_id, table_path = _setup_run(tmp_path, n_rows=MEDIUM_BATCH_SIZE, columns=["Cloning"])

        t0 = time.monotonic()
        result = run_export(run_dir, str(tmp_path), run_id)
        elapsed = time.monotonic() - t0

        assert result["accepted_changes_count"] == MEDIUM_BATCH_SIZE
        assert elapsed < MEDIUM_BATCH_SECONDS, (
            f"Medium batch export took {elapsed:.2f}s, expected < {MEDIUM_BATCH_SECONDS}s"
        )

    def test_medium_batch_proposal_load_within_time(self, tmp_path: pathlib.Path):
        """Listing proposals for a medium batch is fast."""
        run_dir, run_id, _ = _setup_run(tmp_path, n_rows=MEDIUM_BATCH_SIZE, columns=["Cloning"])

        t0 = time.monotonic()
        proposals = list_proposals(run_dir)
        elapsed = time.monotonic() - t0

        assert len(proposals) == MEDIUM_BATCH_SIZE
        assert elapsed < LOAD_SECONDS, (
            f"Medium batch proposal load took {elapsed:.2f}s, expected < {LOAD_SECONDS}s"
        )

    def test_medium_batch_multi_column_export(self, tmp_path: pathlib.Path):
        """Medium batch with multiple columns: export handles all accepted changes."""
        cols = ["Cloning", "Readout", "Species"]
        run_dir, run_id, table_path = _setup_run(
            tmp_path, n_rows=20, columns=cols
        )
        t0 = time.monotonic()
        result = run_export(run_dir, str(tmp_path), run_id)
        elapsed = time.monotonic() - t0

        assert result["accepted_changes_count"] == 20 * len(cols)
        assert elapsed < MEDIUM_BATCH_SECONDS, (
            f"Multi-column export took {elapsed:.2f}s, expected < {MEDIUM_BATCH_SECONDS}s"
        )

    def test_audit_log_generation_medium_batch(self, tmp_path: pathlib.Path):
        """Audit log generation for a medium batch finishes quickly."""
        run_dir, run_id, _ = _setup_run(tmp_path, n_rows=MEDIUM_BATCH_SIZE, columns=["Cloning"])
        candidates = get_export_candidates(run_dir)

        t0 = time.monotonic()
        audit_path = generate_audit_log(run_dir, candidates)
        elapsed = time.monotonic() - t0

        assert audit_path.exists()
        assert elapsed < LOAD_SECONDS, (
            f"Audit log generation took {elapsed:.2f}s, expected < {LOAD_SECONDS}s"
        )

    def test_diagnostics_generation_medium_batch(self, tmp_path: pathlib.Path):
        """Diagnostics generation for a medium batch finishes quickly."""
        run_dir, run_id, _ = _setup_run(tmp_path, n_rows=MEDIUM_BATCH_SIZE, columns=["Cloning"])
        from backend.app.artifacts import read_json
        run_data = read_json(run_dir / "run.json")

        t0 = time.monotonic()
        diag_path = generate_diagnostics(run_dir, run_data, [])
        elapsed = time.monotonic() - t0

        assert diag_path.exists()
        assert elapsed < LOAD_SECONDS, (
            f"Diagnostics generation took {elapsed:.2f}s, expected < {LOAD_SECONDS}s"
        )
