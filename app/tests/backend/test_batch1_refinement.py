from __future__ import annotations

import json
import pathlib
import struct
import zlib
from datetime import datetime, timezone
from types import SimpleNamespace
from unittest.mock import AsyncMock

import openpyxl
import pandas as pd
import pytest

from backend.app.config import RunConfig
from backend.app.extraction import (
    EvidenceRecord,
    ProposalRecord,
    build_text_extraction_prompt,
    extract_cell,
    load_evidence,
    load_proposals,
    persist_evidence,
    persist_proposal,
    rank_evidence,
    run_figure_review,
)
from backend.app.ids import generate_cell_id, generate_proposal_id, generate_row_id
from backend.app.ingest import (
    create_masked_working_dataframe,
    get_eligible_cells,
    load_schema,
    persist_masked_working_copy,
    validate_schema_columns,
)
from backend.app.matching import (
    MatchResult,
    PaperMetadata,
    detect_duplicate_row_conflicts,
    persist_match_artifacts,
    score_against_row,
)
from backend.app.review import get_proposal_detail
from backend.app.runner import run_pipeline
from backend.app.schemas import (
    EvidenceSourceType,
    EvidenceStatus,
    MatchOutcome,
    NumericValueForm,
    ProposalStatus,
    ReviewBucket,
    SchemaFieldType,
    WarningCategory,
)


@pytest.fixture
def run_dir(tmp_path: pathlib.Path) -> pathlib.Path:
    path = tmp_path / "run"
    for subdir in ("proposals", "evidence", "retrieval", "style_profiles", "matching"):
        (path / subdir).mkdir(parents=True, exist_ok=True)
    return path


@pytest.fixture
def minimal_doc_dict() -> dict:
    return {
        "pdf_id": "paper_test",
        "pdf_path": "paper_test.pdf",
        "full_text": (
            "Methods\n"
            "The scaffold was implanted in the tibial defect of Sprague-Dawley rats.\n"
            "Results\n"
            "Bone volume fraction (BVF) was measured as 45.3% at 12 weeks.\n"
        ),
        "blocks": [
            {
                "block_id": "b1",
                "block_type": "section_heading",
                "page_number": 1,
                "text": "Methods",
                "normalized_text": "methods",
                "reading_order": 1,
                "bbox": [10, 720, 200, 740],
                "provenance": "pypdfium2",
            },
            {
                "block_id": "b2",
                "block_type": "paragraph",
                "page_number": 1,
                "text": "The scaffold was implanted in the tibial defect of Sprague-Dawley rats.",
                "normalized_text": "the scaffold was implanted in the tibial defect of sprague dawley rats",
                "reading_order": 2,
                "bbox": [10, 620, 400, 680],
                "provenance": "pypdfium2",
            },
            {
                "block_id": "b3",
                "block_type": "paragraph",
                "page_number": 2,
                "text": "Bone volume fraction (BVF) was measured as 45.3% at 12 weeks.",
                "normalized_text": "bone volume fraction bvf was measured as 45 3 at 12 weeks",
                "reading_order": 3,
                "bbox": [10, 420, 420, 470],
                "provenance": "pypdfium2",
            },
            {
                "block_id": "b4",
                "block_type": "caption",
                "page_number": 2,
                "text": "Figure 1. BVF chart with 45.3% at 12 weeks.",
                "normalized_text": "figure 1 bvf chart with 45 3 at 12 weeks",
                "reading_order": 4,
                "bbox": [10, 350, 420, 390],
                "provenance": "pypdfium2",
            },
        ],
        "figures": [
            {
                "figure_id": "fig_1",
                "page_number": 2,
                "caption_text": "Figure 1. BVF chart with 45.3% at 12 weeks.",
                "bbox": [10, 200, 420, 340],
                "crop_path": None,
                "full_page_path": None,
            }
        ],
    }


def _minimal_png_bytes() -> bytes:
    def chunk(tag: bytes, data: bytes) -> bytes:
        payload = struct.pack(">I", len(data)) + tag + data
        return payload + struct.pack(">I", zlib.crc32(tag + data) & 0xFFFFFFFF)

    width = height = 150
    rows = []
    for y in range(height):
        row = bytearray([0])
        for x in range(width):
            value = 255 if (x // 10 + y // 10) % 2 == 0 else 80
            row.extend([value, value, value])
        rows.append(bytes(row))
    ihdr = struct.pack(">IIBBBBB", width, height, 8, 2, 0, 0, 0)
    idat = zlib.compress(b"".join(rows))
    return b"\x89PNG\r\n\x1a\n" + chunk(b"IHDR", ihdr) + chunk(b"IDAT", idat) + chunk(b"IEND", b"")


class TestSchemaAndEligibilityRefinement:
    def test_load_schema_supports_optional_field_type_and_allowed_values(self, tmp_path: pathlib.Path):
        schema_csv = tmp_path / "schema.csv"
        schema_csv.write_text(
            "column_name,description,field_type,allowed_values\n"
            "Species,Animal species,categorical,\"Rat|Mouse\"\n"
            "Dose,Dose administered,number,\n",
            encoding="utf-8",
        )

        schema = load_schema(str(schema_csv), str(tmp_path / "table.xlsx"))

        assert schema[0]["field_type"] == "categorical"
        assert schema[0]["allowed_values"] == ["Rat", "Mouse"]
        assert schema[1]["field_type"] == "number"
        assert schema[1]["allowed_values"] is None

    def test_validate_schema_rejects_allowed_values_without_categorical(self):
        errors = validate_schema_columns([
            {
                "column_name": "Dose",
                "description": "Dose administered",
                "field_type": "number",
                "allowed_values": ["1", "2"],
            }
        ])
        assert any("allowed_values require field_type='categorical'" in error for error in errors)

    def test_already_filled_cells_stay_out_of_scope_outside_verify_mode(self):
        df = pd.DataFrame(
            [
                {
                    "Title": "Paper A",
                    "Authors": "Smith, J.",
                    "Publication Year": "2024",
                    "Species": "Rat",
                    "Dose": "",
                }
            ]
        )
        schema = [
            {"column_name": "Title", "description": "title"},
            {"column_name": "Authors", "description": "authors"},
            {"column_name": "Publication Year", "description": "year"},
            {"column_name": "Species", "description": "species", "field_type": "categorical", "allowed_values": ["Rat", "Mouse"]},
            {"column_name": "Dose", "description": "dose", "field_type": "number"},
        ]
        cells = get_eligible_cells(df, schema, verify_mode=False)
        assert {cell["column_name"] for cell in cells} == {"Dose"}

    def test_eval_mode_marks_filled_cells_eligible(self):
        df = pd.DataFrame(
            [
                {
                    "Title": "Paper A",
                    "Authors": "Smith, J.",
                    "Publication Year": "2024",
                    "Species": "Rat",
                    "Dose": "5",
                }
            ]
        )
        schema = [
            {"column_name": "Title", "description": "title"},
            {"column_name": "Authors", "description": "authors"},
            {"column_name": "Publication Year", "description": "year"},
            {"column_name": "Species", "description": "species"},
            {"column_name": "Dose", "description": "dose"},
        ]
        cells = get_eligible_cells(df, schema, verify_mode=False, eval_mode=True)
        assert {cell["column_name"] for cell in cells} == {"Species", "Dose"}

    def test_eval_mode_masked_working_copy_blanks_targets_without_mutating_original(self, tmp_path: pathlib.Path):
        workbook_path = tmp_path / "gold.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        sheet.append(["Title", "Authors", "Publication Year", "Species", "Dose"])
        sheet.append(["Paper A", "Smith, J.", "2024", "Rat", "5"])
        sheet.append(["Paper B", "Jones, A.", "2023", "Mouse", "10"])
        workbook.create_sheet("Schema")
        workbook.save(workbook_path)

        df = pd.DataFrame(
            [
                {"Title": "Paper A", "Authors": "Smith, J.", "Publication Year": "2024", "Species": "Rat", "Dose": "5"},
                {"Title": "Paper B", "Authors": "Jones, A.", "Publication Year": "2023", "Species": "Mouse", "Dose": "10"},
            ]
        )
        schema = [
            {"column_name": "Title", "description": "title"},
            {"column_name": "Authors", "description": "authors"},
            {"column_name": "Publication Year", "description": "year"},
            {"column_name": "Species", "description": "species"},
            {"column_name": "Dose", "description": "dose"},
        ]

        masked_df, summary = create_masked_working_dataframe(df, schema)
        masked_path = tmp_path / "masked.xlsx"
        persist_masked_working_copy(str(workbook_path), str(masked_path), schema, masked_df)

        reloaded_original = openpyxl.load_workbook(workbook_path).active
        reloaded_masked = openpyxl.load_workbook(masked_path).active

        assert reloaded_original["D2"].value == "Rat"
        assert reloaded_original["E2"].value == "5"
        assert reloaded_masked["D2"].value in ("", None)
        assert reloaded_masked["E2"].value in ("", None)
        assert reloaded_masked["A2"].value == "Paper A"
        assert summary["target_cell_count"] == 4
        assert summary["masked_non_empty_cell_count"] == 4


class TestMatchingRefinement:
    def test_doi_and_author_signals_dominate_title_similarity(self):
        paper = PaperMetadata(
            title="A completely different title",
            authors=["Jones, A.", "Smith, B."],
            year=2024,
            doi="10.1000/xyz123",
        )
        wrong_title_right_metadata = {
            "Title": "Totally unrelated wording",
            "Authors": "Jones, A.; Smith, B.",
            "Publication Year": "2024",
            "DOI": "https://doi.org/10.1000/xyz123",
        }
        title_only_row = {
            "Title": "A completely different title",
            "Authors": "Other, C.",
            "Publication Year": "2018",
            "DOI": "",
        }

        assert score_against_row(paper, wrong_title_right_metadata) > score_against_row(paper, title_only_row)

    def test_duplicate_row_conflicts_capture_conflicting_pdf_ids_and_artifacts(self, run_dir: pathlib.Path):
        now = datetime.now(timezone.utc).isoformat()
        results = detect_duplicate_row_conflicts(
            [
                MatchResult(
                    pdf_id="paper_a",
                    pdf_path="paper_a.pdf",
                    outcome=MatchOutcome.matched,
                    matched_row_index=3,
                    matched_row_title="Shared Row",
                    score=0.9,
                    runner_up_score=0.2,
                    reasoning="clear match",
                    blocked=False,
                    matched_at=now,
                ),
                MatchResult(
                    pdf_id="paper_b",
                    pdf_path="paper_b.pdf",
                    outcome=MatchOutcome.matched,
                    matched_row_index=3,
                    matched_row_title="Shared Row",
                    score=0.88,
                    runner_up_score=0.1,
                    reasoning="clear match",
                    blocked=False,
                    matched_at=now,
                ),
            ]
        )

        persist_match_artifacts(run_dir, "run_test", results)
        conflicts = json.loads((run_dir / "matching" / "conflicts.json").read_text(encoding="utf-8"))

        assert all(result.conflict_pdf_ids == ["paper_a", "paper_b"] for result in results)
        assert conflicts[0]["conflict_pdf_ids"] == ["paper_a", "paper_b"]


class TestExtractionRefinement:
    async def test_extract_cell_persists_jsonl_index_and_numeric_form(
        self,
        run_dir: pathlib.Path,
        minimal_doc_dict: dict,
    ):
        provider = AsyncMock()
        provider.chat_complete_structured = AsyncMock(
            return_value={
                "proposed_value": "45.3%",
                "state": "found",
                "rationale": "- BVF reported directly.",
                "calculation": None,
                "numeric_value_form": "exact",
                "quotes": [
                    {"text": "Bone volume fraction (BVF) was measured as 45.3% at 12 weeks.", "page": 2, "source_type": "direct_quote"},
                    {"text": "Figure 1. BVF chart with 45.3% at 12 weeks.", "page": 2, "source_type": "direct_quote"},
                ],
            }
        )
        provider.vision_complete_structured = AsyncMock()

        proposal = await extract_cell(
            run_id="run_test",
            pdf_id="paper_test",
            row_id="row_test",
            cell_id="cell_test",
            column_name="Bone volume fraction",
            column_description="BVF measurement",
            row_context={"Title": "Paper", "Authors": "Jones, A.", "Publication Year": "2024"},
            doc_dict=minimal_doc_dict,
            run_dir=run_dir,
            provider=provider,
            text_model_id="text-model",
            field_type=SchemaFieldType.number,
        )

        proposals_jsonl = run_dir / "proposals" / "proposals.jsonl"
        proposal_index = run_dir / "proposals" / "proposal_index.json"
        proposal_lines = proposals_jsonl.read_text(encoding="utf-8").strip().splitlines()
        index = json.loads(proposal_index.read_text(encoding="utf-8"))

        assert proposal.evidence_status == EvidenceStatus.direct_strong
        assert proposal.numeric_value_form == NumericValueForm.exact
        assert len(proposal.evidence_ids) == 2
        assert len(proposal_lines) == 1
        assert index[proposal.proposal_id]["proposal_id"] == proposal.proposal_id

    async def test_direct_evidence_requires_anchored_quote(self, run_dir: pathlib.Path, minimal_doc_dict: dict):
        provider = AsyncMock()
        provider.chat_complete_structured = AsyncMock(
            return_value={
                "proposed_value": "45.3%",
                "state": "found",
                "rationale": "- Quote is loosely related.",
                "calculation": None,
                "numeric_value_form": "exact",
                "quotes": [{"text": "This quote does not exist in the paper", "page": 2, "source_type": "direct_quote"}],
            }
        )
        provider.vision_complete_structured = AsyncMock()

        proposal = await extract_cell(
            run_id="run_test",
            pdf_id="paper_test",
            row_id="row_test",
            cell_id="cell_loose_quote",
            column_name="Bone volume fraction",
            column_description="BVF measurement",
            row_context={},
            doc_dict=minimal_doc_dict,
            run_dir=run_dir,
            provider=provider,
            text_model_id="text-model",
            field_type=SchemaFieldType.number,
        )

        assert proposal.evidence_status == EvidenceStatus.inferred_strong
        assert "fallback_evidence_used" in proposal.warning_flags

    async def test_eval_artifact_metadata_propagates_to_proposal_and_evidence(
        self,
        run_dir: pathlib.Path,
        minimal_doc_dict: dict,
    ):
        provider = AsyncMock()
        provider.chat_complete_structured = AsyncMock(
            return_value={
                "proposed_value": "45.3%",
                "state": "found",
                "rationale": "- BVF reported directly.",
                "calculation": None,
                "numeric_value_form": "exact",
                "quotes": [
                    {
                        "text": "Bone volume fraction (BVF) was measured as 45.3% at 12 weeks.",
                        "page": 2,
                        "source_type": "direct_quote",
                    }
                ],
            }
        )
        proposal = await extract_cell(
            run_id="run_test",
            pdf_id="paper_test",
            row_id="row_test",
            cell_id="cell_eval_metadata",
            column_name="Bone volume fraction",
            column_description="BVF measurement",
            row_context={"Title": "Paper"},
            doc_dict=minimal_doc_dict,
            run_dir=run_dir,
            provider=provider,
            text_model_id="text-model",
            vision_model_id="vision-model",
            artifact_context={
                "run_mode": "eval",
                "prompt_hash": "prompt-hash",
                "config_hash": "config-hash",
                "schema_hash": "schema-hash",
                "parser_identity": "docling",
                "gold_table_source_reference": "/tmp/gold.xlsx",
                "gold_table_hash": "gold-hash",
                "gold_table_snapshot_path": "inputs/gold_table.xlsx",
                "masked_working_table_path": "inputs/masked_working_table.xlsx",
                "masked_working_table_hash": "masked-hash",
            },
        )

        evidence = [item for item in load_evidence(run_dir) if item.proposal_id == proposal.proposal_id]

        assert proposal.run_mode == "eval"
        assert proposal.prompt_hash == "prompt-hash"
        assert proposal.config_hash == "config-hash"
        assert proposal.schema_hash == "schema-hash"
        assert proposal.parser_identity == "docling"
        assert proposal.gold_table_hash == "gold-hash"
        assert proposal.masked_working_table_hash == "masked-hash"
        assert evidence[0].run_mode == "eval"
        assert evidence[0].prompt_hash == "prompt-hash"
        assert evidence[0].text_model_id == "text-model"
        assert evidence[0].vision_model_id == "vision-model"

    async def test_direct_evidence_requires_quote_to_directly_support_value(self, run_dir: pathlib.Path, minimal_doc_dict: dict):
        provider = AsyncMock()
        provider.chat_complete_structured = AsyncMock(
            return_value={
                "proposed_value": "Improved bone regeneration",
                "state": "found",
                "rationale": "- Interpreted from the paper.",
                "calculation": None,
                "numeric_value_form": None,
                "quotes": [
                    {
                        "text": "Bone volume fraction (BVF) was measured as 45.3% at 12 weeks.",
                        "page": 2,
                        "source_type": "direct_quote",
                    }
                ],
            }
        )
        provider.vision_complete_structured = AsyncMock()

        proposal = await extract_cell(
            run_id="run_test",
            pdf_id="paper_test",
            row_id="row_test",
            cell_id="cell_indirect_quote",
            column_name="Outcome summary",
            column_description="Short outcome summary",
            row_context={},
            doc_dict=minimal_doc_dict,
            run_dir=run_dir,
            provider=provider,
            text_model_id="text-model",
            field_type=SchemaFieldType.text,
        )

        # The quote is real and anchorable, but it only reports a BVF measurement.
        # It does not directly state the broader outcome summary value, so support
        # must stay inferred rather than direct evidence.
        assert proposal.evidence_status == EvidenceStatus.inferred_strong

    async def test_unclear_triggers_recall_rescue_and_optional_whole_document(
        self,
        run_dir: pathlib.Path,
        minimal_doc_dict: dict,
    ):
        provider = AsyncMock()
        provider.chat_complete_structured = AsyncMock(
            side_effect=[
                {
                    "proposed_value": None,
                    "state": "unclear",
                    "rationale": None,
                    "calculation": None,
                    "numeric_value_form": None,
                    "quotes": [],
                },
                {
                    "proposed_value": "Sprague-Dawley rat",
                    "state": "found",
                    "rationale": "- Species stated in methods.",
                    "calculation": None,
                    "numeric_value_form": None,
                    "quotes": [{"text": "The scaffold was implanted in the tibial defect of Sprague-Dawley rats.", "page": 1, "source_type": "direct_quote"}],
                },
            ]
        )
        provider.vision_complete_structured = AsyncMock()

        from backend.app.retrieval import run_retrieval_for_cell

        retrieval = run_retrieval_for_cell(
            run_id="run_test",
            pdf_id="paper_test",
            column_name="Species",
            column_description="Animal species used",
            doc_dict=minimal_doc_dict,
            run_dir=run_dir,
            top_k=3,
        )

        proposal = await extract_cell(
            run_id="run_test",
            pdf_id="paper_test",
            row_id="row_test",
            cell_id="cell_species",
            column_name="Species",
            column_description="Animal species used",
            row_context={},
            doc_dict=minimal_doc_dict,
            run_dir=run_dir,
            provider=provider,
            text_model_id="text-model",
            retrieval=retrieval,
            field_type=SchemaFieldType.categorical,
            allowed_values=["Sprague-Dawley rat", "Mouse"],
            whole_document_mode=True,
            whole_document_max_chars=5000,
        )

        assert proposal.recall_rescue_used is True
        assert proposal.whole_document_used is True
        assert provider.chat_complete_structured.await_count == 2

    async def test_figure_review_splits_reviewer_visible_subtypes(self, run_dir: pathlib.Path, minimal_doc_dict: dict):
        crop = run_dir / "figure.png"
        crop.write_bytes(_minimal_png_bytes())
        doc_dict = dict(minimal_doc_dict)
        doc_dict["figures"] = [
            {**minimal_doc_dict["figures"][0], "crop_path": str(crop), "figure_id": "fig_caption"},
            {**minimal_doc_dict["figures"][0], "crop_path": str(crop), "figure_id": "fig_visual"},
        ]

        provider = AsyncMock()
        provider.vision_complete_structured = AsyncMock(
            side_effect=[
                {
                    "proposed_value": "45.3%",
                    "state": "found",
                    "rationale": "- Caption states the value.",
                    "numeric_value_form": "exact",
                    "figure_description": "Caption-supported BVF value",
                    "caption_relevant": True,
                },
                {
                    "proposed_value": "45.3%",
                    "state": "found",
                    "rationale": "- Visual inspection suggests the same value.",
                    "numeric_value_form": "exact",
                    "figure_description": "Visual estimate from bar height",
                    "caption_relevant": False,
                },
            ]
        )

        hits = await run_figure_review(
            proposal_id="prop_test",
            run_id="run_test",
            pdf_id="paper_test",
            column_name="Bone volume fraction",
            column_description="BVF measurement",
            doc_dict=doc_dict,
            run_dir=run_dir,
            provider=provider,
            vision_model_id="vision-model",
        )

        types = [hit.evidence.source_type for hit in hits]
        assert EvidenceSourceType.caption_grounded_figure_evidence in types
        assert EvidenceSourceType.visual_interpretation_figure_evidence in types

        ranked = rank_evidence(
            [
                EvidenceRecord(
                    evidence_id="ev_inferred",
                    run_id="run_test",
                    proposal_id="prop_test",
                    pdf_id="paper_test",
                    source_type=EvidenceSourceType.inferred_reasoning,
                    quote_text="Model reasoning",
                    evidence_rank=99,
                    is_primary=False,
                    created_at=datetime.now(timezone.utc).isoformat(),
                ),
                hits[0].evidence,
                hits[1].evidence,
            ]
        )
        assert ranked[0].source_type == EvidenceSourceType.caption_grounded_figure_evidence

    def test_prompt_is_schema_first_and_supports_whole_document_mode(self):
        messages = build_text_extraction_prompt(
            column_name="Species",
            column_description="Animal species used",
            row_context={"Title": "Paper"},
            retrieval=None,
            style_profile=None,
            field_type=SchemaFieldType.categorical,
            allowed_values=["Rat", "Mouse"],
            whole_document_text="Full document text.",
        )
        prompt = messages[1]["content"]
        assert "allowed_values: Rat, Mouse" in prompt
        assert "Whole-document rescue context" in prompt
        assert "not evidence; do not copy into the answer" in prompt


class TestReviewWarningTruth:
    def test_review_payload_maps_fallback_warning_flags(self, run_dir: pathlib.Path):
        proposal = ProposalRecord(
            proposal_id="prop_test",
            run_id="run_test",
            pdf_id="paper_test",
            row_id="row_test",
            column_name="Dose",
            cell_id="cell_test",
            proposal_status=ProposalStatus.value_proposed,
            evidence_status=EvidenceStatus.inferred_strong,
            review_bucket=ReviewBucket.attention,
            reason_codes=["anchor_fallback"],
            proposed_value="5 mg",
            rationale="- fallback evidence used",
            evidence_ids=["ev_test"],
            warning_flags=["fallback_evidence_used"],
            created_at=datetime.now(timezone.utc).isoformat(),
        )
        evidence = EvidenceRecord(
            evidence_id="ev_test",
            run_id="run_test",
            proposal_id="prop_test",
            pdf_id="paper_test",
            source_type=EvidenceSourceType.quote_plus_page,
            quote_text="Dose reported as 5 mg.",
            page_number=1,
            evidence_rank=1,
            is_primary=True,
            created_at=datetime.now(timezone.utc).isoformat(),
        )
        persist_proposal(run_dir, proposal)
        persist_evidence(run_dir, evidence)

        detail = get_proposal_detail(run_dir, "prop_test")

        assert WarningCategory.fallback_evidence_used.value in detail["warning_categories"]
        assert detail["is_fallback_evidence"] is True


class TestRunnerProviderTruth:
    @pytest.mark.asyncio
    async def test_provider_failure_hard_fails_before_parse(self, tmp_path: pathlib.Path, monkeypatch):
        output_dir = str(tmp_path / "runs")
        config = RunConfig.model_validate(
            {
                "table_path": "../benchmark_datasets/massively_parallel_reporter_assays/table_template.csv",
                "schema_path": "../benchmark_datasets/massively_parallel_reporter_assays/schema.csv",
                "pdf_dir": "../benchmark_datasets/massively_parallel_reporter_assays/pdfs",
                "output_dir": output_dir,
                "provider": {
                    "token": "lm_studio",
                    "base_url": "http://localhost:1234",
                    "text_model": {"model_id": "test-model"},
                },
            }
        )

        parse_called = {"value": False}

        async def _ready_ok(*_args, **_kwargs):
            return SimpleNamespace(ok=True, errors=[])

        async def _provider_fail(*_args, **_kwargs):
            from backend.app.provider import ProviderError

            raise ProviderError("provider offline")

        def _parse_should_not_run(**_kwargs):
            parse_called["value"] = True
            raise AssertionError("parse_pdf should not run after provider init failure")

        monkeypatch.setattr("backend.app.runner.check_readiness", _ready_ok)
        monkeypatch.setattr("backend.app.runner.initialize_provider", _provider_fail)
        monkeypatch.setattr("backend.app.runner.parse_pdf", _parse_should_not_run)

        await run_pipeline("run_provider_fail", config, "config.json", output_dir)

        run_data = json.loads((tmp_path / "runs" / "run_provider_fail" / "run.json").read_text(encoding="utf-8"))
        assert run_data["status"] == "failed"
        assert run_data["error_message"] == "provider offline"
        assert run_data["provider_mode"] == "unavailable"
        assert run_data["provider_readiness_error"] == "provider offline"
        assert parse_called["value"] is False
        provider_mode = json.loads((tmp_path / "runs" / "run_provider_fail" / "summaries" / "provider_mode.json").read_text(encoding="utf-8"))
        reviewer_summary = json.loads((tmp_path / "runs" / "run_provider_fail" / "summaries" / "reviewer_summary.json").read_text(encoding="utf-8"))
        assert provider_mode["mode"] == "unavailable"
        assert reviewer_summary["total_proposals"] == 0

